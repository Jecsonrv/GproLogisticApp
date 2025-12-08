import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
import json

# Load the workbook
file_path = r'C:\Users\jecso\Desktop\Gpro Logistic App\Instrucciones\Copia de OS - GPRO LOGISTIC PRUEBAS (1).xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

print("="*80)
print("EXCEL FILE ANALYSIS - GPRO LOGISTIC")
print("Customs Brokerage Agency System")
print("="*80)
print()

# Analyze each sheet
for sheet_name in wb.sheetnames:
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*80}")

    ws = wb[sheet_name]

    # Get dimensions
    print(f"\nDimensions: {ws.max_row} rows x {ws.max_column} columns")

    # Get headers (assume first row)
    headers = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(1, col).value
        headers.append(str(cell_value) if cell_value is not None else f"Column_{col}")

    print(f"\nColumns ({len(headers)}):")
    for i, header in enumerate(headers, 1):
        print(f"  {i}. {header}")

    # Sample data (rows 2-6)
    print(f"\nSample Data (first 5 rows):")
    sample_data = []
    for row in range(2, min(7, ws.max_row + 1)):
        row_data = []
        for col in range(1, min(ws.max_column + 1, 21)):  # Limit to 20 columns for readability
            cell = ws.cell(row, col)
            value = cell.value
            if value is not None:
                row_data.append(str(value)[:50])  # Truncate long values
            else:
                row_data.append("")
        if any(row_data):  # Only add if row has data
            sample_data.append(row_data)

    if sample_data:
        # Print in tabular format
        for i, row in enumerate(sample_data, 2):
            print(f"\n  Row {i}:")
            for j, (header, value) in enumerate(zip(headers[:20], row)):
                if value:
                    print(f"    {header}: {value}")
    else:
        print("  No data found")

    # Check for formulas
    formula_count = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f':
                formula_count += 1

    if formula_count > 0:
        print(f"\n  Contains {formula_count} formula cells")

    # Check for data validation
    if hasattr(ws, 'data_validations') and ws.data_validations:
        print(f"\n  Contains {len(ws.data_validations.dataValidation)} data validations")

print("\n" + "="*80)
print("END OF ANALYSIS")
print("="*80)
