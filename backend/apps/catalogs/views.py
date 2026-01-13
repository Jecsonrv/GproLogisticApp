from rest_framework import viewsets, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from .models import (
    ProviderCategory, Provider, CustomsAgent, Bank, ShipmentType, Customs, SubClient,
    Service, ClientServicePrice
)
from .serializers import (
    ProviderCategorySerializer, ProviderSerializer, CustomsAgentSerializer, BankSerializer,
    ShipmentTypeSerializer, CustomsSerializer, SubClientSerializer,
    ServiceSerializer, ClientServicePriceSerializer
)
from .permissions import IsAdminOrReadOnly
from apps.users.permissions import IsAdminUser, IsOperativo

class ProviderCategoryViewSet(viewsets.ModelViewSet):
    queryset = ProviderCategory.objects.all()
    serializer_class = ProviderCategorySerializer
    permission_classes = [IsAdminOrReadOnly]
    search_fields = ['name']
    filterset_fields = ['is_active']

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.action == 'list' and self.request.query_params.get('is_active') is None:
            queryset = queryset.filter(is_active=True)
        return queryset.order_by('name')

class ProviderViewSet(viewsets.ModelViewSet):
    queryset = Provider.objects.select_related('category').all()
    serializer_class = ProviderSerializer
    permission_classes = [IsAdminOrReadOnly]
    search_fields = ['name', 'nit', 'email']
    filterset_fields = ['is_active', 'category']

    def get_queryset(self):
        queryset = super().get_queryset()
        # Por defecto mostrar solo activos en listados
        if self.action == 'list' and self.request.query_params.get('is_active') is None:
            queryset = queryset.filter(is_active=True)
        return queryset.order_by('name')

    @action(detail=False, methods=['get'], permission_classes=[IsOperativo])
    def general_summary(self, request):
        """
        Resumen global de Cuentas por Pagar (Proveedores).
        Suma deuda de Transfers y ProviderInvoices.
        """
        from apps.transfers.models import Transfer, ProviderInvoice
        from django.db.models import Sum, Q, F
        from datetime import datetime

        year = request.query_params.get('year')
        
        # Base QuerySets (excluyendo anulados/cancelados)
        # Transfers: Gastos operativos, administrativos, fletes, etc.
        transfers_qs = Transfer.objects.exclude(status='cancelado')
        
        # ProviderInvoices: Costos directos / Facturas de proveedores de servicios
        invoices_qs = ProviderInvoice.objects.exclude(payment_status='cancelada')

        # Filtro de año
        if year:
            try:
                year_int = int(year)
                # Filtro estricto por año de transacción/emisión
                transfers_qs = transfers_qs.filter(transaction_date__year=year_int)
                invoices_qs = invoices_qs.filter(issue_date__year=year_int)
            except ValueError:
                pass

        # === 1. Métricas Financieras Globales ===
        
        # Calcular totales de Transfers
        transfer_stats = transfers_qs.aggregate(
            total_amount=Sum('amount'),
            total_paid=Sum('paid_amount'),
            total_balance=Sum('balance')
        )
        
        # Calcular totales de ProviderInvoices (aquí el balance es calculado)
        # Nota: ProviderInvoice no tiene campo 'balance' persistido como Transfer, se calcula
        invoice_stats = invoices_qs.aggregate(
            total_amount=Sum('total_amount'),
            total_paid=Sum('paid_amount')
        )
        
        inv_total = float(invoice_stats['total_amount'] or 0)
        inv_paid = float(invoice_stats['total_paid'] or 0)
        inv_balance = inv_total - inv_paid

        # Unificar totales
        total_amount = float(transfer_stats['total_amount'] or 0) + inv_total
        total_paid = float(transfer_stats['total_paid'] or 0) + inv_paid
        total_pending = float(transfer_stats['total_balance'] or 0) + inv_balance

        # === 2. Métricas de Proveedores ===
        total_providers = Provider.objects.count()
        
        # Proveedores con deuda (necesitamos identificar IDs únicos con saldo > 0)
        # IDs desde Transfers con saldo
        transfer_debtor_ids = set(Transfer.objects.filter(
            balance__gt=0.01
        ).exclude(status='cancelado').values_list('provider_id', flat=True))
        
        # IDs desde Invoices con saldo (calculado via DB es complejo sin annotate, lo hacemos simple:
        # ProviderInvoices pendientes/parciales implican deuda)
        invoice_debtor_ids = set(ProviderInvoice.objects.filter(
            payment_status__in=['pendiente', 'parcial']
        ).values_list('provider_id', flat=True))
        
        # Unir sets de deudores
        all_debtors = transfer_debtor_ids.union(invoice_debtor_ids)
        # Filtrar None (en caso de transfers sin proveedor asignado, ej. caja chica)
        debtors_count = len([pid for pid in all_debtors if pid is not None])
        
        up_to_date_count = total_providers - debtors_count

        data = {
            'financial': {
                'total_amount': total_amount,
                'total_paid': total_paid,
                'total_pending': total_pending,
            },
            'providers': {
                'total': total_providers,
                'with_debt': debtors_count,
                'up_to_date': up_to_date_count
            }
        }
        
        return Response(data)

    @action(detail=True, methods=['get'], permission_classes=[IsOperativo])
    def account_statement(self, request, pk=None):
        from apps.transfers.models import Transfer, ProviderInvoice
        from django.db.models import Sum, F
        from datetime import datetime

        provider = self.get_object()
        
        # Filtros: Si year viene vacío (?year=), es "Todo el tiempo". Si no viene, default al actual.
        year_param = request.query_params.get('year')
        year = int(year_param) if year_param and year_param.isdigit() else (datetime.now().year if year_param is None else None)
        
        # ========== TRANSFERS (Gastos terceros, propios, admin) ==========
        unpaid_transfers = Transfer.objects.filter(
            provider=provider,
            status__in=['pendiente', 'aprobado', 'provisionada', 'parcial']
        ).order_by('transaction_date')
        
        transfer_debt = unpaid_transfers.aggregate(Sum('balance'))['balance__sum'] or 0
        
        # ========== PROVIDER INVOICES (Costos Directos / Tercerizados) ==========
        unpaid_invoices = ProviderInvoice.objects.filter(
            provider=provider,
            payment_status__in=['pendiente', 'parcial']
        ).order_by('issue_date')
        
        # Calcular balance de facturas de proveedor (total - pagado)
        invoice_debt = sum(
            float(inv.total_amount) - float(inv.paid_amount)
            for inv in unpaid_invoices
        )
        
        total_debt = float(transfer_debt) + invoice_debt
        
        # Historial - TRANSFERS
        transfers_qs = Transfer.objects.filter(provider=provider)
        if year:
            transfers_qs = transfers_qs.filter(transaction_date__year=year)
        history_transfers = transfers_qs.select_related('service_order').order_by('-transaction_date')
        
        # Historial - PROVIDER INVOICES
        invoices_qs = ProviderInvoice.objects.filter(provider=provider)
        if year:
            invoices_qs = invoices_qs.filter(issue_date__year=year)
        history_invoices = invoices_qs.select_related('service_order').order_by('-issue_date')
        
        # Aging Analysis (combinado)
        aging = {
            'current': 0.0,  # 0-30 días
            '1-30': 0.0,     # 31-60 días (1-30 vencido)
            '31-60': 0.0,    # 61-90 días
            '61-90': 0.0,    # 91-120 días
            '90+': 0.0       # > 120 días
        }
        
        today = datetime.now().date()
        
        # Aging de Transfers
        for transfer in unpaid_transfers:
            age_days = (today - transfer.transaction_date).days
            amount = float(transfer.balance)
            
            if age_days <= 30:
                aging['current'] += amount
            elif age_days <= 60:
                aging['1-30'] += amount
            elif age_days <= 90:
                aging['31-60'] += amount
            elif age_days <= 120:
                aging['61-90'] += amount
            else:
                aging['90+'] += amount
        
        # Aging de ProviderInvoices (costos directos)
        for inv in unpaid_invoices:
            age_days = (today - inv.issue_date).days
            amount = float(inv.total_amount) - float(inv.paid_amount)
            
            if age_days <= 30:
                aging['current'] += amount
            elif age_days <= 60:
                aging['1-30'] += amount
            elif age_days <= 90:
                aging['31-60'] += amount
            elif age_days <= 120:
                aging['61-90'] += amount
            else:
                aging['90+'] += amount
        
        # Construir lista unificada de "transfers" (gastos a pagar)
        transfers_data = []
        
        # Agregar Transfers normales
        for t in history_transfers:
            transfers_data.append({
                'id': t.id,
                'source': 'transfer',  # Para identificar el origen
                'transaction_date': t.transaction_date,
                'service_order': t.service_order.order_number if t.service_order else 'Gastos Admin',
                'service_order_id': t.service_order.id if t.service_order else None,
                'purchase_order': t.service_order.purchase_order if t.service_order else '',
                'type': t.get_transfer_type_display(),
                'amount': float(t.amount),
                'balance': float(t.balance),
                'paid_amount': float(t.paid_amount),
                'status': t.status,
                'status_display': t.get_status_display(),
                'description': t.description,
                'invoice_number': t.invoice_number or 'S/N',
                'invoice_file': t.invoice_file.url if t.invoice_file else None
            })
        
        # Agregar ProviderInvoices (costos directos)
        for inv in history_invoices:
            balance = float(inv.total_amount) - float(inv.paid_amount)
            transfers_data.append({
                'id': inv.id,
                'source': 'provider_invoice',  # Para identificar el origen
                'transaction_date': inv.issue_date,
                'service_order': inv.service_order.order_number if inv.service_order else 'Sin OS',
                'service_order_id': inv.service_order.id if inv.service_order else None,
                'purchase_order': inv.service_order.purchase_order if inv.service_order else '',
                'type': 'Costo Directo',  # Tipo especial para costos tercerizados
                'amount': float(inv.total_amount),
                'balance': balance,
                'paid_amount': float(inv.paid_amount),
                'status': inv.payment_status,
                'status_display': dict(ProviderInvoice.PAYMENT_STATUS_CHOICES).get(inv.payment_status, inv.payment_status),
                'description': inv.notes or f'Factura {inv.invoice_number}',
                'invoice_number': inv.invoice_number or 'S/N',
                'invoice_file': inv.invoice_file.url if inv.invoice_file else None
            })
        
        # Ordenar por fecha descendente
        transfers_data.sort(key=lambda x: x['transaction_date'], reverse=True)
        
        return Response({
            'provider': {
                'id': provider.id,
                'name': provider.name,
                'nit': provider.nit
            },
            'total_debt': total_debt,
            'aging': aging,
            'transfers': transfers_data,
            'year': year
        })

    @action(detail=True, methods=['get'], permission_classes=[IsOperativo])
    def export_statement_excel(self, request, pk=None):
        """Exportar estado de cuenta del proveedor a Excel"""
        from apps.transfers.models import Transfer, ProviderInvoice
        from django.http import HttpResponse
        from datetime import datetime
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        provider = self.get_object()
        
        # Filtros: Si year viene vacío (?year=), es "Todo el tiempo". Si no viene, default al actual.
        year_param = request.query_params.get('year')
        year = int(year_param) if year_param and year_param.isdigit() else (datetime.now().year if year_param is None else None)

        # Historial - TRANSFERS
        transfers_qs = Transfer.objects.filter(provider=provider)
        if year:
            transfers_qs = transfers_qs.filter(transaction_date__year=year)
        transfers = transfers_qs.select_related('service_order').order_by('transaction_date')
        
        # Historial - PROVIDER INVOICES (Costos Directos)
        invoices_qs = ProviderInvoice.objects.filter(provider=provider)
        if year:
            invoices_qs = invoices_qs.filter(issue_date__year=year)
        provider_invoices = invoices_qs.select_related('service_order').order_by('issue_date')

        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estado de Cuenta Proveedor"

        # Estilos profesionales - Diseño GPRO
        header_fill = PatternFill(start_color="0F2E4D", end_color="0F2E4D", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        title_font = Font(size=16, bold=True, color="0F2E4D")
        subtitle_font = Font(size=12, bold=True, color="1A4C7A")
        label_font = Font(bold=True, color="404040", size=10)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        currency_format = '"$"#,##0.00'

        # === ENCABEZADO ===
        ws['A1'] = "ESTADO DE CUENTA - PROVEEDOR"
        ws['A1'].font = title_font
        ws.merge_cells('A1:J1')

        ws['A2'] = "GPRO LOGISTIC - Agencia Aduanal"
        ws['A2'].font = Font(size=11, color="666666")
        ws.merge_cells('A2:J2')

        ws['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A3'].font = Font(size=9, italic=True, color="999999")

        # === INFORMACIÓN DEL PROVEEDOR ===
        ws['A5'] = "INFORMACIÓN DEL PROVEEDOR"
        ws['A5'].font = subtitle_font
        ws.merge_cells('A5:J5')

        # Datos del proveedor
        provider_data = [
            ('Proveedor:', provider.name),
            ('NIT:', provider.nit or 'No registrado'),
            ('Registro IVA:', provider.iva_registration or 'No registrado'),
            ('Email:', provider.email or 'No registrado'),
            ('Teléfono:', provider.phone or 'No registrado'),
        ]

        row = 6
        for label, value in provider_data:
            ws.cell(row=row, column=1, value=label).font = label_font
            ws.cell(row=row, column=2, value=value)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            row += 1

        # === TABLA DE MOVIMIENTOS ===
        start_row = 12
        ws.cell(row=start_row, column=1, value=f"MOVIMIENTOS DEL AÑO {year}").font = subtitle_font
        ws.merge_cells(f'A{start_row}:J{start_row}')

        # Headers de la tabla
        headers = ['Fecha', 'Factura Prov.', 'Orden de Servicio', 'PO', 'Tipo', 'Descripción',
                   'Estado', 'Total', 'Pagado', 'Saldo', 'Comprobante']
        header_row = start_row + 1

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Datos
        data_row = header_row + 1
        totals = {'total': 0, 'paid': 0, 'balance': 0}
        
        # Combinar transfers y provider_invoices en una lista unificada
        all_items = []
        
        for transfer in transfers:
            all_items.append({
                'date': transfer.transaction_date,
                'invoice_number': transfer.invoice_number or 'S/N',
                'service_order': transfer.service_order.order_number if transfer.service_order else 'Gastos Admin',
                'purchase_order': transfer.service_order.purchase_order if transfer.service_order else '',
                'type': transfer.get_transfer_type_display(),
                'description': transfer.description or '',
                'status': transfer.get_status_display(),
                'amount': float(transfer.amount),
                'paid': float(transfer.paid_amount),
                'balance': float(transfer.balance),
                'has_file': bool(transfer.invoice_file)
            })
        
        for inv in provider_invoices:
            balance = float(inv.total_amount) - float(inv.paid_amount)
            all_items.append({
                'date': inv.issue_date,
                'invoice_number': inv.invoice_number or 'S/N',
                'service_order': inv.service_order.order_number if inv.service_order else 'Sin OS',
                'purchase_order': inv.service_order.purchase_order if inv.service_order else '',
                'type': 'Costo Directo',
                'description': inv.notes or f'Factura de proveedor',
                'status': dict(ProviderInvoice.PAYMENT_STATUS_CHOICES).get(inv.payment_status, inv.payment_status),
                'amount': float(inv.total_amount),
                'paid': float(inv.paid_amount),
                'balance': balance,
                'has_file': bool(inv.invoice_file)
            })
        
        # Ordenar por fecha
        all_items.sort(key=lambda x: x['date'])

        for item in all_items:
            ws.cell(row=data_row, column=1, value=item['date'].strftime('%d/%m/%Y')).border = thin_border
            ws.cell(row=data_row, column=2, value=item['invoice_number']).border = thin_border
            ws.cell(row=data_row, column=3, value=item['service_order']).border = thin_border
            ws.cell(row=data_row, column=4, value=item['purchase_order']).border = thin_border
            ws.cell(row=data_row, column=5, value=item['type']).border = thin_border
            ws.cell(row=data_row, column=6, value=item['description']).border = thin_border
            ws.cell(row=data_row, column=7, value=item['status']).border = thin_border

            # Montos
            total_cell = ws.cell(row=data_row, column=8, value=item['amount'])
            total_cell.number_format = currency_format
            total_cell.border = thin_border
            total_cell.alignment = Alignment(horizontal='right')

            paid_cell = ws.cell(row=data_row, column=9, value=item['paid'])
            paid_cell.number_format = currency_format
            paid_cell.border = thin_border
            paid_cell.alignment = Alignment(horizontal='right')

            balance_cell = ws.cell(row=data_row, column=10, value=item['balance'])
            balance_cell.number_format = currency_format
            balance_cell.border = thin_border
            balance_cell.alignment = Alignment(horizontal='right')

            ws.cell(row=data_row, column=11, value="Sí" if item['has_file'] else "No").border = thin_border
            ws.cell(row=data_row, column=11).alignment = Alignment(horizontal='center')

            totals['total'] += item['amount']
            totals['paid'] += item['paid']
            totals['balance'] += item['balance']

            data_row += 1

        # Fila de totales
        ws.cell(row=data_row, column=1, value="TOTALES").font = Font(bold=True)
        ws.cell(row=data_row, column=1).border = thin_border
        for col in range(2, 8):
            ws.cell(row=data_row, column=col).border = thin_border

        total_total = ws.cell(row=data_row, column=8, value=totals['total'])
        total_total.number_format = currency_format
        total_total.font = Font(bold=True)
        total_total.border = thin_border
        total_total.alignment = Alignment(horizontal='right')

        total_paid = ws.cell(row=data_row, column=9, value=totals['paid'])
        total_paid.number_format = currency_format
        total_paid.font = Font(bold=True)
        total_paid.border = thin_border
        total_paid.alignment = Alignment(horizontal='right')

        total_balance = ws.cell(row=data_row, column=10, value=totals['balance'])
        total_balance.number_format = currency_format
        total_balance.font = Font(bold=True)
        total_balance.border = thin_border
        total_balance.alignment = Alignment(horizontal='right')

        ws.cell(row=data_row, column=11).border = thin_border

        # Ajustar anchos
        column_widths = [12, 15, 18, 15, 15, 30, 12, 15, 15, 15, 10]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # Generar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'estado_cuenta_proveedor_{provider.name.replace(" ", "_")}_{year}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

class CustomsAgentViewSet(viewsets.ModelViewSet):
    queryset = CustomsAgent.objects.all()
    serializer_class = CustomsAgentSerializer
    permission_classes = [IsAdminOrReadOnly]
    search_fields = ['name', 'email']
    filterset_fields = ['is_active']

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.action == 'list' and self.request.query_params.get('is_active') is None:
            queryset = queryset.filter(is_active=True)
        return queryset.order_by('name')

class BankViewSet(viewsets.ModelViewSet):
    """ViewSet para gestión de bancos"""
    queryset = Bank.objects.all()
    serializer_class = BankSerializer
    permission_classes = [IsAdminOrReadOnly]
    search_fields = ['name']
    filterset_fields = ['is_active']

    def get_queryset(self):
        queryset = super().get_queryset()
        # Por defecto mostrar solo activos en listados
        if self.action == 'list' and self.request.query_params.get('is_active') is None:
            queryset = queryset.filter(is_active=True)
        return queryset.order_by('name')

class ShipmentTypeViewSet(viewsets.ModelViewSet):
    queryset = ShipmentType.objects.all()
    serializer_class = ShipmentTypeSerializer
    permission_classes = [IsAdminOrReadOnly]
    search_fields = ['name', 'code']
    filterset_fields = ['is_active']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        if self.action == 'list' and self.request.query_params.get('is_active') is None:
            queryset = queryset.filter(is_active=True)
        return queryset.order_by('name')

class CustomsViewSet(viewsets.ModelViewSet):
    """ViewSet para gestión de aduanas"""
    queryset = Customs.objects.all()
    serializer_class = CustomsSerializer
    permission_classes = [IsAdminOrReadOnly]
    search_fields = ['name', 'code', 'location']
    filterset_fields = ['is_active']

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.action == 'list' and self.request.query_params.get('is_active') is None:
            queryset = queryset.filter(is_active=True)
        return queryset.order_by('name')


class SubClientViewSet(viewsets.ModelViewSet):
    queryset = SubClient.objects.all()
    serializer_class = SubClientSerializer
    permission_classes = [IsAdminOrReadOnly]
    search_fields = ['name']
    filterset_fields = ['is_active', 'parent_client']

    def get_queryset(self):
        queryset = super().get_queryset()
        if self.action == 'list' and self.request.query_params.get('is_active') is None:
            queryset = queryset.filter(is_active=True)
        return queryset.order_by('name')


class ServiceViewSet(viewsets.ModelViewSet):
    """ViewSet para gestión de servicios"""
    queryset = Service.objects.all()
    serializer_class = ServiceSerializer
    permission_classes = [IsAdminOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'description']
    filterset_fields = ['is_active', 'applies_iva']
    ordering_fields = ['id', 'name', 'default_price']
    ordering = ['name']

    def get_queryset(self):
        queryset = super().get_queryset()
        # Por defecto mostrar solo activos
        if self.action == 'list' and self.request.query_params.get('is_active') is None:
            queryset = queryset.filter(is_active=True)
        return queryset

    @action(detail=False, methods=['get'])
    def activos(self, request):
        """Endpoint para obtener solo servicios activos (para dropdowns)"""
        services = self.queryset.filter(is_active=True).order_by('name')
        serializer = self.get_serializer(services, many=True)
        return Response(serializer.data)


class ClientServicePriceViewSet(viewsets.ModelViewSet):
    """ViewSet para tarifario personalizado de clientes"""
    queryset = ClientServicePrice.objects.select_related('client', 'service').all()
    serializer_class = ClientServicePriceSerializer
    permission_classes = [IsAdminOrReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ['client__name', 'service__name', 'service__code']
    filterset_fields = ['client', 'service', 'is_active']

    def get_queryset(self):
        queryset = super().get_queryset()
        # Por defecto mostrar solo activos
        if self.action == 'list' and self.request.query_params.get('is_active') is None:
            queryset = queryset.filter(is_active=True)
        return queryset

    @action(detail=False, methods=['get'], url_path='by-client/(?P<client_id>[^/.]+)')
    def by_client(self, request, client_id=None):
        """Obtener todos los precios personalizados de un cliente específico"""
        prices = self.queryset.filter(client_id=client_id, is_active=True)
        serializer = self.get_serializer(prices, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """Crear múltiples precios personalizados a la vez"""
        items = request.data.get('items', [])
        created = []
        errors = []

        for item in items:
            serializer = self.get_serializer(data=item)
            if serializer.is_valid():
                serializer.save()
                created.append(serializer.data)
            else:
                errors.append({
                    'item': item,
                    'errors': serializer.errors
                })

        return Response({
            'created': created,
            'errors': errors,
            'total_created': len(created),
            'total_errors': len(errors)
        })