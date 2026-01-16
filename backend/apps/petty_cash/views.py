from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from django.db.models import Sum, Case, When, F, DecimalField
from decimal import Decimal
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import PettyCashTransaction, CashCount
from .serializers import PettyCashTransactionSerializer, CashCountSerializer

class PettyCashTransactionViewSet(viewsets.ModelViewSet):
    queryset = PettyCashTransaction.objects.all()
    serializer_class = PettyCashTransactionSerializer

    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter for "Active" (current box) transactions
        active = self.request.query_params.get('active')
        if active == 'true':
            queryset = queryset.filter(cash_count__isnull=True)
            
        # Filter for specific Cash Count (History)
        cash_count_id = self.request.query_params.get('cash_count_id')
        if cash_count_id:
            queryset = queryset.filter(cash_count_id=cash_count_id)
            
        return queryset

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=['get'])
    def balance(self, request):
        """
        Calculates the current balance: Total Income - Total Expenses
        For the ACTIVE box only (unarchived transactions).
        """
        # Only consider transactions not yet part of a Cash Count
        queryset = PettyCashTransaction.objects.filter(cash_count__isnull=True)
        
        aggregates = queryset.aggregate(
            total_income=Sum(Case(
                When(transaction_type='INCOME', then=F('amount')),
                default=0,
                output_field=DecimalField()
            )),
            total_expenses=Sum(Case(
                When(transaction_type='EXPENSE', then=F('amount')),
                default=0,
                output_field=DecimalField()
            ))
        )
        
        income = aggregates['total_income'] or Decimal('0.00')
        expenses = aggregates['total_expenses'] or Decimal('0.00')
        balance = income - expenses
        
        return Response({
            'balance': balance,
            'total_income': income,
            'total_expenses': expenses
        })

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Exportar movimientos de caja chica a Excel con formato profesional"""
        queryset = self.get_queryset()
        
        # Aplicar filtros
        transaction_type = request.query_params.get('transaction_type')
        category_code = request.query_params.get('category_code')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        
        # Also respect 'active' or 'cash_count_id' from get_queryset logic if passed, 
        # but let's ensure we use the filtered queryset from get_queryset() which we did above.
        
        if transaction_type:
            queryset = queryset.filter(transaction_type=transaction_type)
        if category_code:
            queryset = queryset.filter(category_code=category_code)
        if date_from:
            queryset = queryset.filter(transaction_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(transaction_date__lte=date_to)
        
        queryset = queryset.order_by('-transaction_date')
        
        # Mapeo de categorías
        CATEGORY_NAMES = {
            'INGRESO': 'Ingresos / Reembolsos',
            'ALIMENTOS': 'Alimentación / Cafetería',
            'TRANSPORTE': 'Transporte / Combustible / Taxis',
            'PAPELERIA': 'Papelería y Útiles de Oficina',
            'ASEO': 'Artículos de Aseo y Limpieza',
            'MANTENIM': 'Mantenimiento y Reparaciones',
            'TRAMITES': 'Trámites y Diligencias',
            'OTROS': 'Otros Gastos Varios',
        }
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Caja Chica"
        
        # Estilos profesionales - Diseño GPRO
        header_fill = PatternFill(start_color="0F2E4D", end_color="0F2E4D", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        title_font = Font(size=16, bold=True, color="0F2E4D")
        subtitle_font = Font(size=12, bold=True, color="1A4C7A")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        currency_format = '"$"#,##0.00'
        
        # === ENCABEZADO ===
        ws['A1'] = "CAJA CHICA - REGISTRO DE MOVIMIENTOS"
        ws['A1'].font = title_font
        ws.merge_cells('A1:I1')
        
        ws['A2'] = "GPRO LOGISTIC - Agencia Aduanal"
        ws['A2'].font = Font(size=11, color="666666")
        ws.merge_cells('A2:I2')
        
        ws['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A3'].font = Font(size=9, italic=True, color="999999")
        
        # === TABLA DE DATOS ===
        start_row = 5
        ws.cell(row=start_row, column=1, value="DETALLE DE MOVIMIENTOS").font = subtitle_font
        ws.merge_cells(f'A{start_row}:I{start_row}')
        
        # Headers de tabla
        headers = ['Fecha', 'Tipo', 'Concepto', 'Beneficiario', 'Categoría', 
                   'Monto', 'Factura/Ref', 'OS', 'NIT/DUI']
        header_row = start_row + 1
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Datos
        data_row = header_row + 1
        total_income = 0
        total_expense = 0
        
        for transaction in queryset:
            ws.cell(row=data_row, column=1, value=transaction.transaction_date.strftime('%d/%m/%Y')).border = thin_border
            
            # Tipo
            tipo_text = 'Ingreso' if transaction.transaction_type == 'INCOME' else 'Gasto'
            ws.cell(row=data_row, column=2, value=tipo_text).border = thin_border
            
            ws.cell(row=data_row, column=3, value=transaction.concept or '').border = thin_border
            ws.cell(row=data_row, column=4, value=transaction.beneficiary or '').border = thin_border
            
            # Categoría
            category_name = CATEGORY_NAMES.get(transaction.category_code, '') if transaction.category_code else ''
            ws.cell(row=data_row, column=5, value=category_name).border = thin_border
            
            # Monto con formato
            amount_cell = ws.cell(row=data_row, column=6, value=float(transaction.amount))
            amount_cell.number_format = currency_format
            amount_cell.border = thin_border
            amount_cell.alignment = Alignment(horizontal='right')
            
            ws.cell(row=data_row, column=7, value=transaction.reference_number or '').border = thin_border
            ws.cell(row=data_row, column=8, value=transaction.service_order_ref or '').border = thin_border
            
            # Combinar NIT/DUI si existen
            nit_dui = transaction.nit or transaction.dui or ''
            ws.cell(row=data_row, column=9, value=nit_dui).border = thin_border
            
            if transaction.transaction_type == 'INCOME':
                total_income += float(transaction.amount)
            else:
                total_expense += float(transaction.amount)
            
            data_row += 1
        
        # Filas de totales
        data_row += 1
        
        ws.cell(row=data_row, column=1, value="TOTAL INGRESOS").font = Font(bold=True, color="059669")
        ws.cell(row=data_row, column=1).border = thin_border
        for col in range(2, 6):
            ws.cell(row=data_row, column=col).border = thin_border
        
        income_cell = ws.cell(row=data_row, column=6, value=total_income)
        income_cell.number_format = currency_format
        income_cell.font = Font(bold=True, color="059669")
        income_cell.border = thin_border
        income_cell.alignment = Alignment(horizontal='right')
        
        for col in range(7, 10):
            ws.cell(row=data_row, column=col).border = thin_border
        
        data_row += 1
        
        ws.cell(row=data_row, column=1, value="TOTAL GASTOS").font = Font(bold=True, color="DC2626")
        ws.cell(row=data_row, column=1).border = thin_border
        for col in range(2, 6):
            ws.cell(row=data_row, column=col).border = thin_border
        
        expense_cell = ws.cell(row=data_row, column=6, value=total_expense)
        expense_cell.number_format = currency_format
        expense_cell.font = Font(bold=True, color="DC2626")
        expense_cell.border = thin_border
        expense_cell.alignment = Alignment(horizontal='right')
        
        for col in range(7, 10):
            ws.cell(row=data_row, column=col).border = thin_border
        
        data_row += 1
        
        ws.cell(row=data_row, column=1, value="SALDO").font = Font(bold=True, size=11)
        ws.cell(row=data_row, column=1).border = thin_border
        for col in range(2, 6):
            ws.cell(row=data_row, column=col).border = thin_border
        
        balance = total_income - total_expense
        balance_cell = ws.cell(row=data_row, column=6, value=balance)
        balance_cell.number_format = currency_format
        balance_cell.font = Font(bold=True, size=11)
        balance_cell.border = thin_border
        balance_cell.alignment = Alignment(horizontal='right')
        
        for col in range(7, 10):
            ws.cell(row=data_row, column=col).border = thin_border
        
        # Ajustar anchos de columna
        column_widths = [12, 12, 30, 25, 18, 14, 15, 12, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=GPRO_Caja_Chica_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
        wb.save(response)
        return response

class CashCountViewSet(viewsets.ModelViewSet):
    queryset = CashCount.objects.all()
    serializer_class = CashCountSerializer

    def perform_create(self, serializer):
        # Calculate expected balance based ONLY on Open transactions (not yet archived)
        open_transactions = PettyCashTransaction.objects.filter(cash_count__isnull=True)
        
        aggregates = open_transactions.aggregate(
            total_income=Sum(Case(
                When(transaction_type='INCOME', then=F('amount')),
                default=0,
                output_field=DecimalField()
            )),
            total_expenses=Sum(Case(
                When(transaction_type='EXPENSE', then=F('amount')),
                default=0,
                output_field=DecimalField()
            ))
        )
        income = aggregates['total_income'] or Decimal('0.00')
        expenses = aggregates['total_expenses'] or Decimal('0.00')
        calculated_balance = income - expenses
        
        # Save with calculated fields
        actual_balance = serializer.validated_data.get('actual_balance')
        difference = actual_balance - calculated_balance
        
        instance = serializer.save(
            performed_by=self.request.user,
            calculated_balance=calculated_balance,
            difference=difference
        )
        
        # ARCHIVE TRANSACTIONS: Link all currently open transactions to this CashCount
        open_transactions.update(cash_count=instance)


    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Exportar arqueos de caja chica a Excel con formato profesional"""
        queryset = self.get_queryset().order_by('-date')
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Arqueos de Caja"
        
        # Estilos profesionales - Diseño GPRO
        header_fill = PatternFill(start_color="0F2E4D", end_color="0F2E4D", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        title_font = Font(size=16, bold=True, color="0F2E4D")
        subtitle_font = Font(size=12, bold=True, color="1A4C7A")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        currency_format = '"$"#,##0.00'
        
        # === ENCABEZADO ===
        ws['A1'] = "CAJA CHICA - HISTORIAL DE ARQUEOS"
        ws['A1'].font = title_font
        ws.merge_cells('A1:F1')
        
        ws['A2'] = "GPRO LOGISTIC - Agencia Aduanal"
        ws['A2'].font = Font(size=11, color="666666")
        ws.merge_cells('A2:F2')
        
        ws['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A3'].font = Font(size=9, italic=True, color="999999")
        
        # === TABLA DE DATOS ===
        start_row = 5
        ws.cell(row=start_row, column=1, value="DETALLE DE ARQUEOS").font = subtitle_font
        ws.merge_cells(f'A{start_row}:F{start_row}')
        
        # Headers de tabla
        headers = ['Fecha', 'Saldo Sistema', 'Efectivo Contado', 'Diferencia', 'Usuario', 'Notas']
        header_row = start_row + 1
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Datos
        data_row = header_row + 1
        
        for count in queryset:
            ws.cell(row=data_row, column=1, value=count.date.strftime('%d/%m/%Y')).border = thin_border
            
            # Saldo Sistema
            calc_cell = ws.cell(row=data_row, column=2, value=float(count.calculated_balance))
            calc_cell.number_format = currency_format
            calc_cell.border = thin_border
            calc_cell.alignment = Alignment(horizontal='right')
            
            # Efectivo Contado
            actual_cell = ws.cell(row=data_row, column=3, value=float(count.actual_balance))
            actual_cell.number_format = currency_format
            actual_cell.border = thin_border
            actual_cell.alignment = Alignment(horizontal='right')
            
            # Diferencia (colorear si hay diferencia)
            diff_cell = ws.cell(row=data_row, column=4, value=float(count.difference))
            diff_cell.number_format = currency_format
            diff_cell.border = thin_border
            diff_cell.alignment = Alignment(horizontal='right')
            if float(count.difference) != 0:
                diff_cell.font = Font(color="DC2626", bold=True)
            
            # Usuario
            user_name = count.performed_by.get_full_name() if count.performed_by else 'N/A'
            ws.cell(row=data_row, column=5, value=user_name).border = thin_border
            
            # Notas
            ws.cell(row=data_row, column=6, value=count.notes or '').border = thin_border
            
            data_row += 1
        
        # Ajustar anchos de columna
        column_widths = [12, 16, 16, 14, 20, 40]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=GPRO_Arqueos_Caja_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
        wb.save(response)
        return response

    @action(detail=True, methods=['get'])
    def export_denomination_detail(self, request, pk=None):
        """Exportar detalle de denominaciones de un arqueo específico para depósito bancario"""
        cash_count = self.get_object()
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Detalle Denominaciones"
        
        # Estilos profesionales
        header_fill = PatternFill(start_color="0F2E4D", end_color="0F2E4D", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        title_font = Font(size=16, bold=True, color="0F2E4D")
        subtitle_font = Font(size=11, bold=True, color="1A4C7A")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        currency_format = '"$"#,##0.00'
        
        # === ENCABEZADO ===
        ws['A1'] = "DETALLE DE DENOMINACIONES - ARQUEO DE CAJA"
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        
        ws['A2'] = "GPRO LOGISTIC - Agencia Aduanal"
        ws['A2'].font = Font(size=11, color="666666")
        ws.merge_cells('A2:D2')
        
        ws['A3'] = f"Fecha del Arqueo: {cash_count.date.strftime('%d/%m/%Y')}"
        ws['A3'].font = Font(size=10, color="333333")
        ws.merge_cells('A3:D3')
        
        ws['A4'] = f"Realizado por: {cash_count.performed_by.get_full_name() if cash_count.performed_by else 'N/A'}"
        ws['A4'].font = Font(size=9, italic=True, color="666666")
        ws.merge_cells('A4:D4')
        
        # === RESUMEN ===
        start_row = 6
        ws.cell(row=start_row, column=1, value="RESUMEN").font = subtitle_font
        ws.merge_cells(f'A{start_row}:D{start_row}')
        
        summary_row = start_row + 1
        ws.cell(row=summary_row, column=1, value="Saldo Sistema:")
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        summary_cell = ws.cell(row=summary_row, column=2, value=float(cash_count.calculated_balance))
        summary_cell.number_format = currency_format
        summary_cell.font = Font(bold=True)
        
        summary_row += 1
        ws.cell(row=summary_row, column=1, value="Efectivo Contado:")
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        actual_cell = ws.cell(row=summary_row, column=2, value=float(cash_count.actual_balance))
        actual_cell.number_format = currency_format
        actual_cell.font = Font(bold=True, color="059669")
        
        summary_row += 1
        ws.cell(row=summary_row, column=1, value="Diferencia:")
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        diff_cell = ws.cell(row=summary_row, column=2, value=float(cash_count.difference))
        diff_cell.number_format = currency_format
        diff_cell.font = Font(bold=True, color="DC2626" if float(cash_count.difference) != 0 else "333333")
        
        # === TABLA DE DENOMINACIONES ===
        table_start_row = summary_row + 3
        ws.cell(row=table_start_row, column=1, value="DETALLE DE BILLETES Y MONEDAS").font = subtitle_font
        ws.merge_cells(f'A{table_start_row}:D{table_start_row}')
        
        # Headers
        header_row = table_start_row + 1
        headers = ['Denominación', 'Cantidad', 'Subtotal', '']
        for col_num, header in enumerate(headers, 1):
            if header:  # Skip empty header
                cell = ws.cell(row=header_row, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
        
        # Datos de denominaciones
        data_row = header_row + 1
        details = cash_count.details.all().order_by('-denomination')
        
        for detail in details:
            # Denominación
            denom_cell = ws.cell(row=data_row, column=1, value=float(detail.denomination))
            denom_cell.number_format = currency_format
            denom_cell.border = thin_border
            denom_cell.alignment = Alignment(horizontal='right')
            denom_cell.font = Font(bold=True)
            
            # Cantidad
            qty_cell = ws.cell(row=data_row, column=2, value=detail.quantity)
            qty_cell.border = thin_border
            qty_cell.alignment = Alignment(horizontal='center')
            
            # Subtotal
            total_cell = ws.cell(row=data_row, column=3, value=float(detail.total))
            total_cell.number_format = currency_format
            total_cell.border = thin_border
            total_cell.alignment = Alignment(horizontal='right')
            
            data_row += 1
        
        # Total final
        data_row += 1
        ws.cell(row=data_row, column=1, value="TOTAL EFECTIVO").font = Font(bold=True, size=11)
        ws.cell(row=data_row, column=1).border = thin_border
        ws.cell(row=data_row, column=2).border = thin_border
        
        total_cell = ws.cell(row=data_row, column=3, value=float(cash_count.actual_balance))
        total_cell.number_format = currency_format
        total_cell.font = Font(bold=True, size=11, color="059669")
        total_cell.border = thin_border
        total_cell.alignment = Alignment(horizontal='right')
        
        # Notas
        if cash_count.notes:
            notes_row = data_row + 3
            ws.cell(row=notes_row, column=1, value="OBSERVACIONES:").font = Font(bold=True, size=10)
            ws.merge_cells(f'A{notes_row}:D{notes_row}')
            
            notes_content_row = notes_row + 1
            ws.cell(row=notes_content_row, column=1, value=cash_count.notes)
            ws.cell(row=notes_content_row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(f'A{notes_content_row}:D{notes_content_row}')
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 2
        
        # Respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=GPRO_Detalle_Denominaciones_Arqueo_{cash_count.date.strftime("%Y%m%d")}.xlsx'
        
        wb.save(response)
        return response
