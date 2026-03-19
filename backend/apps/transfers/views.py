from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.core.exceptions import ValidationError
from django.db.models import Q, Sum
from django.db import transaction
from django.http import HttpResponse, FileResponse
from django.utils.text import slugify
from django_filters import rest_framework as filters
from .models import Transfer, TransferPayment, BatchPayment, ProviderCreditNote, CreditNoteApplication, ProviderInvoicePayment
from .serializers import (
    TransferSerializer, TransferListSerializer, TransferPaymentSerializer,
    BatchPaymentSerializer, BatchPaymentDetailSerializer,
    ProviderCreditNoteListSerializer, ProviderCreditNoteDetailSerializer,
    ProviderCreditNoteCreateSerializer, ApplyCreditNoteSerializer, CreditNoteApplicationSerializer
)
from apps.users.permissions import IsAnyOperativo, IsOperativo2OrAdmin, TransferApprovalPermission, IsOperativo
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime
import os
import io
import zipfile

class TransferFilter(filters.FilterSet):
    """Filtros avanzados para transfers"""
    date_from = filters.DateFilter(field_name='transaction_date', lookup_expr='gte')
    date_to = filters.DateFilter(field_name='transaction_date', lookup_expr='lte')
    payment_date_from = filters.DateFilter(field_name='payment_date', lookup_expr='gte')
    payment_date_to = filters.DateFilter(field_name='payment_date', lookup_expr='lte')
    client = filters.NumberFilter(field_name='service_order__client__id')
    min_amount = filters.NumberFilter(field_name='amount', lookup_expr='gte')
    max_amount = filters.NumberFilter(field_name='amount', lookup_expr='lte')
    
    class Meta:
        model = Transfer
        fields = ['transfer_type', 'status', 'service_order', 'provider', 'payment_method']

class TransferViewSet(viewsets.ModelViewSet):
    queryset = Transfer.objects.select_related(
        'service_order',
        'service_order__client',
        'provider',
        'bank',
        'created_by'
    ).all()
    serializer_class = TransferSerializer
    permission_classes = [IsAnyOperativo]
    parser_classes = (MultiPartParser, FormParser, JSONParser)
    filterset_class = TransferFilter
    search_fields = ['description', 'invoice_number', 'service_order__order_number']
    ordering_fields = ['transaction_date', 'amount', 'created_at']
    ordering = ['-transaction_date']
    pagination_class = None
    DEFAULT_EXPORT_MAX_FILES = 20
    
    def get_serializer_class(self):
        if self.action == 'list':
            return TransferListSerializer
        return TransferSerializer

    def _parse_ids(self, raw_value):
        """Normalize list/CSV IDs from request payload."""
        if raw_value is None:
            return []
        if isinstance(raw_value, str):
            raw_items = [item.strip() for item in raw_value.split(',') if item.strip()]
        elif isinstance(raw_value, list):
            raw_items = raw_value
        else:
            raw_items = [raw_value]

        normalized = []
        for item in raw_items:
            try:
                normalized.append(int(item))
            except (TypeError, ValueError):
                continue
        return normalized

    def _zip_unique_name(self, used_names, proposed_name):
        """Prevent duplicate names inside ZIP file."""
        base_name = proposed_name.replace('\\', '/')
        if base_name not in used_names:
            used_names.add(base_name)
            return base_name

        root, ext = os.path.splitext(base_name)
        idx = 2
        while True:
            candidate = f"{root}_{idx}{ext}"
            if candidate not in used_names:
                used_names.add(candidate)
                return candidate
            idx += 1

    def _add_file_to_zip(self, zip_file, used_names, file_field, zip_subpath, only_pdf=False):
        """Add file field to ZIP if available and valid."""
        if not file_field:
            return False

        name = (getattr(file_field, 'name', '') or '').strip()
        if not name:
            return False

        _, ext = os.path.splitext(name)
        ext = ext.lower()
        if only_pdf and ext != '.pdf':
            return False

        storage = getattr(file_field, 'storage', None)
        if not storage or not storage.exists(name):
            return False

        base_name = os.path.basename(name) or 'documento'
        zip_name = self._zip_unique_name(used_names, f"{zip_subpath}/{base_name}")

        with storage.open(name, 'rb') as source:
            zip_file.writestr(zip_name, source.read())

        return True

    def _safe_folder_part(self, value, fallback):
        part = slugify(str(value or '').strip())
        return part or fallback

    def _build_transfer_folder(self, transfer):
        provider_name = None
        if transfer.provider:
            provider_name = transfer.provider.name
        elif transfer.beneficiary_name:
            provider_name = transfer.beneficiary_name

        provider_part = self._safe_folder_part(provider_name, 'sin-proveedor')
        type_part = self._safe_folder_part(transfer.transfer_type, 'sin-tipo')

        if transfer.service_order and transfer.service_order.order_number:
            os_part = self._safe_folder_part(
                transfer.service_order.order_number,
                f'os-{transfer.service_order_id or "na"}'
            )
            return (
                f"gastos/os_{os_part}/"
                f"proveedor_{provider_part}/"
                f"{type_part}_transfer_{transfer.id}"
            )

        return (
            f"gastos/operacion/"
            f"proveedor_{provider_part}/"
            f"{type_part}_transfer_{transfer.id}"
        )

    def _build_provider_invoice_folder(self, invoice):
        provider_part = self._safe_folder_part(
            invoice.provider.name if invoice.provider else None,
            'sin-proveedor'
        )
        invoice_part = self._safe_folder_part(
            invoice.invoice_number,
            f'factura-{invoice.id}'
        )

        if invoice.service_order and invoice.service_order.order_number:
            os_part = self._safe_folder_part(
                invoice.service_order.order_number,
                f'os-{invoice.service_order_id or "na"}'
            )
            return (
                f"costos_directos/os_{os_part}/"
                f"proveedor_{provider_part}/"
                f"factura_{invoice_part}_{invoice.id}"
            )

        return (
            f"costos_directos/operacion/"
            f"proveedor_{provider_part}/"
            f"factura_{invoice_part}_{invoice.id}"
        )

    def _validate_export_file(self, file_field, only_pdf=False):
        if not file_field:
            return False

        name = (getattr(file_field, 'name', '') or '').strip()
        if not name:
            return False

        _, ext = os.path.splitext(name)
        ext = ext.lower()
        if only_pdf and ext != '.pdf':
            return False

        storage = getattr(file_field, 'storage', None)
        if not storage or not storage.exists(name):
            return False

        return True

    def _collect_export_candidates(
        self,
        transfers,
        provider_invoices,
        only_pdf=False,
        include_payment_proofs=True,
        payment_proofs_only_pdf=False,
    ):
        candidates = []

        for transfer in transfers:
            transfer_folder = self._build_transfer_folder(transfer)

            if self._validate_export_file(transfer.invoice_file, only_pdf=only_pdf):
                candidates.append((
                    transfer.invoice_file,
                    f"{transfer_folder}/soporte",
                    only_pdf,
                ))

            if include_payment_proofs:
                proof_only_pdf = only_pdf and payment_proofs_only_pdf
                for payment in transfer.payments.all():
                    if getattr(payment, 'is_deleted', False):
                        continue
                    if self._validate_export_file(payment.proof_file, only_pdf=proof_only_pdf):
                        candidates.append((
                            payment.proof_file,
                            f"{transfer_folder}/pagos/pago_{payment.id}",
                            proof_only_pdf,
                        ))

        for invoice in provider_invoices:
            invoice_folder = self._build_provider_invoice_folder(invoice)

            if self._validate_export_file(invoice.invoice_file, only_pdf=only_pdf):
                candidates.append((
                    invoice.invoice_file,
                    f"{invoice_folder}/soporte",
                    only_pdf,
                ))

            if include_payment_proofs:
                proof_only_pdf = only_pdf and payment_proofs_only_pdf
                for payment in invoice.invoice_payments.all():
                    if getattr(payment, 'is_deleted', False):
                        continue
                    if self._validate_export_file(payment.proof_file, only_pdf=proof_only_pdf):
                        candidates.append((
                            payment.proof_file,
                            f"{invoice_folder}/pagos/pago_{payment.id}",
                            proof_only_pdf,
                        ))

        return candidates
    
    def _validate_transfer_edit(self, transfer, request):
        """
        Validar si un gasto puede ser editado.

        RESTRICCIONES:
        1. Si la orden está cerrada, no se pueden modificar campos financieros
        2. Si está facturado con DTE emitido, solo campos limitados
        3. Si está pagado, no se puede editar el monto
        4. Solo admin/operativo2 puede aprobar
        """
        errors = []
        new_status = request.data.get('status')
        new_amount = request.data.get('amount')

        # 0. Validar orden cerrada (Bloquear edición financiera si OS cerrada)
        if transfer.service_order and transfer.service_order.status == 'cerrada':
            financial_fields = {'amount', 'transfer_type', 'provider', 'currency', 'exchange_rate'}
            requested_fields = set(request.data.keys())
            blocked_fields = requested_fields & financial_fields
            if blocked_fields:
                return {
                    'error': 'La orden de servicio está cerrada.',
                    'detail': f'No se pueden modificar campos financieros de un gasto en una orden cerrada: {", ".join(blocked_fields)}',
                    'code': 'ORDER_CLOSED'
                }

        # 1. Validar factura con DTE emitido
        if transfer.invoice and transfer.invoice.is_dte_issued:
            # Solo permitir editar campos no financieros
            financial_fields = {'amount', 'customer_markup_percentage', 'iva_type', 'customer_iva_type'}
            requested_fields = set(request.data.keys())
            blocked_fields = requested_fields & financial_fields

            if blocked_fields:
                return {
                    'error': 'Este gasto está vinculado a una factura con DTE emitido.',
                    'detail': f'No se pueden modificar los campos financieros: {", ".join(blocked_fields)}',
                    'code': 'DTE_ISSUED'
                }

        # 2. Validar cambio de monto si tiene pagos
        if new_amount is not None:
            from decimal import Decimal
            new_amount_decimal = Decimal(str(new_amount))
            if transfer.paid_amount > 0 and new_amount_decimal < transfer.paid_amount:
                return {
                    'error': 'No se puede reducir el monto por debajo de lo ya pagado.',
                    'detail': f'Monto pagado: ${transfer.paid_amount}',
                    'code': 'AMOUNT_BELOW_PAID'
                }

        # 3. Validar permiso de aprobación
        if new_status == 'aprobado':
            if request.user.role not in ['admin', 'operativo2']:
                return {
                    'error': 'No tiene permisos para aprobar gastos.',
                    'detail': 'Contacte a un supervisor.',
                    'code': 'APPROVAL_FORBIDDEN'
                }

        # 4. Validar cambio de estado desde pagado
        if transfer.status == 'pagado' and new_status and new_status != 'pagado':
            if request.user.role != 'admin':
                return {
                    'error': 'No se puede cambiar el estado de un gasto pagado.',
                    'detail': 'Solo un administrador puede realizar esta acción.',
                    'code': 'PAID_STATUS_LOCKED'
                }

        return None  # Sin errores

    def update(self, request, *args, **kwargs):
        """
        Override update con validaciones de integridad.
        """
        transfer = self.get_object()

        validation_error = self._validate_transfer_edit(transfer, request)
        if validation_error:
            return Response(validation_error, status=status.HTTP_400_BAD_REQUEST)

        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        """
        Override partial_update con validaciones de integridad.
        """
        transfer = self.get_object()

        validation_error = self._validate_transfer_edit(transfer, request)
        if validation_error:
            return Response(validation_error, status=status.HTTP_400_BAD_REQUEST)

        return super().partial_update(request, *args, **kwargs)
    
    def create(self, request, *args, **kwargs):
        """
        Override create para agregar warning de margen negativo.

        AUDITORÍA #8: Advertencia cuando el margen de utilidad es negativo
        (pérdida), pero sin bloquear la operación.
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        transfer = self.perform_create(serializer)

        # Calcular si hay margen negativo
        warning_message = None
        profit = transfer.get_profit()
        if profit < 0:
            cost = transfer.amount * (transfer.exchange_rate or 1)
            sale_price = transfer.get_customer_base_price()
            markup = transfer.customer_markup_percentage or 0
            warning_message = (
                f'ADVERTENCIA: Margen negativo detectado. '
                f'Costo: ${cost:.2f}, Precio venta: ${sale_price:.2f}, '
                f'Pérdida: ${abs(profit):.2f}. '
                f'Margen configurado: {markup}%. '
                f'Verifique la configuración de cobro al cliente.'
            )

        response_data = self.get_serializer(transfer).data

        if warning_message:
            response_data['warning'] = warning_message
            response_data['has_negative_margin'] = True

        headers = self.get_success_headers(response_data)
        return Response(response_data, status=status.HTTP_201_CREATED, headers=headers)

    def perform_create(self, serializer):
        """Asignar usuario que crea la transferencia con validación de OS cerrada"""
        from apps.orders.models import ServiceOrder
        from rest_framework.exceptions import ValidationError

        # Obtener OS desde los datos validados del serializer
        service_order = serializer.validated_data.get('service_order')

        # Si no está en validados (poco común), buscar en request data
        if not service_order:
            order_id = self.request.data.get('service_order')
            if order_id:
                try:
                    service_order = ServiceOrder.objects.get(id=order_id)
                except ServiceOrder.DoesNotExist:
                    pass

        # Validar bloqueo si la OS está cerrada: NO SE PUEDEN AGREGAR GASTOS NUEVOS
        if service_order and service_order.status == 'cerrada':
            raise ValidationError({
                'error': 'No se pueden agregar nuevos gastos a una orden de servicio cerrada.',
                'code': 'ORDER_CLOSED'
            })

        serializer.context['request'] = self.request
        transfer = serializer.save(created_by=self.request.user)
        transfer._current_user = self.request.user
        return transfer
    
    def perform_update(self, serializer):
        """Set current user for signal on update"""
        serializer.context['request'] = self.request
        transfer = serializer.save()
        transfer._current_user = self.request.user
        return transfer
    
    def destroy(self, request, *args, **kwargs):
        """
        Eliminar un gasto con validaciones de integridad.
        """
        transfer = self.get_object()

        # 0. Validar orden cerrada
        if transfer.service_order and transfer.service_order.status == 'cerrada':
            return Response(
                {
                    'error': 'No se puede eliminar gastos de una orden cerrada.',
                    'code': 'ORDER_CLOSED'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # 1. Validar si está facturado al cliente
        if transfer.invoice:
            invoice_number = transfer.invoice.invoice_number or f"#{transfer.invoice.id}"
            return Response(
                {
                    'error': f'No se puede eliminar este gasto porque ya fue facturado al cliente.',
                    'detail': f'Factura asociada: {invoice_number}',
                    'code': 'INVOICED'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # 2. Validar si tiene pagos registrados
        payments_count = transfer.payments.filter(is_deleted=False).count()
        if payments_count > 0:
            return Response(
                {
                    'error': f'No se puede eliminar este gasto porque tiene {payments_count} pago(s) registrado(s).',
                    'detail': 'Primero debe eliminar o reversar los pagos asociados.',
                    'code': 'HAS_PAYMENTS'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # 3. Validar si está pagado
        if transfer.status == 'pagado':
            return Response(
                {
                    'error': 'No se puede eliminar un gasto que ya está marcado como pagado.',
                    'detail': 'Cambie el estado antes de eliminar o contacte a un administrador.',
                    'code': 'ALREADY_PAID'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # 4. Solo admin puede eliminar gastos aprobados
        if transfer.status == 'aprobado':
            if request.user.role != 'admin':
                return Response(
                    {
                        'error': 'Solo un administrador puede eliminar gastos aprobados.',
                        'detail': 'Contacte a un administrador para realizar esta acción.',
                        'code': 'APPROVED_REQUIRES_ADMIN'
                    },
                    status=status.HTTP_403_FORBIDDEN
                )

        # Proceder con la eliminación
        transfer._current_user = request.user
        transfer.delete()

        return Response(
            {'message': 'Gasto eliminado correctamente.'},
            status=status.HTTP_204_NO_CONTENT
        )

    @action(detail=False, methods=['get'], permission_classes=[IsAnyOperativo])
    def export_excel(self, request):
        """Exportar transfers y provider_invoices a Excel con formato profesional"""
        from .models import ProviderInvoice

        transfer_queryset = self.filter_queryset(self.get_queryset())
        
        # Obtener filtros de la query
        query_params = request.query_params
        
        # Mapear filtros para ProviderInvoice
        invoice_filters = Q()
        if query_params.get('date_from'):
            invoice_filters &= Q(issue_date__gte=query_params.get('date_from'))
        if query_params.get('date_to'):
            invoice_filters &= Q(issue_date__lte=query_params.get('date_to'))
        if query_params.get('provider'):
            invoice_filters &= Q(provider_id=query_params.get('provider'))
        if query_params.get('service_order'):
            invoice_filters &= Q(service_order_id=query_params.get('service_order'))
        if query_params.get('status'):
            invoice_filters &= Q(payment_status=query_params.get('status'))
            
        provider_invoice_queryset = ProviderInvoice.objects.filter(invoice_filters).select_related(
            'provider', 'service_order'
        )

        # Unificar datos
        all_data = []

        # Transfers
        for transfer in transfer_queryset:
            all_data.append({
                'fecha': transfer.transaction_date,
                'tipo': transfer.get_transfer_type_display(),
                'estado': transfer.get_status_display(),
                'monto': float(transfer.amount),
                'descripcion': transfer.description or '',
                'os': transfer.service_order.order_number if transfer.service_order else '',
                'po': transfer.service_order.purchase_order if transfer.service_order else '',
                'proveedor': transfer.provider.name if transfer.provider else transfer.beneficiary_name or '',
                'metodo_pago': transfer.get_payment_method_display() if transfer.payment_method else '',
                'factura': transfer.invoice_number or '',
                'fecha_pago': transfer.payment_date,
                'generation_code': transfer.generation_code or '',
                'reception_stamp': transfer.reception_stamp or '',
            })

        # Provider Invoices
        for invoice in provider_invoice_queryset:
            all_data.append({
                'fecha': invoice.issue_date,
                'tipo': "Costo Directo",
                'estado': dict(ProviderInvoice.PAYMENT_STATUS_CHOICES).get(invoice.payment_status),
                'monto': float(invoice.total_amount),
                'descripcion': invoice.notes or f"Factura de {invoice.provider.name}",
                'os': invoice.service_order.order_number if invoice.service_order else '',
                'po': invoice.service_order.purchase_order if invoice.service_order else '',
                'proveedor': invoice.provider.name,
                'metodo_pago': '', # ProviderInvoice no tiene un método de pago directo
                'factura': invoice.invoice_number,
                'fecha_pago': invoice.payment_date,
                'generation_code': invoice.generation_code or '',
                'reception_stamp': invoice.reception_stamp or '',
            })
            
        # Ordenar por fecha
        all_data.sort(key=lambda x: x['fecha'], reverse=True)


        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cuentas por Pagar"

        # Estilos profesionales - Diseño GPRO
        header_fill = PatternFill(start_color="0F2E4D", end_color="0F2E4D", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        title_font = Font(size=16, bold=True, color="0F2E4D")
        subtitle_font = Font(size=12, bold=True, color="1A4C7A")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        currency_format = '"$"#,##0.00'

        # === ENCABEZADO ===
        ws['A1'] = "CUENTAS POR PAGAR (UNIFICADO)"
        ws['A1'].font = title_font
        ws.merge_cells('A1:M1')

        ws['A2'] = "GPRO LOGISTIC - Agencia Aduanal"
        ws['A2'].font = Font(size=11, color="666666")
        ws.merge_cells('A2:M2')

        ws['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A3'].font = Font(size=9, italic=True, color="999999")

        # === TABLA DE DATOS ===
        start_row = 5
        ws.cell(row=start_row, column=1, value="DETALLE DE PAGOS").font = subtitle_font
        ws.merge_cells(f'A{start_row}:M{start_row}')

        # Headers de tabla
        headers = ['Fecha', 'Tipo', 'Estado', 'Monto', 'Descripción', 'OS', 'PO',
                   'Proveedor', 'Método Pago', 'Factura', 'Fecha Pago', 'Cód. Generación', 'Sello Recepción']
        header_row = start_row + 1

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Datos
        data_row = header_row + 1
        total_amount = 0

        for item in all_data:
            ws.cell(row=data_row, column=1, value=item['fecha'].strftime('%d/%m/%Y')).border = thin_border
            ws.cell(row=data_row, column=2, value=item['tipo']).border = thin_border
            ws.cell(row=data_row, column=3, value=item['estado']).border = thin_border

            # Columna de monto con formato
            amount_cell = ws.cell(row=data_row, column=4, value=item['monto'])
            amount_cell.number_format = currency_format
            amount_cell.border = thin_border
            amount_cell.alignment = Alignment(horizontal='right')

            ws.cell(row=data_row, column=5, value=item['descripcion']).border = thin_border
            ws.cell(row=data_row, column=6, value=item['os']).border = thin_border
            ws.cell(row=data_row, column=7, value=item['po']).border = thin_border
            ws.cell(row=data_row, column=8, value=item['proveedor']).border = thin_border
            ws.cell(row=data_row, column=9, value=item['metodo_pago']).border = thin_border
            ws.cell(row=data_row, column=10, value=item['factura']).border = thin_border
            ws.cell(row=data_row, column=11, value=item['fecha_pago'].strftime('%d/%m/%Y') if item['fecha_pago'] else '').border = thin_border
            ws.cell(row=data_row, column=12, value=item['generation_code']).border = thin_border
            ws.cell(row=data_row, column=13, value=item['reception_stamp']).border = thin_border

            total_amount += item['monto']
            data_row += 1

        # Fila de totales
        ws.cell(row=data_row, column=1, value="TOTALES").font = Font(bold=True)
        ws.cell(row=data_row, column=1).border = thin_border
        for col in range(2, 4):
            ws.cell(row=data_row, column=col).border = thin_border

        total_cell = ws.cell(row=data_row, column=4, value=total_amount)
        total_cell.number_format = currency_format
        total_cell.font = Font(bold=True)
        total_cell.border = thin_border
        total_cell.alignment = Alignment(horizontal='right')

        for col in range(5, 14):
            ws.cell(row=data_row, column=col).border = thin_border

        # Ajustar anchos de columna
        column_widths = [12, 18, 12, 14, 40, 15, 15, 25, 16, 15, 12, 35, 35]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # Respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=GPRO_Cuentas_Por_Pagar_{datetime.now().strftime("%Y%m%d")}.xlsx'

        wb.save(response)
        return response

    @action(detail=False, methods=['post'], permission_classes=[IsAnyOperativo])
    def export_documents(self, request):
        """
        Exporta documentos de soporte (facturas/comprobantes) a ZIP.

        Payload esperado:
        {
            "transfer_ids": [1,2,3],
            "provider_invoice_ids": [10,11],
            "only_pdf": true,
            "include_payment_proofs": true
        }
        """
        from .models import ProviderInvoice

        transfer_ids = self._parse_ids(request.data.get('transfer_ids'))
        provider_invoice_ids = self._parse_ids(request.data.get('provider_invoice_ids'))

        only_pdf = str(request.data.get('only_pdf', 'true')).lower() in ('1', 'true', 'yes', 'on')
        include_payment_proofs = str(request.data.get('include_payment_proofs', 'true')).lower() in ('1', 'true', 'yes', 'on')
        payment_proofs_only_pdf = str(request.data.get('payment_proofs_only_pdf', 'false')).lower() in ('1', 'true', 'yes', 'on')
        preview_only = str(request.data.get('preview_only', 'false')).lower() in ('1', 'true', 'yes', 'on')

        raw_max_files = request.data.get('max_files', self.DEFAULT_EXPORT_MAX_FILES)
        try:
            max_files = int(raw_max_files)
        except (TypeError, ValueError):
            max_files = self.DEFAULT_EXPORT_MAX_FILES

        if max_files <= 0:
            max_files = self.DEFAULT_EXPORT_MAX_FILES

        if not transfer_ids and not provider_invoice_ids:
            return Response(
                {'error': 'Debe enviar al menos un ID de transferencia o factura de proveedor.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        transfers = Transfer.objects.filter(id__in=transfer_ids).select_related(
            'provider', 'service_order'
        ).prefetch_related('payments')

        provider_invoices = ProviderInvoice.objects.filter(id__in=provider_invoice_ids).select_related(
            'provider', 'service_order'
        ).prefetch_related('invoice_payments')

        candidates = self._collect_export_candidates(
            transfers,
            provider_invoices,
            only_pdf=only_pdf,
            include_payment_proofs=include_payment_proofs,
            payment_proofs_only_pdf=payment_proofs_only_pdf,
        )

        files_found = len(candidates)

        if preview_only:
            return Response({
                'total_files': files_found,
                'selected_transfers': transfers.count(),
                'selected_provider_invoices': provider_invoices.count(),
                'max_files': max_files,
                'exceeds_limit': files_found > max_files,
            })

        if files_found == 0:
            return Response(
                {
                    'error': 'No se encontraron documentos para exportar con los criterios enviados.',
                    'detail': 'Verifique que los registros tengan soporte adjunto y coincidan con el tipo de archivo solicitado.'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        if files_found > max_files:
            return Response(
                {
                    'error': f'Se excedió el límite máximo de {max_files} PDF por exportación.',
                    'detail': f'Se detectaron {files_found} archivos para exportar. Reduzca su selección e intente nuevamente.',
                    'code': 'MAX_FILES_EXCEEDED',
                    'total_files': files_found,
                    'max_files': max_files,
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        buffer = io.BytesIO()
        used_names = set()
        files_added = 0

        with zipfile.ZipFile(buffer, 'w', compression=zipfile.ZIP_DEFLATED) as zip_file:
            for file_field, zip_subpath, file_only_pdf in candidates:
                if self._add_file_to_zip(
                    zip_file,
                    used_names,
                    file_field,
                    zip_subpath,
                    only_pdf=file_only_pdf,
                ):
                    files_added += 1

        buffer.seek(0)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"GPRO_Documentos_Pagos_{timestamp}.zip"

        response = HttpResponse(buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Resumen de transfers por tipo y estado"""
        queryset = self.filter_queryset(self.get_queryset())
        
        summary = {
            'total_transfers': queryset.count(),
            'total_amount': queryset.aggregate(Sum('amount'))['amount__sum'] or 0,
            'by_type': {},
            'by_status': {}
        }
        
        # Por tipo
        for type_choice in Transfer.TYPE_CHOICES:
            type_code = type_choice[0]
            type_data = queryset.filter(transfer_type=type_code)
            summary['by_type'][type_code] = {
                'label': type_choice[1],
                'count': type_data.count(),
                'amount': type_data.aggregate(Sum('amount'))['amount__sum'] or 0
            }
        
        # Por estado
        for status_choice in Transfer.STATUS_CHOICES:
            status_code = status_choice[0]
            status_data = queryset.filter(status=status_code)
            summary['by_status'][status_code] = {
                'label': status_choice[1],
                'count': status_data.count(),
                'amount': status_data.aggregate(Sum('amount'))['amount__sum'] or 0
            }
        
        return Response(summary)
    
    @action(detail=True, methods=['get'], permission_classes=[IsOperativo])
    def download_invoice(self, request, pk=None):
        """Descargar el comprobante adjunto de una transferencia"""
        transfer = self.get_object()
        
        if not transfer.invoice_file:
            return Response(
                {'error': 'Esta transferencia no tiene comprobante adjunto'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        try:
            file_path = transfer.invoice_file.path
            
            if not os.path.exists(file_path):
                return Response(
                    {'error': 'El archivo no se encuentra en el servidor'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Detectar tipo de contenido basado en la extensión
            filename = os.path.basename(file_path)
            ext = filename.lower().split('.')[-1]
            
            content_types = {
                'pdf': 'application/pdf',
                'jpg': 'image/jpeg',
                'jpeg': 'image/jpeg',
                'png': 'image/png'
            }
            content_type = content_types.get(ext, 'application/octet-stream')
            
            response = FileResponse(
                open(file_path, 'rb'),
                content_type=content_type
            )
            
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
            
        except Exception as e:
            return Response(
                {'error': f'Error al descargar el archivo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'], permission_classes=[IsOperativo])
    def register_payment(self, request, pk=None):
        """Registrar un pago parcial o total a una transferencia (gasto a proveedor)"""
        from .models import TransferPayment
        from decimal import Decimal
        
        transfer = self.get_object()
        
        # Validar que la transferencia no esté completamente pagada
        if transfer.status == 'pagado' and transfer.balance <= 0:
            return Response(
                {'error': 'Esta transferencia ya está completamente pagada'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar datos requeridos
        amount = request.data.get('amount')
        if not amount:
            return Response(
                {'error': 'El monto es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            amount = Decimal(str(amount))
            if amount <= 0:
                return Response(
                    {'error': 'El monto debe ser mayor a cero'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validar que el monto no exceda el saldo pendiente
            if amount > transfer.balance:
                return Response(
                    {'error': f'El monto excede el saldo pendiente de ${transfer.balance}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except (ValueError, TypeError):
            return Response(
                {'error': 'Monto inválido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Actualizar método de pago y banco en el transfer si es el primer pago
        if transfer.paid_amount == 0:
            payment_method = request.data.get('payment_method', 'transferencia')
            transfer.payment_method = payment_method
            
            # Solo actualizar banco si el método requiere banco
            if payment_method in ['transferencia', 'cheque', 'tarjeta']:
                bank_id = request.data.get('bank')
                if bank_id:
                    transfer.bank_id = bank_id
            
            transfer.save(update_fields=['payment_method', 'bank'])
        
        # Crear el pago
        payment = TransferPayment(
            transfer=transfer,
            amount=amount,
            payment_date=request.data.get('payment_date'),
            payment_method=request.data.get('payment_method', 'transferencia'),
            reference_number=request.data.get('reference', ''),
            notes=request.data.get('notes', ''),
            created_by=request.user
        )
        
        # Manejar archivo de comprobante si existe
        if 'proof_file' in request.FILES:
            payment.proof_file = request.FILES['proof_file']
        
        payment.save()
        
        # Refrescar el transfer para obtener los valores actualizados
        transfer.refresh_from_db()
        
        return Response({
            'message': 'Pago registrado exitosamente',
            'payment': {
                'id': payment.id,
                'amount': str(payment.amount),
                'payment_date': payment.payment_date,
                'payment_method': payment.payment_method,
                'reference_number': payment.reference_number
            },
            'transfer': {
                'id': transfer.id,
                'paid_amount': str(transfer.paid_amount),
                'balance': str(transfer.balance),
                'status': transfer.status,
                'status_display': transfer.get_status_display()
            }
        }, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['post'], permission_classes=[IsOperativo])
    def register_credit_note(self, request, pk=None):
        """Registrar una nota de crédito del proveedor"""
        from .models import TransferPayment
        from decimal import Decimal
        
        transfer = self.get_object()
        
        # Validar que la transferencia no esté completamente pagada
        if transfer.status == 'pagado' and transfer.balance <= 0:
            return Response(
                {'error': 'Esta transferencia ya está completamente pagada'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar datos requeridos
        amount = request.data.get('amount')
        note_number = request.data.get('note_number')
        reason = request.data.get('reason', '')
        
        if not amount or not note_number:
            return Response(
                {'error': 'El monto y número de nota de crédito son requeridos'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            amount = Decimal(str(amount))
            if amount <= 0:
                return Response(
                    {'error': 'El monto debe ser mayor a cero'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validar que el monto no exceda el saldo pendiente
            if amount > transfer.balance:
                return Response(
                    {'error': f'El monto excede el saldo pendiente de ${transfer.balance}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except (ValueError, TypeError):
            return Response(
                {'error': 'Monto inválido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Crear el pago como nota de crédito
        payment = TransferPayment(
            transfer=transfer,
            amount=amount,
            payment_date=request.data.get('issue_date'),
            payment_method='nota_credito',
            reference_number=note_number,
            notes=f"Nota de Crédito: {reason}" if reason else "Nota de Crédito",
            created_by=request.user
        )
        
        # Manejar archivo PDF de la nota de crédito si existe
        if 'pdf_file' in request.FILES:
            payment.proof_file = request.FILES['pdf_file']
        
        payment.save()
        
        # Refrescar el transfer
        transfer.refresh_from_db()
        
        return Response({
            'message': 'Nota de crédito registrada exitosamente',
            'credit_note': {
                'id': payment.id,
                'amount': str(payment.amount),
                'note_number': payment.reference_number,
                'issue_date': payment.payment_date,
                'reason': reason
            },
            'transfer': {
                'id': transfer.id,
                'paid_amount': str(transfer.paid_amount),
                'balance': str(transfer.balance),
                'status': transfer.status,
                'status_display': transfer.get_status_display()
            }
        }, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['get'], permission_classes=[IsOperativo])
    def detail_with_payments(self, request, pk=None):
        """Obtener detalle completo de una transferencia incluyendo historial de pagos"""
        from .models import TransferPayment
        from decimal import Decimal
        
        transfer = self.get_object()
        
        # Serializar transferencia
        transfer_data = TransferSerializer(transfer, context={'request': request}).data
        
        # Obtener historial de pagos
        payments = TransferPayment.objects.filter(
            transfer=transfer,
            is_deleted=False
        ).order_by('-payment_date')
        
        payments_data = [{
            'id': p.id,
            'amount': str(p.amount),
            'payment_date': p.payment_date,
            'payment_method': p.payment_method,
            'payment_method_display': p.get_payment_method_display(),
            'reference_number': p.reference_number,
            'notes': p.notes,
            'proof_file': p.proof_file.url if p.proof_file else None,
            'created_by': p.created_by.username if p.created_by else None,
            'created_at': p.created_at
        } for p in payments]
        
        # Identificar notas de crédito (pagos con método nota_credito)
        credit_notes = [p for p in payments_data if p['payment_method'] == 'nota_credito']
        
        # Enriquecer notas de crédito con el archivo PDF original si no lo tienen en el pago
        for cn in credit_notes:
            if not cn['proof_file']:
                try:
                    # Buscar la NC original usando el número de referencia (que es el número de nota)
                    provider_nc = ProviderCreditNote.objects.filter(
                        provider=transfer.provider,
                        note_number=cn['reference_number'],
                        is_deleted=False
                    ).first()
                    
                    if provider_nc and provider_nc.pdf_file:
                        cn['proof_file'] = provider_nc.pdf_file.url
                except:
                    pass

        regular_payments = [p for p in payments_data if p['payment_method'] != 'nota_credito']
        
        transfer_data['payments'] = regular_payments
        transfer_data['credit_notes'] = credit_notes
        transfer_data['total_payments_count'] = len(regular_payments)
        transfer_data['total_credit_notes_count'] = len(credit_notes)
        transfer_data['credited_amount'] = str(sum(Decimal(cn['amount']) for cn in credit_notes))

        return Response(transfer_data)


class BatchPaymentViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar pagos agrupados a proveedores.

    Endpoints principales:
    - POST /batch-payments/create_batch_payment/ - Crear pago múltiple con distribución FIFO
    - GET /batch-payments/ - Listar todos los pagos agrupados
    - GET /batch-payments/{id}/ - Detalle de un pago agrupado
    - DELETE /batch-payments/{id}/ - Eliminar pago agrupado (revierte pagos individuales)
    """
    queryset = BatchPayment.objects.select_related(
        'provider', 'bank', 'created_by'
    ).all()
    serializer_class = BatchPaymentSerializer
    permission_classes = [IsOperativo]
    parser_classes = (MultiPartParser, FormParser, JSONParser)

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return BatchPaymentDetailSerializer
        return BatchPaymentSerializer

    @action(detail=False, methods=['post'], permission_classes=[IsOperativo])
    def create_batch_payment(self, request):
        """
        Crea un pago agrupado con asignación manual de montos a cada factura.

        Body esperado:
        {
            "allocations": [
                {"transfer_id": 1, "amount": "150.00"},
                {"transfer_id": 2, "amount": "50.00"}
            ],
            "payment_method": "transferencia",
            "payment_date": "2025-12-21",
            "bank": 1,
            "reference_number": "TRANS-12345",
            "notes": "Pago quincenal",
            "proof_file": <archivo>
        }
        """
        from django.db import transaction
        from decimal import Decimal
        import json

        # Parsear allocations
        allocations_data = request.data.get('allocations', '[]')
        try:
            allocations = json.loads(allocations_data) if isinstance(allocations_data, str) else allocations_data
        except:
            return Response({'error': 'Formato inválido en allocations'}, status=status.HTTP_400_BAD_REQUEST)

        if not allocations or len(allocations) == 0:
            return Response({'error': 'Debe especificar los montos a pagar por factura'}, status=status.HTTP_400_BAD_REQUEST)

        payment_method = request.data.get('payment_method')
        payment_date = request.data.get('payment_date')

        if not payment_method or not payment_date:
            return Response({'error': 'Método de pago y fecha son requeridos'}, status=status.HTTP_400_BAD_REQUEST)

        # Validar y preparar datos
        valid_allocations = []
        total_payment_amount = Decimal('0.00')
        provider_id = None
        service_orders_affected = set()

        # Manejo robusto de bank_id (convertir '' a None)
        bank_id = request.data.get('bank')
        if bank_id == '' or bank_id == 'null' or bank_id == 'undefined':
            bank_id = None

        for item in allocations:
            transfer_id = item.get('transfer_id')
            amount_str = str(item.get('amount', 0))
            
            try:
                amount = Decimal(amount_str)
                if amount <= 0:
                    continue # Ignorar montos 0 o negativos
            except:
                return Response({'error': f'Monto inválido para factura ID {transfer_id}'}, status=status.HTTP_400_BAD_REQUEST)

            try:
                transfer = Transfer.objects.select_related('provider', 'service_order').get(id=transfer_id)
            except Transfer.DoesNotExist:
                return Response({'error': f'Factura ID {transfer_id} no encontrada'}, status=status.HTTP_400_BAD_REQUEST)

            # Validar proveedor único
            if provider_id is None:
                provider_id = transfer.provider_id
            elif provider_id != transfer.provider_id:
                return Response({'error': 'Todas las facturas deben ser del mismo proveedor'}, status=status.HTTP_400_BAD_REQUEST)

            # Validar saldo
            if amount > transfer.balance:
                return Response({
                    'error': f'El monto ${amount} excede el saldo pendiente (${transfer.balance}) de la factura {transfer.invoice_number or transfer.id}'
                }, status=status.HTTP_400_BAD_REQUEST)

            valid_allocations.append({
                'transfer': transfer,
                'amount': amount
            })
            total_payment_amount += amount
            if transfer.service_order:
                service_orders_affected.add(transfer.service_order)

        if total_payment_amount <= 0:
            return Response({'error': 'El monto total a pagar debe ser mayor a cero'}, status=status.HTTP_400_BAD_REQUEST)

        # Crear registros
        try:
            with transaction.atomic():
                # 1. Crear BatchPayment
                batch_payment = BatchPayment(
                    provider_id=provider_id,
                    total_amount=total_payment_amount,
                    payment_method=payment_method,
                    payment_date=payment_date,
                    bank_id=bank_id,
                    reference_number=request.data.get('reference_number', ''),
                    notes=request.data.get('notes', ''),
                    created_by=request.user
                )

                if 'proof_file' in request.FILES:
                    batch_payment.proof_file = request.FILES['proof_file']

                batch_payment.save()

                # 2. Crear TransferPayments individuales
                payments_created = []
                for alloc in valid_allocations:
                    transfer_payment = TransferPayment(
                        transfer=alloc['transfer'],
                        batch_payment=batch_payment,
                        amount=alloc['amount'],
                        payment_date=payment_date,
                        payment_method=payment_method,
                        reference_number=request.data.get('reference_number', ''),
                        notes=f"Pago agrupado {batch_payment.batch_number}",
                        created_by=request.user
                    )
                    
                    if batch_payment.proof_file:
                        transfer_payment.proof_file = batch_payment.proof_file
                        
                    transfer_payment.save()
                    payments_created.append(transfer_payment)

                # 3. Documentos en Service Orders: ya gestionado por signal sync_batch_payment_documents

                serializer = BatchPaymentDetailSerializer(batch_payment)
                return Response({
                    'message': f'Pago agrupado registrado exitosamente. Total: ${total_payment_amount}',
                    'batch_payment': serializer.data
                }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({'error': f'Error al procesar el pago: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def destroy(self, request, *args, **kwargs):
        """
        Elimina un pago agrupado y revierte todos los pagos individuales asociados.
        Usa soft delete para mantener historial.
        """
        batch_payment = self.get_object()
        service_orders_affected = list(batch_payment.get_service_orders())

        try:
            with transaction.atomic():
                # Los TransferPayment se eliminarán en cascada por el on_delete=CASCADE
                # Esto disparará el signal post_delete que recalculará los paid_amount de los Transfers
                batch_payment._current_user = request.user
                batch_payment.delete()  # Soft delete

                # Registrar auditoría explícita por cada OS impactada.
                from apps.orders.models import OrderHistory
                for service_order in service_orders_affected:
                    OrderHistory.objects.create(
                        service_order=service_order,
                        user=request.user,
                        event_type='payment_deleted',
                        description=f'Pago agrupado eliminado: {batch_payment.batch_number}',
                        metadata={
                            'source': 'batch_payment',
                            'batch_number': batch_payment.batch_number,
                            'batch_payment_id': batch_payment.id,
                            'provider': batch_payment.provider.name if batch_payment.provider else None,
                            'amount': float(batch_payment.total_amount),
                        }
                    )

                return Response(
                    {'message': 'Pago agrupado eliminado correctamente. Los pagos individuales fueron revertidos.'},
                    status=status.HTTP_204_NO_CONTENT
                )
        except Exception as e:
            return Response(
                {'error': f'Error al eliminar el pago agrupado: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ProviderCreditNoteFilter(filters.FilterSet):
    """Filtros para notas de crédito de proveedores"""
    issue_date_from = filters.DateFilter(field_name='issue_date', lookup_expr='gte')
    issue_date_to = filters.DateFilter(field_name='issue_date', lookup_expr='lte')
    min_amount = filters.NumberFilter(field_name='amount', lookup_expr='gte')
    max_amount = filters.NumberFilter(field_name='amount', lookup_expr='lte')
    has_available = filters.BooleanFilter(method='filter_has_available')

    class Meta:
        model = ProviderCreditNote
        fields = ['provider', 'status', 'reason']

    def filter_has_available(self, queryset, name, value):
        if value:
            return queryset.filter(available_amount__gt=0)
        return queryset


class ProviderCreditNoteViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar Notas de Crédito de Proveedores.

    Sistema ERP profesional para manejar notas de crédito que anulan
    parcial o totalmente facturas de proveedores.

    ENDPOINTS:
    - GET    /provider-credit-notes/              - Listar NC
    - POST   /provider-credit-notes/              - Crear NC
    - GET    /provider-credit-notes/{id}/         - Detalle de NC
    - PUT    /provider-credit-notes/{id}/         - Actualizar NC
    - DELETE /provider-credit-notes/{id}/         - Eliminar NC (solo si no tiene aplicaciones)
    - POST   /provider-credit-notes/{id}/apply/   - Aplicar NC a facturas
    - POST   /provider-credit-notes/{id}/void/    - Anular NC
    - GET    /provider-credit-notes/summary/      - Resumen estadístico
    - GET    /provider-credit-notes/by-provider/  - NC agrupadas por proveedor
    """
    queryset = ProviderCreditNote.objects.select_related(
        'provider', 'created_by', 'voided_by'
    ).all()
    permission_classes = [IsOperativo]
    parser_classes = (MultiPartParser, FormParser, JSONParser)
    filterset_class = ProviderCreditNoteFilter
    search_fields = ['note_number', 'provider__name', 'reason_detail']
    ordering_fields = ['issue_date', 'amount', 'available_amount', 'created_at']
    ordering = ['-issue_date']

    def get_serializer_class(self):
        if self.action == 'list':
            return ProviderCreditNoteListSerializer
        elif self.action == 'create':
            return ProviderCreditNoteCreateSerializer
        elif self.action == 'apply':
            return ApplyCreditNoteSerializer
        return ProviderCreditNoteDetailSerializer

    def destroy(self, request, *args, **kwargs):
        """
        Eliminar una NC solo si no tiene aplicaciones.
        """
        credit_note = self.get_object()

        # Validar que no tenga aplicaciones
        if credit_note.applications.filter(is_deleted=False).exists():
            return Response(
                {
                    'error': 'No se puede eliminar esta nota de crédito porque ya tiene aplicaciones.',
                    'detail': 'Primero debe anular la NC para revertir las aplicaciones.',
                    'code': 'HAS_APPLICATIONS'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar que no esté anulada
        if credit_note.status == 'anulada':
            return Response(
                {
                    'error': 'No se puede eliminar una nota de crédito anulada.',
                    'detail': 'Las NC anuladas se mantienen para auditoría.',
                    'code': 'ALREADY_VOIDED'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        credit_note._current_user = request.user
        credit_note.delete()
        return Response(
            {'message': 'Nota de crédito eliminada correctamente.'},
            status=status.HTTP_204_NO_CONTENT
        )

    @action(detail=True, methods=['post'], permission_classes=[IsOperativo])
    def apply(self, request, pk=None):
        """
        Aplicar una Nota de Crédito a una o más facturas del proveedor.

        Body esperado:
        {
            "applications": [
                {"transfer_id": 1, "amount": 500.00, "notes": "Aplicación parcial"},
                {"transfer_id": 2, "amount": 300.00, "notes": ""}
            ]
        }

        VALIDACIONES:
        - La NC debe tener saldo disponible
        - Los Transfers deben ser del mismo proveedor
        - El monto total no puede exceder el saldo disponible de la NC
        - Cada monto no puede exceder el saldo del Transfer
        """
        from decimal import Decimal

        credit_note = self.get_object()

        # Validar estado de la NC
        if not credit_note.can_apply():
            return Response(
                {
                    'error': 'Esta nota de crédito no puede aplicarse.',
                    'detail': f'Estado: {credit_note.get_status_display()}. Saldo disponible: ${credit_note.available_amount}',
                    'code': 'CANNOT_APPLY'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar datos de entrada
        serializer = ApplyCreditNoteSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        applications_data = serializer.validated_data['applications']

        # Calcular monto total a aplicar
        total_to_apply = Decimal('0.00')
        for app_data in applications_data:
            total_to_apply += Decimal(str(app_data['amount']))

        if total_to_apply > credit_note.available_amount:
            return Response(
                {
                    'error': 'El monto total excede el saldo disponible de la NC.',
                    'detail': f'Intentando aplicar ${total_to_apply}, pero solo hay ${credit_note.available_amount} disponibles.',
                    'code': 'EXCEEDS_AVAILABLE'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar cada Transfer y recopilar errores
        transfers_to_apply = []
        errors = []

        for i, app_data in enumerate(applications_data):
            transfer_id = app_data['transfer_id']
            amount = Decimal(str(app_data['amount']))

            try:
                transfer = Transfer.objects.get(id=transfer_id, is_deleted=False)
            except Transfer.DoesNotExist:
                errors.append(f"Factura #{transfer_id} no encontrada")
                continue

            # Validar mismo proveedor
            if transfer.provider_id != credit_note.provider_id:
                errors.append(f"Factura #{transfer_id} no pertenece al proveedor {credit_note.provider.name}")
                continue

            # Validar saldo disponible en el Transfer
            if amount > transfer.balance:
                errors.append(f"Factura #{transfer_id}: monto ${amount} excede saldo pendiente ${transfer.balance}")
                continue

            transfers_to_apply.append({
                'transfer': transfer,
                'amount': amount,
                'notes': app_data.get('notes', '')
            })

        if errors:
            return Response(
                {
                    'error': 'Errores en las aplicaciones.',
                    'detail': errors,
                    'code': 'VALIDATION_ERRORS'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # Aplicar todas las NC en una transacción
        try:
            with transaction.atomic():
                applications_created = []

                for app_data in transfers_to_apply:
                    application = CreditNoteApplication(
                        credit_note=credit_note,
                        transfer=app_data['transfer'],
                        amount=app_data['amount'],
                        applied_by=request.user,
                        notes=app_data['notes']
                    )
                    application.save()
                    applications_created.append({
                        'id': application.id,
                        'transfer_id': app_data['transfer'].id,
                        'transfer_invoice': app_data['transfer'].invoice_number,
                        'amount': str(app_data['amount'])
                    })

                # Refrescar la NC para obtener valores actualizados
                credit_note.refresh_from_db()

                return Response({
                    'message': f'Nota de crédito aplicada exitosamente a {len(applications_created)} factura(s).',
                    'credit_note': {
                        'id': credit_note.id,
                        'note_number': credit_note.note_number,
                        'status': credit_note.status,
                        'status_display': credit_note.get_status_display(),
                        'applied_amount': str(credit_note.applied_amount),
                        'available_amount': str(credit_note.available_amount)
                    },
                    'applications': applications_created
                }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {
                    'error': 'Error al aplicar la nota de crédito.',
                    'detail': str(e),
                    'code': 'APPLICATION_ERROR'
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'], permission_classes=[IsOperativo2OrAdmin])
    def void(self, request, pk=None):
        """
        Anular una Nota de Crédito y revertir todas sus aplicaciones.

        Body esperado:
        {
            "reason": "Motivo de la anulación"
        }

        EFECTOS:
        - Revierte todos los TransferPayments creados por la NC
        - Restaura los saldos de los Transfers afectados
        - Marca la NC como anulada con registro de auditoría

        RESTRICCIONES:
        - Solo usuarios con rol operativo2 o admin pueden anular
        - No se puede anular una NC ya anulada
        """
        credit_note = self.get_object()

        # Validar que no esté ya anulada
        if credit_note.status == 'anulada':
            return Response(
                {
                    'error': 'Esta nota de crédito ya está anulada.',
                    'detail': f'Anulada el {credit_note.voided_at} por {credit_note.voided_by.username if credit_note.voided_by else "N/A"}',
                    'code': 'ALREADY_VOIDED'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # Obtener motivo
        reason = request.data.get('reason', '').strip()
        if not reason:
            return Response(
                {
                    'error': 'Debe proporcionar un motivo de anulación.',
                    'code': 'REASON_REQUIRED'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            with transaction.atomic():
                # Obtener Transfers afectados antes de anular
                affected_transfers = list(credit_note.get_affected_transfers().values_list('id', flat=True))

                # Anular la NC
                credit_note.void(user=request.user, reason=reason)

                return Response({
                    'message': 'Nota de crédito anulada correctamente. Las aplicaciones han sido revertidas.',
                    'credit_note': {
                        'id': credit_note.id,
                        'note_number': credit_note.note_number,
                        'status': credit_note.status,
                        'status_display': credit_note.get_status_display(),
                        'voided_at': credit_note.voided_at,
                        'void_reason': credit_note.void_reason
                    },
                    'affected_transfers': affected_transfers
                }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {
                    'error': 'Error al anular la nota de crédito.',
                    'detail': str(e),
                    'code': 'VOID_ERROR'
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], permission_classes=[IsOperativo])
    def summary(self, request):
        """
        Resumen estadístico de notas de crédito.

        Retorna:
        - Total de NC registradas
        - Monto total en NC
        - Monto total aplicado
        - Saldo disponible total
        - Desglose por estado
        - Desglose por proveedor (top 10)
        """
        queryset = self.filter_queryset(self.get_queryset())

        # Totales generales
        totals = queryset.aggregate(
            total_amount=Sum('amount'),
            total_applied=Sum('applied_amount'),
            total_available=Sum('available_amount')
        )

        # Por estado
        by_status = {}
        for status_choice in ProviderCreditNote.STATUS_CHOICES:
            status_code = status_choice[0]
            status_data = queryset.filter(status=status_code)
            by_status[status_code] = {
                'label': status_choice[1],
                'count': status_data.count(),
                'amount': status_data.aggregate(total=Sum('amount'))['total'] or 0
            }

        # Por proveedor (top 10)
        from django.db.models import Count
        by_provider = queryset.values(
            'provider', 'provider__name'
        ).annotate(
            count=Count('id'),
            total_amount=Sum('amount'),
            total_available=Sum('available_amount')
        ).order_by('-total_amount')[:10]

        return Response({
            'total_count': queryset.count(),
            'total_amount': totals['total_amount'] or 0,
            'total_applied': totals['total_applied'] or 0,
            'total_available': totals['total_available'] or 0,
            'by_status': by_status,
            'by_provider': list(by_provider)
        })

    @action(detail=False, methods=['get'], permission_classes=[IsOperativo])
    def by_provider(self, request):
        """
        Lista NC agrupadas por proveedor.

        Query params opcionales:
        - provider_id: Filtrar por proveedor específico
        - only_available: Solo NC con saldo disponible
        """
        provider_id = request.query_params.get('provider_id')
        only_available = request.query_params.get('only_available', 'false').lower() == 'true'

        queryset = self.filter_queryset(self.get_queryset())

        if provider_id:
            queryset = queryset.filter(provider_id=provider_id)

        if only_available:
            queryset = queryset.filter(available_amount__gt=0, status__in=['pendiente', 'parcial'])

        # Agrupar por proveedor
        from django.db.models import Count
        providers_data = queryset.values(
            'provider', 'provider__name'
        ).annotate(
            count=Count('id'),
            total_amount=Sum('amount'),
            total_available=Sum('available_amount')
        ).order_by('provider__name')

        # Obtener NC detalladas por proveedor
        result = []
        for provider_data in providers_data:
            provider_notes = queryset.filter(provider_id=provider_data['provider'])
            notes_list = ProviderCreditNoteListSerializer(provider_notes, many=True).data

            result.append({
                'provider_id': provider_data['provider'],
                'provider_name': provider_data['provider__name'],
                'count': provider_data['count'],
                'total_amount': provider_data['total_amount'],
                'total_available': provider_data['total_available'],
                'credit_notes': notes_list
            })

        return Response(result)

    @action(detail=False, methods=['get'], permission_classes=[IsOperativo])
    def pending_for_provider(self, request):
        """
        Obtener facturas pendientes de un proveedor para aplicar NC.

        Query params requeridos:
        - provider_id: ID del proveedor

        Retorna facturas (Transfers) del proveedor con saldo pendiente.
        """
        provider_id = request.query_params.get('provider_id')

        if not provider_id:
            return Response(
                {'error': 'Se requiere el parámetro provider_id'},
                status=status.HTTP_400_BAD_REQUEST
            )

        transfers = Transfer.objects.filter(
            provider_id=provider_id,
            balance__gt=0,
            is_deleted=False
        ).exclude(
            status='pagado'
        ).select_related('service_order').order_by('-transaction_date')

        data = [{
            'id': t.id,
            'description': t.description,
            'invoice_number': t.invoice_number,
            'amount': str(t.amount),
            'balance': str(t.balance),
            'paid_amount': str(t.paid_amount),
            'service_order_number': t.service_order.order_number if t.service_order else None,
            'transaction_date': t.transaction_date,
            'status': t.status,
            'status_display': t.get_status_display()
        } for t in transfers]

        return Response({
            'provider_id': provider_id,
            'pending_transfers': data,
            'total_count': len(data),
            'total_balance': sum(t.balance for t in transfers)
        })


class ProviderInvoicePaymentViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar pagos individuales a facturas de proveedor (costos directos).
    Permite eliminar pagos individuales.
    """
    queryset = ProviderInvoicePayment.objects.select_related(
        'provider_invoice', 'created_by'
    ).all()
    permission_classes = [IsOperativo]
    http_method_names = ['get', 'delete']

    def destroy(self, request, *args, **kwargs):
        payment = self.get_object()
        payment._current_user = request.user
        payment.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class TransferPaymentViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar pagos individuales a transferencias/gastos.

    Permite eliminar pagos individuales (no agrupados) realizados a proveedores.
    Los pagos que son parte de un pago agrupado (batch_payment) deben eliminarse
    a través del BatchPaymentViewSet.
    """
    queryset = TransferPayment.objects.select_related(
        'transfer', 'created_by', 'batch_payment'
    ).all()
    serializer_class = TransferPaymentSerializer
    permission_classes = [IsOperativo]
    http_method_names = ['get', 'delete']  # Solo lectura y eliminación

    def destroy(self, request, *args, **kwargs):
        """
        Elimina un pago individual.
        No permite eliminar pagos que son parte de un pago agrupado.
        """
        payment = self.get_object()

        # Validar que no sea parte de un pago agrupado
        if payment.batch_payment:
            return Response(
                {
                    'error': f'Este pago es parte del pago agrupado #{payment.batch_payment.batch_number}. Para eliminarlo, debe eliminar el pago agrupado completo.',
                    'code': 'BATCH_PAYMENT_PROTECTED'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        # Eliminar el pago (el modelo se encarga de actualizar paid_amount del Transfer)
        payment._current_user = request.user
        payment.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ============================================
# ViewSets para COSTOS DIRECTOS (ProviderInvoice)
# ============================================

from .models import ProviderInvoice, DirectCostAllocation
from .serializers import (
    ProviderInvoiceListSerializer, ProviderInvoiceDetailSerializer,
    ProviderInvoiceCreateSerializer, DirectCostAllocationSerializer,
    DirectCostAllocationCreateSerializer
)


class ProviderInvoiceFilter(filters.FilterSet):
    """Filtros para facturas de proveedor"""
    date_from = filters.DateFilter(field_name='issue_date', lookup_expr='gte')
    date_to = filters.DateFilter(field_name='issue_date', lookup_expr='lte')

    class Meta:
        model = ProviderInvoice
        fields = ['service_order', 'provider', 'status', 'payment_status']


class ProviderInvoiceViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar Facturas de Proveedor (Costos Directos).

    Estas facturas representan costos de servicios tercerizados que se
    revenderán al cliente con margen de ganancia.

    Acciones adicionales:
    - allocate_cost: Asignar una porción del costo a un servicio
    - get_available_charges: Obtener servicios disponibles para asignar
    """
    queryset = ProviderInvoice.objects.select_related(
        'provider', 'service_order', 'created_by'
    ).prefetch_related('allocations').all()
    permission_classes = [IsOperativo2OrAdmin]
    parser_classes = (MultiPartParser, FormParser, JSONParser)
    filterset_class = ProviderInvoiceFilter
    search_fields = ['invoice_number', 'provider__name', 'service_order__order_number']
    ordering_fields = ['issue_date', 'total_amount', 'created_at']
    ordering = ['-created_at']
    pagination_class = None

    def get_serializer_class(self):
        if self.action == 'list':
            return ProviderInvoiceListSerializer
        elif self.action == 'create':
            return ProviderInvoiceCreateSerializer
        return ProviderInvoiceDetailSerializer

    def destroy(self, request, *args, **kwargs):
        """
        Eliminar una factura de proveedor con manejo de errores de validación.
        """
        instance = self.get_object()
        try:
            instance._current_user = request.user
            instance.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except ValidationError as e:
            error_message = e.messages[0] if hasattr(e, 'messages') and e.messages else str(e)
            return Response(
                {'error': error_message, 'code': 'VALIDATION_ERROR'},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['post'])
    def allocate_cost(self, request, pk=None):
        """
        Asignar una porción del costo de la factura a un servicio.

        Body:
        {
            "order_charge_id": 123,
            "cost_amount": 250.00,
            "description": "Cuadrilla portuaria"
        }
        """
        from apps.orders.models import OrderCharge
        from decimal import Decimal

        provider_invoice = self.get_object()

        # Validar datos
        order_charge_id = request.data.get('order_charge_id')
        cost_amount = request.data.get('cost_amount')
        description = request.data.get('description', '')

        if not order_charge_id or not cost_amount:
            return Response(
                {'error': 'order_charge_id y cost_amount son requeridos'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            cost_amount = Decimal(str(cost_amount))
            if cost_amount <= 0:
                return Response(
                    {'error': 'El monto debe ser mayor a cero'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except:
            return Response(
                {'error': 'Monto inválido'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar disponibilidad
        available = provider_invoice.total_amount - provider_invoice.allocated_amount
        if cost_amount > available:
            return Response(
                {'error': f'El monto excede el disponible (${available})'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Obtener el OrderCharge
        try:
            order_charge = OrderCharge.objects.get(
                id=order_charge_id,
                service_order=provider_invoice.service_order,
                is_deleted=False
            )
        except OrderCharge.DoesNotExist:
            return Response(
                {'error': 'Servicio no encontrado o no pertenece a esta OS'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar que no tenga ya una asignación
        if hasattr(order_charge, 'cost_allocation') and order_charge.cost_allocation:
            return Response(
                {'error': 'Este servicio ya tiene un costo directo asignado'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Advertencia si el precio de venta es menor al costo (margen negativo)
        # Pero permitir la operación - el usuario puede tener sus razones
        warning_message = None
        if order_charge.subtotal < cost_amount:
            profit = order_charge.subtotal - cost_amount
            margin = ((profit / cost_amount) * 100) if cost_amount > 0 else 0
            warning_message = (
                f'Margen negativo: el precio de venta (${order_charge.subtotal}) '
                f'es menor al costo (${cost_amount}). '
                f'Pérdida: ${abs(profit)} ({margin:.1f}%). '
                f'Revise si el margen de ganancia es correcto.'
            )

        # Crear la asignación
        with transaction.atomic():
            allocation = DirectCostAllocation.objects.create(
                provider_invoice=provider_invoice,
                order_charge=order_charge,
                cost_amount=cost_amount,
                description=description,
                created_by=request.user
            )

        response_data = {
            'message': 'Costo asignado exitosamente',
            'allocation': DirectCostAllocationSerializer(allocation).data,
            'provider_invoice': ProviderInvoiceDetailSerializer(provider_invoice).data
        }
        
        if warning_message:
            response_data['warning'] = warning_message

        return Response(response_data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['get'])
    def available_charges(self, request, pk=None):
        """
        Obtener los servicios de la OS que no tienen costo asignado.
        """
        from apps.orders.models import OrderCharge

        provider_invoice = self.get_object()

        # Servicios de la misma OS sin asignación de costo
        charges = OrderCharge.objects.filter(
            service_order=provider_invoice.service_order,
            is_deleted=False,
            is_third_party_service=False  # Solo los que no son tercerizados
        ).select_related('service')

        result = [{
            'id': c.id,
            'service_name': c.service.name,
            'description': c.description,
            'subtotal': float(c.subtotal),
            'total': float(c.total),
            'billing_status': c.billing_status
        } for c in charges]

        return Response({
            'available_charges': result,
            'unallocated_amount': float(provider_invoice.unallocated_amount)
        })

    @action(detail=False, methods=['get'])
    def by_service_order(self, request):
        """Obtener todas las facturas de proveedor de una OS"""
        service_order_id = request.query_params.get('service_order')
        if not service_order_id:
            return Response(
                {'error': 'service_order es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )

        invoices = self.queryset.filter(
            service_order_id=service_order_id,
            is_deleted=False
        )

        serializer = ProviderInvoiceListSerializer(invoices, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'], permission_classes=[IsOperativo])
    def register_payment(self, request, pk=None):
        """Registrar un pago parcial o total a una factura de proveedor (costo directo)"""
        from decimal import Decimal
        
        provider_invoice = self.get_object()
        
        # Calcular saldo pendiente
        balance = provider_invoice.total_amount - provider_invoice.paid_amount
        
        # Validar que no esté completamente pagada
        if provider_invoice.payment_status == 'pagado' and balance <= 0:
            return Response(
                {'error': 'Esta factura ya está completamente pagada'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar datos requeridos
        amount = request.data.get('amount')
        if not amount:
            return Response(
                {'error': 'El monto es requerido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            amount = Decimal(str(amount))
            if amount <= 0:
                return Response(
                    {'error': 'El monto debe ser mayor a cero'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validar que el monto no exceda el saldo pendiente
            if amount > balance:
                return Response(
                    {'error': f'El monto excede el saldo pendiente de ${balance}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except (ValueError, TypeError):
            return Response(
                {'error': 'Monto inválido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Crear registro de pago individual
        payment = ProviderInvoicePayment(
            provider_invoice=provider_invoice,
            amount=amount,
            payment_date=request.data.get('payment_date'),
            payment_method=request.data.get('payment_method', 'transferencia'),
            reference_number=request.data.get('reference', ''),
            notes=request.data.get('notes', ''),
            created_by=request.user
        )
        
        # Guardar archivo de comprobante
        if 'proof_file' in request.FILES:
            payment.proof_file = request.FILES['proof_file']
        
        payment.save()  # El modelo actualiza paid_amount automáticamente
        
        # Refrescar para obtener el status actualizado
        provider_invoice.refresh_from_db()
        
        return Response({
            'message': 'Pago registrado exitosamente',
            'payment': {
                'id': payment.id,
                'amount': str(payment.amount),
                'payment_date': payment.payment_date,
                'payment_method': payment.payment_method,
                'reference_number': payment.reference_number,
                'proof_file': payment.proof_file.url if payment.proof_file else None
            },
            'provider_invoice': {
                'id': provider_invoice.id,
                'paid_amount': str(provider_invoice.paid_amount),
                'balance': str(provider_invoice.total_amount - provider_invoice.paid_amount),
                'payment_status': provider_invoice.payment_status,
                'status_display': dict(ProviderInvoice.PAYMENT_STATUS_CHOICES).get(
                    provider_invoice.payment_status, provider_invoice.payment_status
                )
            }
        }, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['get'])
    def detail_with_payments(self, request, pk=None):
        """Obtener detalle de la factura de proveedor (formato compatible con Transfer)"""
        provider_invoice = self.get_object()
        balance = float(provider_invoice.total_amount) - float(provider_invoice.paid_amount)
        
        # Obtener historial de pagos
        payments = ProviderInvoicePayment.objects.filter(
            provider_invoice=provider_invoice,
            is_deleted=False
        ).order_by('-payment_date')
        
        payments_data = [{
            'id': p.id,
            'amount': str(p.amount),
            'payment_date': p.payment_date,
            'payment_method': p.payment_method,
            'payment_method_display': p.get_payment_method_display(),
            'reference_number': p.reference_number,
            'notes': p.notes,
            'proof_file': p.proof_file.url if p.proof_file else None,
            'created_by': p.created_by.username if p.created_by else None,
            'created_at': p.created_at
        } for p in payments]
        
        return Response({
            'id': provider_invoice.id,
            'source': 'provider_invoice',
            'invoice_number': provider_invoice.invoice_number,
            'provider': {
                'id': provider_invoice.provider.id,
                'name': provider_invoice.provider.name
            },
            'service_order': {
                'id': provider_invoice.service_order.id,
                'order_number': provider_invoice.service_order.order_number,
                'purchase_order': provider_invoice.service_order.purchase_order
            } if provider_invoice.service_order else None,
            'amount': float(provider_invoice.total_amount),
            'paid_amount': float(provider_invoice.paid_amount),
            'balance': balance,
            'status': provider_invoice.payment_status,
            'status_display': dict(ProviderInvoice.PAYMENT_STATUS_CHOICES).get(
                provider_invoice.payment_status, provider_invoice.payment_status
            ),
            'description': provider_invoice.notes or f'Factura {provider_invoice.invoice_number}',
            'transaction_date': provider_invoice.issue_date,
            'invoice_file': provider_invoice.invoice_file.url if provider_invoice.invoice_file else None,
            'generation_code': provider_invoice.generation_code,
            'reception_stamp': provider_invoice.reception_stamp,
            'payments': payments_data,
            'credit_notes': [],
            'total_payments_count': len(payments_data),
        })


class DirectCostAllocationViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar Asignaciones de Costo Directo.

    Permite crear, editar y eliminar asignaciones de costo a servicios.
    """
    queryset = DirectCostAllocation.objects.select_related(
        'provider_invoice', 'provider_invoice__provider',
        'order_charge', 'order_charge__service',
        'created_by'
    ).all()
    permission_classes = [IsOperativo2OrAdmin]
    search_fields = ['description', 'provider_invoice__invoice_number']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'create':
            return DirectCostAllocationCreateSerializer
        return DirectCostAllocationSerializer

    def destroy(self, request, *args, **kwargs):
        """
        Eliminar una asignación de costo con manejo de errores de validación.
        No permite eliminar si el servicio ya fue facturado al cliente.
        """
        allocation = self.get_object()

        try:
            allocation._current_user = request.user
            allocation.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except ValidationError as e:
            error_message = e.messages[0] if hasattr(e, 'messages') and e.messages else str(e)
            return Response(
                {'error': error_message, 'code': 'VALIDATION_ERROR'},
                status=status.HTTP_400_BAD_REQUEST
            )