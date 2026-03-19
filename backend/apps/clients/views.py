from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Sum, Q, F, ExpressionWrapper, DecimalField
from django.http import HttpResponse
from .models import Client
from .serializers import ClientSerializer, ClientListSerializer
from apps.users.permissions import IsOperativo, IsOperativo2, IsAdminUser
from apps.orders.models import ServiceOrder
from apps.transfers.models import Transfer
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta

class ClientViewSet(viewsets.ModelViewSet):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer
    permission_classes = [IsOperativo]
    search_fields = ['name', 'nit', 'email']
    filterset_fields = ['payment_condition', 'is_active']

    def get_queryset(self):
        """
        Sobreescribir queryset para anotar el saldo pendiente real (CXC)
        basado en facturas con balance > 0.
        """
        queryset = super().get_queryset()
        
        if self.action in ['list', 'retrieve']:
            from apps.orders.models import Invoice
            from django.db.models import OuterRef, Subquery, Sum
            
            # Obtener año opcional
            year = self.request.query_params.get('year')

            # Base de facturas pendientes
            pending_invoices = Invoice.objects.filter(
                service_order__client=OuterRef('pk'),
                balance__gt=0.01
            ).exclude(
                status__in=['paid', 'cancelled']
            )

            # Aplicar filtro de año si existe (lógica estricta)
            if year and year != '0':
                pending_invoices = pending_invoices.filter(issue_date__year=year)

            # Anotamos el saldo pendiente real sumando el balance
            pending_balance_sq = pending_invoices.values(
                'service_order__client'
            ).annotate(
                total=Sum('balance')
            ).values('total')
            
            queryset = queryset.annotate(
                annotated_total_pending=Subquery(pending_balance_sq)
            )
            
        return queryset

    def get_serializer_class(self):
        if self.action == 'list':
            return ClientListSerializer
        return ClientSerializer

    def update(self, request, *args, **kwargs):
        """Validar antes de actualizar un cliente"""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()

        # Si se intenta desactivar el cliente
        if 'is_active' in request.data and not request.data['is_active'] and instance.is_active:
            # Verificar si tiene órdenes de servicio no cerradas
            active_orders = ServiceOrder.objects.filter(
                client=instance
            ).exclude(status='cerrada').count()

            if active_orders > 0:
                return Response(
                    {'error': f'No se puede desactivar el cliente porque tiene {active_orders} orden(es) de servicio activa(s). Cierre todas las órdenes antes de desactivar el cliente.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Verificar si tiene facturas pendientes
            from apps.orders.models import Invoice
            pending_invoices = Invoice.objects.filter(
                service_order__client=instance,
                balance__gt=0
            ).exclude(status__in=['paid', 'cancelled']).count()

            if pending_invoices > 0:
                return Response(
                    {'error': f'No se puede desactivar el cliente porque tiene {pending_invoices} factura(s) con saldo pendiente. Liquide todas las facturas antes de desactivar el cliente.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)

        return Response(serializer.data)

    def destroy(self, request, *args, **kwargs):
        """Validar antes de eliminar un cliente"""
        instance = self.get_object()

        # Verificar si tiene órdenes de servicio asociadas
        orders_count = ServiceOrder.objects.filter(client=instance).count()
        if orders_count > 0:
            return Response(
                {'error': f'No se puede eliminar el cliente porque tiene {orders_count} orden(es) de servicio asociada(s). Los clientes con historial no pueden ser eliminados, solo desactivados.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verificar si tiene facturas asociadas
        from apps.orders.models import Invoice
        invoices_count = Invoice.objects.filter(service_order__client=instance).count()
        if invoices_count > 0:
            return Response(
                {'error': f'No se puede eliminar el cliente porque tiene {invoices_count} factura(s) asociada(s). Los clientes con historial de facturación no pueden ser eliminados.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verificar si tiene transfers asociados
        transfers_count = Transfer.objects.filter(client=instance).count()
        if transfers_count > 0:
            return Response(
                {'error': f'No se puede eliminar el cliente porque tiene {transfers_count} transferencia(s) asociada(s).'},
                status=status.HTTP_400_BAD_REQUEST
            )

        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['get'], permission_classes=[IsOperativo])
    def general_summary(self, request):
        """
        Resumen global financiero y de clientes para el dashboard de Estados de Cuenta.
        Calcula totales directamente de la base de datos para máxima precisión.
        """
        from apps.orders.models import Invoice
        
        # Filtro de año opcional
        year = request.query_params.get('year')
        current_year = datetime.now().year
        
        # Base de facturas válidas (no anuladas)
        invoices_qs = Invoice.objects.exclude(status='cancelled')
        
        if year:
            try:
                year_int = int(year)
                # Filtro estricto por año de emisión para reportes históricos precisos
                invoices_qs = invoices_qs.filter(issue_date__year=year_int)
            except ValueError:
                pass

        # === 1. Métricas Financieras Globales ===
        collected_expression = ExpressionWrapper(
            F('total_amount') - F('balance'),
            output_field=DecimalField(max_digits=15, decimal_places=2)
        )

        financial_stats = invoices_qs.aggregate(
            total_invoiced=Sum('total_amount'),
            total_services=Sum('total_services'),
            total_third_party=Sum('total_third_party'),
            total_paid=Sum('paid_amount'),
            total_balance=Sum('balance'),
            total_credited=Sum('credited_amount'),
            total_retention=Sum('retencion'),
            total_collected=Sum(collected_expression)
        )

        # === 2. Métricas de Clientes ===
        # Clientes totales
        total_clients = Client.objects.count()
        
        # Clientes con crédito
        credit_clients = Client.objects.filter(payment_condition='credito').count()
        
        # Para calcular clientes con saldo vs al día, necesitamos ver quiénes tienen facturas pendientes
        # Esto es más eficiente hacerlo con una query de agregación de facturas agrupada por cliente
        clients_with_balance_ids = Invoice.objects.filter(
            balance__gt=0.01
        ).exclude(
            status__in=['paid', 'cancelled']
        ).values_list('service_order__client_id', flat=True).distinct()
        
        pending_clients_count = clients_with_balance_ids.count()
        up_to_date_clients_count = total_clients - pending_clients_count

        data = {
            'financial': {
                'total_invoiced': float(financial_stats['total_invoiced'] or 0),
                'total_services': float(financial_stats['total_services'] or 0),
                'total_third_party': float(financial_stats['total_third_party'] or 0),
                'total_paid': float(financial_stats['total_paid'] or 0),
                'total_pending': float(financial_stats['total_balance'] or 0),
                'total_credited': float(financial_stats['total_credited'] or 0),
                'total_retention': float(financial_stats['total_retention'] or 0),
                'total_collected': float(financial_stats['total_collected'] or 0),
            },
            'clients': {
                'total': total_clients,
                'credit': credit_clients,
                'up_to_date': up_to_date_clients_count,
                'pending': pending_clients_count
            }
        }
        
        return Response(data)

    @action(detail=False, methods=['get'], permission_classes=[IsOperativo])
    def active(self, request):
        """Obtener solo clientes activos - útil para selectores en formularios"""
        active_clients = Client.objects.filter(is_active=True).order_by('name')
        serializer = ClientListSerializer(active_clients, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'], permission_classes=[IsOperativo])
    def account_statement(self, request, pk=None):
        """Estado de cuenta detallado de un cliente con facturas y pagos"""
        from apps.orders.models import Invoice, InvoicePayment
        from apps.orders.serializers import InvoiceListSerializer

        client = self.get_object()

        # Get year filter
        year = request.query_params.get('year', datetime.now().year)

        # Calculate Credit Used: Sum of balances from unpaid invoices
        unpaid_invoices = Invoice.objects.filter(
            service_order__client=client,
            balance__gt=0
        ).exclude(status__in=['paid', 'cancelled'])

        credit_used = unpaid_invoices.aggregate(Sum('balance'))['balance__sum'] or 0
        available_credit = max(0, float(client.credit_limit) - float(credit_used))

        # Get all invoices for the client (optionally filtered by year)
        invoices_qs = Invoice.objects.filter(
            service_order__client=client
        ).select_related('service_order').prefetch_related('payments')

        if year:
            # Mostrar facturas del año seleccionado O facturas con saldo pendiente de cualquier año
            invoices_qs = invoices_qs.filter(
                Q(issue_date__year=year) | Q(balance__gt=0)
            ).distinct()

        # Serializar facturas con información completa
        invoices_data = InvoiceListSerializer(invoices_qs, many=True).data

        # Facturas válidas para métricas financieras (excluye anuladas)
        valid_invoices_qs = invoices_qs.exclude(status='cancelled')

        collected_expression = ExpressionWrapper(
            F('total_amount') - F('balance'),
            output_field=DecimalField(max_digits=15, decimal_places=2)
        )

        # Calcular estadísticas de facturas
        total_invoiced = valid_invoices_qs.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        total_paid = valid_invoices_qs.aggregate(Sum('paid_amount'))['paid_amount__sum'] or 0
        total_collected = valid_invoices_qs.aggregate(
            total_collected=Sum(collected_expression)
        )['total_collected'] or 0
        total_pending = valid_invoices_qs.filter(
            status__in=['pending', 'partial', 'overdue']
        ).aggregate(Sum('balance'))['balance__sum'] or 0

        # Facturas por estado
        invoices_by_status = {
            'pending': invoices_qs.filter(status='pending').count(),
            'partial': invoices_qs.filter(status='partial').count(),
            'paid': invoices_qs.filter(status='paid').count(),
            'overdue': invoices_qs.filter(status='overdue').count(),
            'cancelled': invoices_qs.filter(status='cancelled').count(),
        }

        # Facturas vencidas y próximas a vencer
        today = datetime.now().date()
        overdue_invoices = valid_invoices_qs.filter(
            due_date__lt=today,
            balance__gt=0
        ).exclude(status='paid').count()

        upcoming_due = valid_invoices_qs.filter(
            due_date__gte=today,
            due_date__lte=today + timedelta(days=7),
            balance__gt=0
        ).exclude(status='paid').count()

        # Aging Analysis (Antigüedad de Saldos)
        aging = {
            'current': 0.0,  # Corriente (No vencido)
            '1-30': 0.0,     # 1 a 30 días de vencido
            '31-60': 0.0,    # 31 a 60 días
            '61-90': 0.0,    # 61 a 90 días
            '90+': 0.0       # Más de 90 días
        }

        # Iterar sobre facturas con saldo pendiente para calcular aging
        # Usamos unpaid_invoices que ya filtramos arriba (balance > 0, no pagadas/anuladas)
        for inv in unpaid_invoices:
            bal = float(inv.balance)
            if not inv.due_date:
                aging['current'] += bal
                continue

            days_overdue = (today - inv.due_date).days

            if days_overdue <= 0:
                aging['current'] += bal
            elif days_overdue <= 30:
                aging['1-30'] += bal
            elif days_overdue <= 60:
                aging['31-60'] += bal
            elif days_overdue <= 90:
                aging['61-90'] += bal
            else:
                aging['90+'] += bal

        # Historial de pagos recientes (últimos 10)
        recent_payments = InvoicePayment.objects.filter(
            invoice__service_order__client=client
        ).select_related('invoice').order_by('-payment_date')[:10]

        payments_data = [{
            'id': payment.id,
            'invoice_number': payment.invoice.invoice_number,
            'amount': float(payment.amount),
            'payment_date': payment.payment_date,
            'payment_method': payment.get_payment_method_display(),
            'reference': payment.reference_number or '',
            'notes': payment.notes or ''
        } for payment in recent_payments]

        # Órdenes de Servicio pendientes de facturar
        pending_orders_qs = ServiceOrder.objects.filter(
            client=client,
            facturado=False
        ).exclude(status='cancelada').order_by('created_at')

        pending_orders_data = [{
            'id': order.id,
            'order_number': order.order_number,
            'purchase_order': order.purchase_order,
            'date': order.created_at.date(),
            'eta': order.eta,
            'duca': order.duca,
            'amount': float(order.get_total_amount())
        } for order in pending_orders_qs]

        data = {
            'client_id': client.id,
            'client': client.name,
            'nit': client.nit,
            'payment_condition': client.get_payment_condition_display(),
            'credit_days': client.credit_days,
            'credit_limit': float(client.credit_limit),
            'credit_used': float(credit_used),
            'available_credit': float(available_credit),
            'credit_percentage': round((float(credit_used) / float(client.credit_limit) * 100), 2) if client.credit_limit > 0 else 0,

            # Estadísticas de facturación
            'total_invoiced': float(total_invoiced),
            'total_paid': float(total_paid),
            'total_collected': float(total_collected),
            'total_pending': float(total_pending),
            'total_pending_orders': pending_orders_qs.count(),

            # Análisis de Antigüedad
            'aging': aging,

            # Facturas por estado
            'invoices_by_status': invoices_by_status,
            'overdue_count': overdue_invoices,
            'upcoming_due_count': upcoming_due,

            # Listas detalladas
            'invoices': invoices_data,
            'pending_invoices': pending_orders_data,
            'recent_payments': payments_data,

            'year': year
        }

        return Response(data)
    
    @action(detail=True, methods=['get'], permission_classes=[IsOperativo])
    def export_statement_excel(self, request, pk=None):
        """Exportar estado de cuenta completo a Excel con facturas y pagos"""
        from apps.orders.models import Invoice, InvoicePayment

        client = self.get_object()
        year = request.query_params.get('year', datetime.now().year)
        
        # Debug: verificar parámetros
        print(f"DEBUG: Cliente: {client.name} (ID: {client.id})")
        print(f"DEBUG: Año filtrado: {year}")
        
        # Filtros adicionales
        status_filter = request.query_params.get('status')
        date_from = request.query_params.get('dateFrom')
        date_to = request.query_params.get('dateTo')
        min_amount = request.query_params.get('minAmount')
        max_amount = request.query_params.get('maxAmount')
        invoice_type = request.query_params.get('invoiceType')
        search_query = request.query_params.get('search')

        # Obtener facturas del cliente
        invoices = Invoice.objects.filter(
            service_order__client=client
        ).select_related('service_order').prefetch_related('payments').order_by('-issue_date')

        if year:
            # Mostrar facturas del año seleccionado O facturas con saldo pendiente de cualquier año
            invoices = invoices.filter(
                Q(issue_date__year=year) | Q(balance__gt=0)
            ).distinct()
            
        if status_filter:
            invoices = invoices.filter(status=status_filter)
        
        if date_from:
            invoices = invoices.filter(issue_date__gte=date_from)
            
        if date_to:
            invoices = invoices.filter(issue_date__lte=date_to)

        if min_amount:
            invoices = invoices.filter(total_amount__gte=min_amount)

        if max_amount:
            invoices = invoices.filter(total_amount__lte=max_amount)

        if invoice_type:
            invoices = invoices.filter(invoice_type=invoice_type)

        if search_query:
            invoices = invoices.filter(
                Q(invoice_number__icontains=search_query) |
                Q(service_order__order_number__icontains=search_query)
            )

        # Debug: verificar cantidad de facturas
        print(f"DEBUG: Total de facturas encontradas: {invoices.count()}")
        if invoices.exists():
            print(f"DEBUG: Primera factura: {invoices.first().invoice_number}")
            print(f"DEBUG: Total factura: {invoices.first().total_amount}")
            print(f"DEBUG: Total servicios: {invoices.first().total_services}")
            print(f"DEBUG: Total terceros: {invoices.first().total_third_party}")

        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estado de Cuenta"

        # Estilos (usando un azul más oscuro - #0F2E4D)
        header_fill = PatternFill(start_color="0F2E4D", end_color="0F2E4D", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        title_font = Font(size=16, bold=True, color="0F2E4D")
        subtitle_font = Font(size=12, bold=True, color="1A4C7A")
        label_font = Font(bold=True, color="404040", size=10)
        currency_format = '"$"#,##0.00'
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # === ENCABEZADO ===
        ws['A1'] = "ESTADO DE CUENTA"
        ws['A1'].font = title_font
        ws.merge_cells('A1:L1')

        ws['A2'] = "GPRO LOGISTIC - Agencia Aduanal"
        ws['A2'].font = Font(size=11, color="666666")
        ws.merge_cells('A2:L2')

        ws['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A3'].font = Font(size=9, italic=True, color="999999")

        # === INFORMACIÓN DEL CLIENTE ===
        ws['A5'] = "INFORMACIÓN DEL CLIENTE"
        ws['A5'].font = subtitle_font
        ws.merge_cells('A5:L5')

        # Datos del cliente
        client_data = [
            ('Cliente:', client.name),
            ('NIT:', client.nit),
            ('Dirección:', client.address or 'No especificada'),
            ('Teléfono:', client.phone or 'No especificado'),
            ('Email:', client.email or 'No especificado'),
            ('Contacto:', client.contact_person or 'No especificado'),
            ('Condición de Pago:', client.get_payment_condition_display()),
        ]

        row = 6
        for label, value in client_data:
            ws.cell(row=row, column=1, value=label).font = label_font
            ws.cell(row=row, column=2, value=value)
            # Unir celdas para el valor para que tenga más espacio
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            row += 1

        # === TABLA DE FACTURAS ===
        start_row = 14
        ws.cell(row=start_row, column=1, value="DETALLE DE FACTURAS").font = subtitle_font
        ws.merge_cells(f'A{start_row}:L{start_row}')

        # Headers de la tabla (añadimos Registro IVA, DUCA, BL, PO, Cód. Generación, Sello Recepción, Total Servicios, Total Gastos)
        headers = ['No. Factura', 'Registro IVA', 'Cód. Generación', 'Sello Recepción', 'Orden de Servicio', 'PO', 'DUCA', 'BL', 'Fecha Emisión', 'Vencimiento',
                   'Total Servicios', 'Total Gastos', 'Total Factura', 'Pagado', 'Saldo', 'Estado']
        header_row = start_row + 1

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Status display
        status_display = {
            'pending': 'Pendiente',
            'partial': 'Pago Parcial',
            'paid': 'Pagada',
            'overdue': 'Vencida',
            'cancelled': 'Anulada',
        }

        # Datos de facturas
        data_row = header_row + 1
        totals = {'total_services': 0, 'total_third_party': 0, 'total': 0, 'paid': 0, 'balance': 0}
        today = datetime.now().date()

        for invoice in invoices:
            days_overdue = ''
            if invoice.due_date and invoice.balance > 0 and today > invoice.due_date:
                days_overdue = f' ({(today - invoice.due_date).days}d)'

            # Columna 1: No. Factura
            ws.cell(row=data_row, column=1, value=invoice.invoice_number).border = thin_border
            
            # Columna 2: Registro IVA (Nuevo)
            ws.cell(row=data_row, column=2, value=client.iva_registration or '').border = thin_border
            
            # Columna 3: Código de Generación
            ws.cell(row=data_row, column=3, value=invoice.generation_code or '').border = thin_border
            
            # Columna 4: Sello de Recepción
            ws.cell(row=data_row, column=4, value=invoice.reception_stamp or '').border = thin_border
            
            # Columna 5: Orden de Servicio
            ws.cell(row=data_row, column=5, value=invoice.service_order.order_number if invoice.service_order else '').border = thin_border
            
            # Columna 6: PO (desde la orden de servicio)
            po_value = invoice.service_order.purchase_order if invoice.service_order and invoice.service_order.purchase_order else ''
            ws.cell(row=data_row, column=6, value=po_value).border = thin_border
            
            # Columna 7: DUCA (desde la orden de servicio)
            duca_value = invoice.service_order.duca if invoice.service_order and invoice.service_order.duca else ''
            ws.cell(row=data_row, column=7, value=duca_value).border = thin_border
            
            # Columna 8: BL (desde la orden de servicio)
            bl_value = invoice.service_order.bl_reference if invoice.service_order and invoice.service_order.bl_reference else ''
            ws.cell(row=data_row, column=8, value=bl_value).border = thin_border
            
            # Columna 9: Fecha Emisión
            ws.cell(row=data_row, column=9, value=invoice.issue_date.strftime('%d/%m/%Y') if invoice.issue_date else '').border = thin_border
            
            # Columna 10: Vencimiento
            ws.cell(row=data_row, column=10, value=invoice.due_date.strftime('%d/%m/%Y') if invoice.due_date else '').border = thin_border

            # Columna 11: Total Servicios
            total_services_cell = ws.cell(row=data_row, column=11, value=float(invoice.total_services))
            total_services_cell.number_format = currency_format
            total_services_cell.border = thin_border

            # Columna 12: Total Gastos a Terceros
            total_third_party_cell = ws.cell(row=data_row, column=12, value=float(invoice.total_third_party))
            total_third_party_cell.number_format = currency_format
            total_third_party_cell.border = thin_border

            # Columna 13: Total Factura
            total_cell = ws.cell(row=data_row, column=13, value=float(invoice.total_amount))
            total_cell.number_format = currency_format
            total_cell.border = thin_border

            # Columna 14: Pagado
            paid_cell = ws.cell(row=data_row, column=14, value=float(invoice.paid_amount))
            paid_cell.number_format = currency_format
            paid_cell.border = thin_border

            # Columna 15: Saldo
            balance_cell = ws.cell(row=data_row, column=15, value=float(invoice.balance))
            balance_cell.number_format = currency_format
            balance_cell.border = thin_border

            # Columna 16: Estado
            status_text = status_display.get(invoice.status, invoice.status) + days_overdue
            ws.cell(row=data_row, column=16, value=status_text).border = thin_border

            # Sumar totales
            totals['total_services'] += float(invoice.total_services)
            totals['total_third_party'] += float(invoice.total_third_party)
            totals['total'] += float(invoice.total_amount)
            totals['paid'] += float(invoice.paid_amount)
            totals['balance'] += float(invoice.balance)

            data_row += 1

        # Fila de totales
        ws.cell(row=data_row, column=1, value="TOTALES").font = Font(bold=True)
        ws.cell(row=data_row, column=1).border = thin_border
        for col in range(2, 11):
            ws.cell(row=data_row, column=col).border = thin_border

        # Total Servicios
        total_services_sum = ws.cell(row=data_row, column=11, value=totals['total_services'])
        total_services_sum.number_format = currency_format
        total_services_sum.font = Font(bold=True)
        total_services_sum.border = thin_border

        # Total Gastos a Terceros
        total_third_party_sum = ws.cell(row=data_row, column=12, value=totals['total_third_party'])
        total_third_party_sum.number_format = currency_format
        total_third_party_sum.font = Font(bold=True)
        total_third_party_sum.border = thin_border

        # Total Factura
        total_total = ws.cell(row=data_row, column=13, value=totals['total'])
        total_total.number_format = currency_format
        total_total.font = Font(bold=True)
        total_total.border = thin_border

        # Total Pagado
        total_paid = ws.cell(row=data_row, column=14, value=totals['paid'])
        total_paid.number_format = currency_format
        total_paid.font = Font(bold=True)
        total_paid.border = thin_border

        # Total Saldo
        total_balance = ws.cell(row=data_row, column=15, value=totals['balance'])
        total_balance.number_format = currency_format
        total_balance.font = Font(bold=True)
        total_balance.border = thin_border

        ws.cell(row=data_row, column=16).border = thin_border

        # Ajustar anchos de columna
        column_widths = [18, 15, 35, 20, 18, 16, 16, 14, 14, 16, 16, 16, 14, 14, 16, 16]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # Generar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'estado_cuenta_{client.name.replace(" ", "_")}_{year}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    @action(detail=False, methods=['get'], permission_classes=[IsOperativo])
    def export_clients_excel(self, request):
        """Exportar listado de clientes a Excel"""
        # Filtros
        queryset = self.filter_queryset(self.get_queryset())

        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Listado de Clientes"

        # Estilos (usando un azul más oscuro - #0F2E4D)
        header_fill = PatternFill(start_color="0F2E4D", end_color="0F2E4D", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        title_font = Font(size=16, bold=True, color="0F2E4D")
        subtitle_font = Font(size=12, bold=True, color="1A4C7A")
        currency_format = '"$"#,##0.00'
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # === ENCABEZADO ===
        ws['A1'] = "LISTADO GENERAL DE CLIENTES"
        ws['A1'].font = title_font
        ws.merge_cells('A1:J1')

        ws['A2'] = "GPRO LOGISTIC - Agencia Aduanal"
        ws['A2'].font = Font(size=11, color="666666")
        ws.merge_cells('A2:J2')

        ws['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A3'].font = Font(size=9, italic=True, color="999999")

        # Headers
        headers = [
            'ID', 'Nombre / Razón Social', 'NIT', 'Teléfono', 'Email',
            'Dirección', 'Contacto', 'Condición Pago', 'Límite Crédito', 'Estado'
        ]

        start_row = 5
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Datos
        data_row = start_row + 1
        for client in queryset:
            ws.cell(row=data_row, column=1, value=client.id).border = thin_border
            ws.cell(row=data_row, column=2, value=client.name).border = thin_border
            ws.cell(row=data_row, column=3, value=client.nit).border = thin_border
            ws.cell(row=data_row, column=4, value=client.phone).border = thin_border
            ws.cell(row=data_row, column=5, value=client.email).border = thin_border
            ws.cell(row=data_row, column=6, value=client.address).border = thin_border
            ws.cell(row=data_row, column=7, value=client.contact_person).border = thin_border

            payment_cond = client.get_payment_condition_display()
            if client.payment_condition == 'credito' and client.credit_days:
                payment_cond += f" ({client.credit_days} días)"

            ws.cell(row=data_row, column=8, value=payment_cond).border = thin_border

            credit_cell = ws.cell(row=data_row, column=9, value=float(client.credit_limit))
            credit_cell.number_format = currency_format
            credit_cell.border = thin_border

            status = "Activo" if client.is_active else "Inactivo"
            ws.cell(row=data_row, column=10, value=status).border = thin_border

            data_row += 1

        # Ajustar anchos
        column_widths = [8, 35, 15, 15, 25, 40, 25, 20, 15, 10]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # Generar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f'clientes_gpro_{datetime.now().strftime("%Y%m%d")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response