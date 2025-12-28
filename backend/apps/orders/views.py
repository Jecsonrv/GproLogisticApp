from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.http import HttpResponse
from django.core.exceptions import ValidationError
from .models import ServiceOrder, OrderDocument, OrderCharge
from .serializers import ServiceOrderSerializer, OrderDocumentSerializer
from .serializers_new import ServiceOrderDetailSerializer
from apps.users.permissions import IsOperativo, IsOperativo2
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import zipfile
import io
import os
from datetime import datetime

class ServiceOrderViewSet(viewsets.ModelViewSet):
    serializer_class = ServiceOrderSerializer
    permission_classes = [IsOperativo]
    filterset_fields = ['status', 'client', 'provider']
    search_fields = ['order_number', 'duca', 'purchase_order']

    def get_queryset(self):
        """
        Optimized queryset with select_related/prefetch_related to prevent N+1 queries.
        Implements Row-Level Security (IDOR protection):
        - Admins/Operativo2: See all orders.
        - Operativo: See only orders assigned to them (customs_agent) or created by them.
        """
        from django.db.models import Prefetch, Q
        user = self.request.user

        queryset = ServiceOrder.objects.select_related(
            'client',
            'sub_client',
            'shipment_type',
            'provider',
            'customs_agent',
            'created_by',
            'closed_by'
        )

        # Para el detalle, prefetch los objetos relacionados
        if self.action == 'retrieve':
            queryset = queryset.prefetch_related(
                Prefetch('documents', queryset=OrderDocument.objects.select_related('uploaded_by')),
                Prefetch('charges', queryset=OrderCharge.objects.select_related('service')),
                'transfers'
            )

        # Row-Level Security (IDOR protection)
        if user.is_authenticated:
            # Si es admin o operativo2, ver todo
            if user.role in ['admin', 'operativo2']:
                return queryset
            # Si es operativo básico, filtrar por órdenes asignadas o creadas
            if user.role == 'operativo':
                return queryset.filter(Q(customs_agent=user) | Q(created_by=user))

        # Fallback de seguridad
        return queryset.none()

    def perform_create(self, serializer):
        """Assign current user as customs_agent when creating order"""
        serializer.save(created_by=self.request.user, customs_agent=self.request.user)

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return ServiceOrderDetailSerializer
        return ServiceOrderSerializer

    @action(detail=True, methods=['get'])
    def billable_items(self, request, pk=None):
        """
        Obtiene todos los items facturables para el asistente de facturación.

        Combina:
        1. OrderCharges (Servicios manuales) - sin factura asociada
        2. Transfers (Gastos reembolsables) - sin factura asociada

        Incluye información detallada sobre:
        - Tipo de IVA (gravado/exento/no_sujeto)
        - Estado de facturación
        - Desglose de montos (neto, IVA, total)
        """
        order = self.get_object()

        items = []

        # 1. Servicios (OrderCharge) - Calculadora de Servicios
        charges = order.charges.filter(
            invoice__isnull=True,
            is_deleted=False
        ).select_related('service')

        for charge in charges:
            iva_type = getattr(charge, 'iva_type', 'gravado')
            items.append({
                'id': f'charge_{charge.id}',
                'original_id': charge.id,
                'type': 'service',
                'description': f"{charge.service.name} - {charge.description or ''}",
                'service_name': charge.service.name,
                # Desglose de montos para cuadre con DTE
                'subtotal_neto': float(charge.subtotal),
                'iva_amount': float(charge.iva_amount),
                'total': float(charge.total),
                # Compatibilidad con frontend existente
                'amount': float(charge.subtotal),
                'iva': float(charge.iva_amount),
                # Información fiscal
                'iva_type': iva_type,
                'iva_type_display': charge.get_iva_type_display_short() if hasattr(charge, 'get_iva_type_display_short') else ('IVA 13%' if charge.service.applies_iva else 'Exento'),
                # Estado
                'billing_status': getattr(charge, 'billing_status', 'disponible'),
                'is_editable': charge.is_editable() if hasattr(charge, 'is_editable') else True,
                # Metadatos
                'quantity': charge.quantity,
                'unit_price': float(charge.unit_price),
                'discount': float(charge.discount),
                'notes': charge.description,
                'source_model': 'OrderCharge'
            })

        # 2. Gastos Reembolsables (Transfer) - Calculadora de Gastos
        from apps.transfers.models import Transfer
        from decimal import Decimal

        transfers = Transfer.objects.filter(
            service_order=order,
            transfer_type__in=['cargos', 'costos', 'terceros'],
            invoice__isnull=True,
            is_deleted=False
        ).select_related('provider')

        for t in transfers:
            # Usar métodos del modelo para cálculos consistentes
            base_price = t.get_customer_base_price()
            iva_amount = t.get_customer_iva_amount()
            total = t.get_customer_total()
            profit = t.get_profit()

            iva_type = getattr(t, 'customer_iva_type', 'exento')

            items.append({
                'id': f'expense_{t.id}',
                'original_id': t.id,
                'type': 'expense',
                'description': f"{t.description} ({t.provider.name if t.provider else 'N/A'})",
                'provider_name': t.provider.name if t.provider else 'N/A',
                # Desglose de montos para cuadre con DTE
                'cost_original': float(t.amount),
                'markup_percentage': float(t.customer_markup_percentage),
                'subtotal_neto': float(base_price),
                'iva_amount': float(iva_amount),
                'total': float(total),
                'profit': float(profit),
                # Compatibilidad con frontend existente
                'amount': float(base_price),
                'iva': float(iva_amount),
                # Información fiscal
                'iva_type': iva_type,
                'iva_type_display': t.get_iva_type_display_short(),
                # Estado y restricciones
                'billing_status': getattr(t, 'billing_status', 'disponible'),
                'is_amount_editable': t.is_amount_editable(),
                'is_config_editable': t.is_billing_config_editable(),
                'amount_locked': getattr(t, 'amount_locked', False),
                # Metadatos
                'notes': f"Costo: ${t.amount} | Margen: {t.customer_markup_percentage}%",
                'source_model': 'Transfer'
            })

        # Calcular resumen
        summary = {
            'services': {
                'count': len([i for i in items if i['type'] == 'service']),
                'subtotal': sum(i['subtotal_neto'] for i in items if i['type'] == 'service'),
                'iva': sum(i['iva_amount'] for i in items if i['type'] == 'service'),
                'total': sum(i['total'] for i in items if i['type'] == 'service')
            },
            'expenses': {
                'count': len([i for i in items if i['type'] == 'expense']),
                'subtotal': sum(i['subtotal_neto'] for i in items if i['type'] == 'expense'),
                'iva': sum(i['iva_amount'] for i in items if i['type'] == 'expense'),
                'total': sum(i['total'] for i in items if i['type'] == 'expense')
            },
            'grand_total': {
                'subtotal_neto': sum(i['subtotal_neto'] for i in items),
                'iva_total': sum(i['iva_amount'] for i in items),
                'total': sum(i['total'] for i in items)
            }
        }

        return Response({
            'items': items,
            'summary': summary,
            'order_number': order.order_number,
            'client_name': order.client.name if order.client else None
        })

    @action(detail=True, methods=['get'])
    def charges(self, request, pk=None):
        """Get all charges for a specific order"""
        order = self.get_object()
        charges = order.charges.all().select_related('service')

        charges_data = []
        for charge in charges:
            charges_data.append({
                'id': charge.id,
                'service': charge.service.id,
                'service_name': charge.service.name,
                'quantity': str(charge.quantity),
                'unit_price': str(charge.unit_price),
                'applies_iva': charge.service.applies_iva,
                'subtotal': str(charge.subtotal),
                'iva_amount': str(charge.iva_amount),
                'total': str(charge.total),
                'notes': charge.description,
                'invoice_id': charge.invoice.id if charge.invoice else None,
                'invoice_number': charge.invoice.invoice_number if charge.invoice else None,
            })

        return Response(charges_data)

    @action(detail=True, methods=['post'])
    def add_charge(self, request, pk=None):
        """Add a charge to a specific order"""
        order = self.get_object()

        if order.status == 'cerrada':
            return Response(
                {'error': 'No se pueden agregar cargos a una orden cerrada'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        service_id = request.data.get('service')
        if not service_id:
            return Response(
                {'error': 'Debe seleccionar un servicio válido'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from apps.catalogs.models import Service
            from decimal import Decimal
            try:
                service = Service.objects.get(id=service_id)
            except (ValueError, TypeError):
                 return Response(
                    {'error': 'ID de servicio inválido'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Convertir valores numéricos correctamente (pueden venir como strings desde el frontend)
            try:
                quantity = int(request.data.get('quantity', 1))
                unit_price = Decimal(str(request.data.get('unit_price', service.default_price)))
                discount = Decimal(str(request.data.get('discount', 0)))
            except (ValueError, TypeError) as e:
                return Response(
                    {'error': f'Valores numéricos inválidos: {str(e)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # The frontend sends 'notes' but model has 'description'.
            # Validar iva_type
            iva_type = request.data.get('iva_type', 'gravado')
            if iva_type not in ['gravado', 'no_sujeto']:
                iva_type = 'gravado'

            charge = OrderCharge.objects.create(
                service_order=order,
                service=service,
                quantity=quantity,
                unit_price=unit_price,
                discount=discount,
                description=request.data.get('notes', ''),
                iva_type=iva_type
            )
            
            # Set current user for signal
            charge._current_user = request.user

            return Response({
                'id': charge.id,
                'message': 'Cargo agregado exitosamente'
            }, status=status.HTTP_201_CREATED)

        except Service.DoesNotExist:
            return Response(
                {'error': 'Servicio no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )
        except ValidationError as e:
            # Extraer mensaje limpio del ValidationError
            if hasattr(e, 'message'):
                error_msg = e.message
            elif hasattr(e, 'messages'):
                error_msg = e.messages[0] if e.messages else str(e)
            else:
                error_msg = str(e).strip("[]'\"")
            return Response(
                {'error': error_msg},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['patch'])
    def update_charge(self, request, pk=None):
        """
        Actualiza un cargo existente.
        Permite modificar: quantity, unit_price, discount, notes, iva_type
        NO permite cambiar el servicio (service)
        """
        order = self.get_object()

        if order.status == 'cerrada':
            return Response(
                {'error': 'No se pueden modificar cargos de una orden cerrada'},
                status=status.HTTP_400_BAD_REQUEST
            )

        charge_id = request.data.get('charge_id')
        if not charge_id:
            return Response(
                {'error': 'Se requiere el ID del cargo'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            charge = OrderCharge.objects.get(id=charge_id, service_order=order)

            # Verificar que el cargo no esté facturado
            if charge.invoice_id:
                return Response(
                    {'error': 'No se puede editar un cargo que ya está facturado'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Validar y convertir valores numéricos
            from decimal import Decimal
            try:
                quantity = int(request.data.get('quantity', charge.quantity))
                unit_price = Decimal(str(request.data.get('unit_price', charge.unit_price)))
                discount = Decimal(str(request.data.get('discount', charge.discount)))
            except (ValueError, TypeError) as e:
                return Response(
                    {'error': f'Valores numéricos inválidos: {str(e)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Validar iva_type
            iva_type = request.data.get('iva_type', charge.iva_type)
            if iva_type not in ['gravado', 'no_sujeto']:
                iva_type = 'gravado'

            # Actualizar campos
            charge.quantity = quantity
            charge.unit_price = unit_price
            charge.discount = discount
            charge.description = request.data.get('notes', charge.description)
            charge.iva_type = iva_type

            # Set current user for signal
            charge._current_user = request.user
            charge.save()

            return Response({
                'id': charge.id,
                'message': 'Cargo actualizado exitosamente'
            }, status=status.HTTP_200_OK)

        except OrderCharge.DoesNotExist:
            return Response(
                {'error': 'Cargo no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )
        except ValidationError as e:
            # Extraer mensaje limpio del ValidationError
            if hasattr(e, 'message'):
                error_msg = e.message
            elif hasattr(e, 'messages'):
                error_msg = e.messages[0] if e.messages else str(e)
            else:
                error_msg = str(e).strip("[]'\"")
            return Response(
                {'error': error_msg},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['post'])
    def update_expense_configurations(self, request, pk=None):
        """
        Actualiza la configuración de cobro para gastos (Transfers) en la Calculadora de Gastos.

        CAMPOS PERMITIDOS POR GASTO:
        - markup_percentage: Margen de utilidad (%)
        - iva_type: Tratamiento fiscal (gravado/exento/no_sujeto)
        - applies_iva: Legacy boolean (se sincroniza con iva_type)

        RESTRICCIONES:
        - El monto base (amount) NO es editable desde aquí
        - Si el gasto ya está facturado con DTE, no se puede modificar

        SINCRONIZACIÓN:
        - Los cambios se reflejan automáticamente en CXC si el gasto está vinculado a una factura
        """
        order = self.get_object()
        configs = request.data.get('configs', [])

        if not configs:
            return Response({'message': 'No hay configuraciones para guardar'}, status=status.HTTP_200_OK)

        try:
            from apps.transfers.models import Transfer
            from apps.orders.models import OrderHistory
            from decimal import Decimal
            from django.db import transaction

            updated_count = 0
            synced_invoices = set()
            changes_log = []

            with transaction.atomic():
                for item in configs:
                    transfer_id = item.get('expense_id')
                    if not transfer_id:
                        continue

                    try:
                        transfer = Transfer.objects.select_for_update().get(
                            id=transfer_id,
                            service_order=order
                        )
                    except Transfer.DoesNotExist:
                        continue

                    # Verificar si el gasto puede ser editado
                    if not transfer.is_billing_config_editable():
                        changes_log.append({
                            'transfer_id': transfer_id,
                            'status': 'error',
                            'message': 'Gasto con DTE emitido, no editable'
                        })
                        continue

                    # Obtener valores anteriores
                    old_markup = transfer.customer_markup_percentage
                    old_iva_type = getattr(transfer, 'customer_iva_type', 'exento')

                    # Actualizar valores
                    new_markup = Decimal(str(item.get('markup_percentage', 0)))
                    new_applies_iva = item.get('applies_iva', False)
                    new_iva_type = item.get('iva_type', 'gravado' if new_applies_iva else 'exento')

                    # Aplicar cambios
                    transfer.customer_markup_percentage = new_markup
                    transfer.customer_iva_type = new_iva_type
                    # customer_applies_iva se sincroniza automáticamente en save()
                    transfer.save()

                    updated_count += 1

                    # Registrar cambio
                    change_entry = {
                        'transfer_id': transfer_id,
                        'description': transfer.description[:50],
                        'old_markup': str(old_markup),
                        'new_markup': str(new_markup),
                        'old_iva_type': old_iva_type,
                        'new_iva_type': new_iva_type
                    }
                    changes_log.append(change_entry)

                    # Si el gasto está vinculado a una factura, marcar para sincronizar
                    if transfer.invoice:
                        synced_invoices.add(transfer.invoice_id)

                # Sincronizar facturas afectadas (bidireccional)
                from apps.orders.models import Invoice
                for invoice_id in synced_invoices:
                    try:
                        invoice = Invoice.objects.get(id=invoice_id)
                        invoice.calculate_totals()
                    except Invoice.DoesNotExist:
                        pass

                # Registrar en historial de la OS
                if updated_count > 0:
                    OrderHistory.objects.create(
                        service_order=order,
                        event_type='updated',
                        description=f'Configuración de cobro actualizada para {updated_count} gastos',
                        user=request.user,
                        metadata={'changes': changes_log}
                    )

            return Response({
                'message': f'Configuración de gastos guardada correctamente',
                'updated_count': updated_count,
                'synced_invoices': len(synced_invoices),
                'changes': changes_log
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Exportar órdenes de servicio a Excel con formato profesional"""
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Órdenes de Servicio"

        # Estilos profesionales - Diseño GPRO
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        title_font = Font(size=16, bold=True, color="1F4E79")
        subtitle_font = Font(size=12, bold=True, color="2F5496")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # === ENCABEZADO ===
        ws['A1'] = "ÓRDENES DE SERVICIO"
        ws['A1'].font = title_font
        ws.merge_cells('A1:J1')

        ws['A2'] = "GPRO LOGISTIC - Agencia Aduanal"
        ws['A2'].font = Font(size=11, color="666666")
        ws.merge_cells('A2:J2')

        ws['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A3'].font = Font(size=9, italic=True, color="999999")

        # === TABLA DE DATOS ===
        start_row = 5
        ws.cell(row=start_row, column=1, value="DETALLE DE ÓRDENES").font = subtitle_font
        ws.merge_cells(f'A{start_row}:J{start_row}')

        # Headers de la tabla
        headers = ['Número Orden', 'Cliente', 'Subcliente', 'Tipo Embarque',
                   'Proveedor', 'PO', 'ETA', 'DUCA', 'Estado', 'Fecha Creación']
        header_row = start_row + 1

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Datos
        queryset = self.filter_queryset(self.get_queryset())
        data_row = header_row + 1

        for order in queryset:
            ws.cell(row=data_row, column=1, value=order.order_number).border = thin_border
            ws.cell(row=data_row, column=2, value=order.client.name if order.client else '').border = thin_border
            ws.cell(row=data_row, column=3, value=order.sub_client.name if order.sub_client else '').border = thin_border
            ws.cell(row=data_row, column=4, value=order.shipment_type.name if order.shipment_type else '').border = thin_border
            ws.cell(row=data_row, column=5, value=order.provider.name if order.provider else '').border = thin_border
            ws.cell(row=data_row, column=6, value=order.purchase_order or '').border = thin_border
            ws.cell(row=data_row, column=7, value=order.eta.strftime('%d/%m/%Y') if order.eta else '').border = thin_border
            ws.cell(row=data_row, column=8, value=order.duca or '').border = thin_border
            ws.cell(row=data_row, column=9, value=order.get_status_display()).border = thin_border
            ws.cell(row=data_row, column=10, value=order.created_at.strftime('%d/%m/%Y')).border = thin_border
            data_row += 1

        # Ajustar anchos de columna
        column_widths = [18, 25, 20, 18, 25, 15, 12, 15, 15, 14]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # Generar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=GPRO_Ordenes_Servicio.xlsx'

        wb.save(response)
        return response

    @action(detail=False, methods=['post'], permission_classes=[IsOperativo2])
    def download_zip(self, request):
        order_ids = request.data.get('order_ids', [])
        if not order_ids:
            return Response({'error': 'No order IDs provided'}, status=400)

        orders = ServiceOrder.objects.filter(id__in=order_ids)
        
        # Create a zip file in memory
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for order in orders:
                documents = order.documents.all()
                for doc in documents:
                    if doc.file:
                        try:
                            file_path = doc.file.path
                            if os.path.exists(file_path):
                                # Structure: OrderNumber/Filename
                                archive_name = f"{order.order_number}/{os.path.basename(file_path)}"
                                zip_file.write(file_path, archive_name)
                        except Exception as e:
                            pass  # Silently skip files that can't be added

        response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename=documentos_ordenes.zip'
        return response

    @action(detail=True, methods=['get'])
    def all_documents(self, request, pk=None):
        """
        Endpoint unificado que retorna TODOS los documentos relacionados a una OS:
        - Documentos directos (OrderDocument)
        - Facturas emitidas (Invoice.pdf_file, Invoice.dte_file)
        - Comprobantes de pago de clientes (InvoicePayment.receipt_file)
        - Notas de crédito a clientes (CreditNote.pdf_file)
        - Facturas de proveedores (Transfer.invoice_file)
        - Comprobantes de pago a proveedores (TransferPayment.proof_file)
        - Notas de crédito de proveedores (ProviderCreditNote.pdf_file)
        """
        from .models import Invoice, InvoicePayment, CreditNote
        from apps.transfers.models import Transfer, TransferPayment, ProviderCreditNote
        
        order = self.get_object()
        documents = []
        
        # 1. Documentos directos de la OS (OrderDocument)
        # IMPORTANTE: Solo mostrar documentos de tipo 'tramite' u 'otros' aquí
        # Los documentos tipo 'factura_costo' se muestran desde Transfer model
        for doc in order.documents.filter(document_type__in=['tramite', 'otros']):
            if doc.file:
                documents.append({
                    'id': f'doc_{doc.id}',
                    'category': doc.document_type,  # Respetar el tipo de documento real
                    'category_label': doc.get_document_type_display(),
                    'subcategory': doc.get_document_type_display(),
                    'file_url': doc.file.url,
                    'file_name': doc.file.name.split('/')[-1] if doc.file else None,
                    'description': doc.description or doc.get_document_type_display(),
                    'reference': None,
                    'reference_label': None,
                    'uploaded_by': doc.uploaded_by.get_full_name() if doc.uploaded_by else None,
                    'uploaded_at': doc.uploaded_at.isoformat() if doc.uploaded_at else None,
                    'amount': None,
                    'deletable': True,
                    'source_model': 'OrderDocument',
                    'source_id': doc.id,
                })
        
        # 2. Facturas emitidas al cliente (Invoice)
        invoices = Invoice.objects.filter(service_order=order)
        for inv in invoices:
            if inv.pdf_file:
                documents.append({
                    'id': f'inv_pdf_{inv.id}',
                    'category': 'factura_venta',
                    'category_label': 'Facturas de Venta',
                    'subcategory': 'PDF Factura',
                    'file_url': inv.pdf_file.url,
                    'file_name': inv.pdf_file.name.split('/')[-1] if inv.pdf_file else None,
                    'description': f'Factura {inv.invoice_number}',
                    'reference': inv.invoice_number,
                    'reference_label': f'Factura #{inv.invoice_number}',
                    'uploaded_by': inv.created_by.get_full_name() if inv.created_by else None,
                    'uploaded_at': inv.created_at.isoformat() if inv.created_at else None,
                    'amount': float(inv.total_amount),
                    'deletable': False,
                    'source_model': 'Invoice',
                    'source_id': inv.id,
                })
            if inv.dte_file:
                documents.append({
                    'id': f'inv_dte_{inv.id}',
                    'category': 'factura_venta',
                    'category_label': 'Facturas de Venta',
                    'subcategory': 'Archivo DTE',
                    'file_url': inv.dte_file.url,
                    'file_name': inv.dte_file.name.split('/')[-1] if inv.dte_file else None,
                    'description': f'DTE Factura {inv.invoice_number}',
                    'reference': inv.invoice_number,
                    'reference_label': f'Factura #{inv.invoice_number}',
                    'uploaded_by': inv.created_by.get_full_name() if inv.created_by else None,
                    'uploaded_at': inv.created_at.isoformat() if inv.created_at else None,
                    'amount': None,
                    'deletable': False,
                    'source_model': 'Invoice',
                    'source_id': inv.id,
                })
            
            # 3. Comprobantes de pago de clientes (InvoicePayment)
            for payment in inv.payments.filter(is_deleted=False):
                if payment.receipt_file:
                    documents.append({
                        'id': f'inv_pay_{payment.id}',
                        'category': 'pago_cliente',
                        'category_label': 'Pagos de Clientes',
                        'subcategory': 'Comprobante de Pago',
                        'file_url': payment.receipt_file.url,
                        'file_name': payment.receipt_file.name.split('/')[-1] if payment.receipt_file else None,
                        'description': f'Pago Factura {inv.invoice_number}',
                        'reference': inv.invoice_number,
                        'reference_label': f'Factura #{inv.invoice_number}',
                        'uploaded_by': payment.created_by.get_full_name() if payment.created_by else None,
                        'uploaded_at': payment.created_at.isoformat() if payment.created_at else None,
                        'amount': float(payment.amount),
                        'deletable': False,
                        'source_model': 'InvoicePayment',
                        'source_id': payment.id,
                    })
            
            # 4. Notas de crédito (CreditNote)
            for cn in inv.credit_notes.filter(is_deleted=False):
                if cn.pdf_file:
                    documents.append({
                        'id': f'cn_{cn.id}',
                        'category': 'nota_credito',
                        'category_label': 'Notas de Crédito',
                        'subcategory': 'Nota de Crédito',
                        'file_url': cn.pdf_file.url,
                        'file_name': cn.pdf_file.name.split('/')[-1] if cn.pdf_file else None,
                        'description': f'NC {cn.note_number} - {cn.reason}',
                        'reference': cn.note_number,
                        'reference_label': f'NC #{cn.note_number} → Factura #{inv.invoice_number}',
                        'uploaded_by': cn.created_by.get_full_name() if cn.created_by else None,
                        'uploaded_at': cn.created_at.isoformat() if cn.created_at else None,
                        'amount': float(cn.amount),
                        'deletable': False,
                        'source_model': 'CreditNote',
                        'source_id': cn.id,
                    })
        
        # 5. Facturas de proveedores / Gastos (Transfer)
        transfers = Transfer.objects.filter(service_order=order, is_deleted=False)
        for tr in transfers:
            if tr.invoice_file:
                documents.append({
                    'id': f'tr_{tr.id}',
                    'category': 'factura_costo',
                    'category_label': 'Facturas de Costo / Proveedores',
                    'subcategory': tr.get_transfer_type_display(),
                    'file_url': tr.invoice_file.url,
                    'file_name': tr.invoice_file.name.split('/')[-1] if tr.invoice_file else None,
                    'description': tr.description[:100] if tr.description else f'Gasto {tr.get_transfer_type_display()}',
                    'reference': tr.invoice_number or tr.ccf or f'ID:{tr.id}',
                    'reference_label': f'{tr.get_transfer_type_display()} - {tr.provider.name if tr.provider else "Sin proveedor"}',
                    'uploaded_by': tr.created_by.get_full_name() if tr.created_by else None,
                    'uploaded_at': tr.created_at.isoformat() if tr.created_at else None,
                    'amount': float(tr.amount),
                    'deletable': False,
                    'source_model': 'Transfer',
                    'source_id': tr.id,
                })
            
            # 6. Comprobantes de pago a proveedores (TransferPayment)
            for tp in tr.payments.filter(is_deleted=False):
                if tp.proof_file:
                    documents.append({
                        'id': f'tr_pay_{tp.id}',
                        'category': 'pago_proveedor',
                        'category_label': 'Pagos a Proveedores',
                        'subcategory': 'Comprobante de Pago',
                        'file_url': tp.proof_file.url,
                        'file_name': tp.proof_file.name.split('/')[-1] if tp.proof_file else None,
                        'description': f'Pago a {tr.provider.name if tr.provider else "proveedor"} - {tr.description[:50] if tr.description else ""}',
                        'reference': tp.reference_number or f'ID:{tp.id}',
                        'reference_label': f'Pago → {tr.get_transfer_type_display()}',
                        'uploaded_by': tp.created_by.get_full_name() if tp.created_by else None,
                        'uploaded_at': tp.created_at.isoformat() if tp.created_at else None,
                        'amount': float(tp.amount),
                        'deletable': False,
                        'source_model': 'TransferPayment',
                        'source_id': tp.id,
                    })

            # 7. Notas de crédito de proveedores vinculadas a este Transfer
            # Se buscan NC que tengan este transfer como original_transfer
            provider_credit_notes = ProviderCreditNote.objects.filter(
                original_transfer=tr,
                is_deleted=False
            )
            for pcn in provider_credit_notes:
                if pcn.pdf_file:
                    documents.append({
                        'id': f'pcn_{pcn.id}',
                        'category': 'nc_proveedor',
                        'category_label': 'NC de Proveedores',
                        'subcategory': 'Nota de Crédito Proveedor',
                        'file_url': pcn.pdf_file.url,
                        'file_name': pcn.pdf_file.name.split('/')[-1] if pcn.pdf_file else None,
                        'description': f'NC {pcn.note_number} - {pcn.get_reason_display()}',
                        'reference': pcn.note_number,
                        'reference_label': f'NC #{pcn.note_number} → Fact. {tr.invoice_number or tr.id}',
                        'uploaded_by': pcn.created_by.get_full_name() if pcn.created_by else None,
                        'uploaded_at': pcn.created_at.isoformat() if pcn.created_at else None,
                        'amount': float(pcn.amount),
                        'deletable': False,
                        'source_model': 'ProviderCreditNote',
                        'source_id': pcn.id,
                    })

        # Ordenar por fecha (más reciente primero)
        documents.sort(key=lambda x: x['uploaded_at'] or '', reverse=True)
        
        # Generar resumen por categoría
        categories_summary = {}
        for doc in documents:
            cat = doc['category']
            if cat not in categories_summary:
                categories_summary[cat] = {
                    'label': doc['category_label'],
                    'count': 0,
                    'total_amount': 0
                }
            categories_summary[cat]['count'] += 1
            if doc['amount']:
                categories_summary[cat]['total_amount'] += doc['amount']
        
        return Response({
            'order_number': order.order_number,
            'total_documents': len(documents),
            'categories_summary': categories_summary,
            'documents': documents
        })

class OrderDocumentViewSet(viewsets.ModelViewSet):
    queryset = OrderDocument.objects.all()
    serializer_class = OrderDocumentSerializer
    permission_classes = [IsOperativo]
    filterset_fields = ['order', 'document_type']
    
    def perform_destroy(self, instance):
        """Set current user before deletion for signal"""
        instance._current_user = self.request.user
        instance.delete()


class OrderChargeViewSet(viewsets.ModelViewSet):
    """ViewSet for managing order charges"""
    queryset = OrderCharge.objects.all()
    permission_classes = [IsOperativo]

    def get_queryset(self):
        return self.queryset.select_related('service_order', 'service')

    def destroy(self, request, *args, **kwargs):
        """Delete a charge only if the order is still open (not closed/cerrada)"""
        charge = self.get_object()

        # CORREGIDO: Usar estado valido 'cerrada' en lugar de 'abierta' que no existe
        # Estados validos: pendiente, en_puerto, en_transito, en_almacen, finalizada, cerrada
        if charge.service_order.status == 'cerrada':
            return Response(
                {'error': 'No se pueden eliminar cargos de una orden cerrada'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar tambien si la orden esta facturada
        if charge.service_order.facturado:
            return Response(
                {'error': 'No se pueden eliminar cargos de una orden facturada'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Set current user for signal
        charge._current_user = request.user
        charge.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class OrderHistoryViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for order history (read-only)"""
    from .serializers_new import OrderHistorySerializer
    serializer_class = OrderHistorySerializer
    permission_classes = [IsOperativo]
    
    def get_queryset(self):
        from .models import OrderHistory
        queryset = OrderHistory.objects.select_related('user', 'service_order')
        
        # Filtrar por service_order si se proporciona
        service_order_id = self.request.query_params.get('service_order', None)
        if service_order_id:
            queryset = queryset.filter(service_order_id=service_order_id)
        
        return queryset