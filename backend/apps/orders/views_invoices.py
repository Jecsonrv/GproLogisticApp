from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Sum, Q, Count
from django.http import HttpResponse
from django.db import transaction
from django.core.files.base import ContentFile
from decimal import Decimal
import openpyxl
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .models import Invoice, InvoicePayment, ServiceOrder, CreditNote
from apps.catalogs.models import Bank
from .serializers import InvoiceListSerializer, InvoicePaymentSerializer, CreditNoteSerializer
from .serializers_new import InvoiceDetailSerializer, InvoiceCreateSerializer
from apps.orders.pdf_generator import generate_invoice_pdf
from apps.users.permissions import IsOperativo, IsOperativo2

# Import distributed lock utilities (only active when Redis is configured)
try:
    from apps.core.cache import distributed_lock, LockAcquisitionError
    LOCKS_ENABLED = True
except ImportError:
    LOCKS_ENABLED = False
    distributed_lock = None
    LockAcquisitionError = Exception


class InvoiceViewSet(viewsets.ModelViewSet):
    """ViewSet for managing invoices (CXC)"""
    permission_classes = [IsOperativo]
    serializer_class = InvoiceDetailSerializer
    filterset_fields = ['status']
    search_fields = ['invoice_number', 'service_order__client__name', 'ccf']
    ordering_fields = ['issue_date', 'due_date', 'total_amount', 'balance']
    ordering = ['-issue_date']

    def get_serializer_class(self):
        if self.action == 'list':
            return InvoiceListSerializer
        if self.action == 'create':
            return InvoiceCreateSerializer
        return InvoiceDetailSerializer

    def retrieve(self, request, *args, **kwargs):
        """
        Retrieve details of an invoice with automatic PDF self-healing.
        If the PDF record exists but the file is missing from S3, it regenerates it.
        """
        instance = self.get_object()
        
        # --- SELF-HEALING LOGIC ---
        # Verificar integridad del archivo PDF si se supone que existe
        if instance.pdf_file:
            try:
                # Verificar si existe físicamente en el storage (S3/Local)
                if not instance.pdf_file.storage.exists(instance.pdf_file.name):
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.warning(f"PDF missing for Invoice {instance.invoice_number}. Regenerating...")
                    
                    # Regenerar PDF
                    pdf_buffer = generate_invoice_pdf(instance)
                    filename = os.path.basename(instance.pdf_file.name) or f"Invoice_{instance.invoice_number}.pdf"
                    
                    # Guardar (esto sube al storage)
                    instance.pdf_file.save(filename, ContentFile(pdf_buffer.getvalue()), save=True)
                    logger.info(f"PDF successfully regenerated for Invoice {instance.invoice_number}")
            except Exception as e:
                # No bloquear la vista si falla la reparación, solo loggear
                print(f"Error in PDF self-healing: {e}")
        # --------------------------

        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def download_pdf(self, request, pk=None):
        """
        Endpoint seguro para descargar/ver el PDF.
        Garantiza que el archivo exista antes de redirigir.
        """
        invoice = self.get_object()
        
        if not invoice.pdf_file:
            # Si no tiene PDF asignado, intentar generarlo
            try:
                buffer = generate_invoice_pdf(invoice)
                filename = f"Invoice_{invoice.invoice_number}.pdf"
                invoice.pdf_file.save(filename, ContentFile(buffer.getvalue()), save=True)
            except Exception as e:
                return Response(
                    {'error': f'La factura no tiene PDF y falló la generación: {str(e)}'},
                    status=status.HTTP_404_NOT_FOUND
                )

        # Verificar existencia física y reparar si es necesario
        try:
            if not invoice.pdf_file.storage.exists(invoice.pdf_file.name):
                buffer = generate_invoice_pdf(invoice)
                filename = os.path.basename(invoice.pdf_file.name)
                invoice.pdf_file.save(filename, ContentFile(buffer.getvalue()), save=True)
        except Exception as e:
             return Response(
                {'error': f'El archivo no se encuentra y no se pudo regenerar: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
            
        # Redirigir a la URL del archivo
        return HttpResponse(status=302, headers={'Location': invoice.pdf_file.url})

    def create(self, request, *args, **kwargs):
        """Override create to ensure response includes invoice_number and linked items"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        try:
            self.perform_create(serializer)
        except ValueError as e:
            # Capturar errores de validación de negocio (incluyendo falla de subida)
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            # Capturar errores inesperados
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error creating invoice: {str(e)}", exc_info=True)
            return Response(
                {'error': 'Ocurrió un error inesperado al crear la factura. Por favor intente nuevamente.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        # Get the created invoice with the auto-generated invoice_number
        # Refresh from DB to get auto-generated fields
        invoice = serializer.instance
        invoice.refresh_from_db()

        # Use detail serializer for response to include all fields
        response_serializer = InvoiceDetailSerializer(invoice)
        headers = self.get_success_headers(response_serializer.data)
        return Response(response_serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def check_storage_integrity(self, invoice):
        """
        Verifica que el archivo PDF realmente exista en el storage (S3).
        Si no existe, lanza una excepción para revertir la transacción.
        """
        if invoice.pdf_file:
            try:
                # Forzar verificación con el storage backend
                if not invoice.pdf_file.storage.exists(invoice.pdf_file.name):
                    raise ValueError(
                        f"Error crítico: El archivo PDF '{invoice.pdf_file.name}' no se guardó correctamente en el almacenamiento. "
                        "La operación ha sido cancelada para evitar inconsistencias."
                    )
            except Exception as e:
                # Si falla la conexión con S3 o cualquier otro error de storage
                raise ValueError(f"Error de verificación de almacenamiento: {str(e)}")

    def perform_create(self, serializer):
        """Create invoice from service order with distributed lock"""
        service_order_id = self.request.data.get('service_order')
        
        if not service_order_id:
            raise ValueError("Se requiere una orden de servicio")

        # Use distributed lock to prevent concurrent invoicing of the same order
        def create_invoice():
            total_amount = Decimal(str(self.request.data.get('total_amount', 0)))
            invoice_number = self.request.data.get('invoice_number', '')

            # If invoice_number is empty or just whitespace, set to empty string
            # so the model's save() method will auto-generate it
            if invoice_number and isinstance(invoice_number, str):
                invoice_number = invoice_number.strip()
                if not invoice_number:
                    invoice_number = ''
            
            # VALIDACIÓN CRÍTICA: Verificar que el invoice_number no exista ya
            # para evitar errores de UniqueViolation
            if invoice_number:
                from .models import Invoice
                if Invoice.objects.filter(invoice_number=invoice_number).exists():
                    from rest_framework.exceptions import ValidationError
                    raise ValidationError({
                        'invoice_number': f'Ya existe una factura con el número {invoice_number}. '
                                         f'Por favor, verifique o use un número diferente.'
                    })

            # Get dates
            from datetime import datetime
            issue_date = self.request.data.get('issue_date')
            if isinstance(issue_date, str):
                issue_date = datetime.strptime(issue_date, '%Y-%m-%d').date()

            due_date = self.request.data.get('due_date')
            if due_date and isinstance(due_date, str):
                due_date = datetime.strptime(due_date, '%Y-%m-%d').date()

            try:
                service_order = ServiceOrder.objects.select_for_update().get(id=service_order_id)
            except ServiceOrder.DoesNotExist:
                raise ValueError("Orden de servicio no encontrada")

            # Remove single invoice restriction to allow partial billing
            # if service_order.facturado:
            #    raise ValueError("Esta orden de servicio ya fue facturada")

            # Get file if provided
            pdf_file = self.request.FILES.get('pdf_file')
            dte_file = self.request.FILES.get('dte_file')

            # Create invoice with values
            invoice_data = {
                'created_by': self.request.user,
                'invoice_number': invoice_number if invoice_number else '',
                'issue_date': issue_date,
                'due_date': due_date,
                'total_amount': total_amount,
                'balance': total_amount,
                'paid_amount': Decimal('0.00')
            }

            # Add files if provided
            if pdf_file:
                invoice_data['pdf_file'] = pdf_file
            if dte_file:
                invoice_data['dte_file'] = dte_file

            invoice = serializer.save(**invoice_data)
            
            # --- VALIDACIÓN DE INTEGRIDAD DE ARCHIVO (SOLUCIÓN PERMANENTE) ---
            # Si se proporcionó un PDF, verificar inmediatamente que se haya subido a S3/Storage.
            # Si falla, la excepción hará rollback de toda la transacción atómica.
            if pdf_file:
                self.check_storage_integrity(invoice)
            # ---------------------------------------------------------------

            # 1. Link selected MANUAL charges to this invoice
            # Use getlist() for FormData arrays, fallback to get() for JSON
            charge_ids_raw = self.request.data.getlist('charge_ids', []) if hasattr(self.request.data, 'getlist') else self.request.data.get('charge_ids', [])

            # Debug logging
            import logging
            logger = logging.getLogger(__name__)
            logger.info(f"Raw charge_ids received: {charge_ids_raw}")

            # Convert to integers, filtering out empty strings and None
            charge_ids = []
            if charge_ids_raw:
                from .models import OrderCharge
                for id_val in charge_ids_raw:
                    # Handle potential list if nested (shouldn't happen but defensive)
                    if isinstance(id_val, list):
                        for sub_val in id_val:
                            if sub_val and str(sub_val).strip():
                                try:
                                    # Handle comma-separated strings
                                    if isinstance(sub_val, str) and ',' in sub_val:
                                        for part in sub_val.split(','):
                                            if part.strip():
                                                charge_ids.append(int(part.strip()))
                                    else:
                                        charge_ids.append(int(sub_val))
                                except (ValueError, TypeError):
                                    logger.warning(f"Invalid charge_id: {sub_val}")
                    # Handle single value
                    elif id_val and str(id_val).strip():
                        try:
                            # Handle comma-separated strings
                            if isinstance(id_val, str) and ',' in id_val:
                                for part in id_val.split(','):
                                    if part.strip():
                                        charge_ids.append(int(part.strip()))
                            else:
                                charge_ids.append(int(id_val))
                        except (ValueError, TypeError):
                            logger.warning(f"Invalid charge_id: {id_val}")

                logger.info(f"Processed charge_ids: {charge_ids}")

                if charge_ids:
                    updated = OrderCharge.objects.filter(
                        id__in=charge_ids,
                        service_order=service_order,
                        invoice__isnull=True
                    ).update(invoice=invoice, billing_status='facturado')
                    logger.info(f"Updated {updated} charges with invoice {invoice.id}")

            # 2. Link selected TRANSFERS (Expenses) directly to invoice (without creating OrderCharges)
            # Los gastos se quedan en su tabla, solo se marcan como facturados
            # Use getlist() for FormData arrays, fallback to get() for JSON
            transfer_ids_raw = self.request.data.getlist('transfer_ids', []) if hasattr(self.request.data, 'getlist') else self.request.data.get('transfer_ids', [])

            logger.info(f"Raw transfer_ids received: {transfer_ids_raw}")

            # Convert to integers, filtering out empty strings and None
            transfer_ids = []
            if transfer_ids_raw:
                from apps.transfers.models import Transfer
                for id_val in transfer_ids_raw:
                    # Handle potential list
                    if isinstance(id_val, list):
                        for sub_val in id_val:
                            if sub_val and str(sub_val).strip():
                                try:
                                    if isinstance(sub_val, str) and ',' in sub_val:
                                        for part in sub_val.split(','):
                                            if part.strip():
                                                transfer_ids.append(int(part.strip()))
                                    else:
                                        transfer_ids.append(int(sub_val))
                                except (ValueError, TypeError):
                                    logger.warning(f"Invalid transfer_id: {sub_val}")
                    # Handle single value
                    elif id_val and str(id_val).strip():
                        try:
                            if isinstance(id_val, str) and ',' in id_val:
                                for part in id_val.split(','):
                                    if part.strip():
                                        transfer_ids.append(int(part.strip()))
                            else:
                                transfer_ids.append(int(id_val))
                        except (ValueError, TypeError):
                            logger.warning(f"Invalid transfer_id: {id_val}")

                logger.info(f"Processed transfer_ids: {transfer_ids}")

                if transfer_ids:
                    # Marcar los transfers como facturados (sin moverlos a OrderCharge)
                    updated = Transfer.objects.filter(
                        id__in=transfer_ids,
                        service_order=service_order,
                        invoice__isnull=True  # Solo los no facturados
                    ).update(invoice=invoice, billing_status='facturado')
                    logger.info(f"Updated {updated} transfers with invoice {invoice.id}")

            # Recalculate totals based on linked charges AND transfers
            invoice.calculate_totals()

            # Mark service order as invoiced (billing started)
            service_order.facturado = True
            service_order.save()

            return invoice

        # Apply distributed lock if Redis is available
        if LOCKS_ENABLED and distributed_lock:
            try:
                with distributed_lock(f'invoice_os_{service_order_id}', timeout=30):
                    with transaction.atomic():
                        return create_invoice()
            except LockAcquisitionError:
                raise ValueError("Otro usuario está procesando esta orden. Intente nuevamente.")
        else:
            # Fallback to database-level lock only
            with transaction.atomic():
                return create_invoice()

    def get_queryset(self):
        from django.db.models import Prefetch
        queryset = Invoice.objects.all().select_related(
            'service_order', 'service_order__client', 'service_order__sub_client', 'created_by'
        ).prefetch_related(
            Prefetch('payments', queryset=InvoicePayment.objects.filter(is_deleted=False).select_related('bank', 'created_by')),
            Prefetch('credit_notes', queryset=CreditNote.objects.filter(is_deleted=False).select_related('created_by'))
        )

        # Filter by status
        status_filter = self.request.query_params.get('status')
        if status_filter == 'pendiente':
            queryset = queryset.filter(balance__gt=0)
        elif status_filter == 'pagada':
            queryset = queryset.filter(balance=0)
        elif status_filter == 'vencida':
            from django.utils import timezone
            queryset = queryset.filter(
                due_date__lt=timezone.now().date(),
                balance__gt=0
            )

        # Filter by client (through service_order)
        client_id = self.request.query_params.get('client')
        if client_id:
            queryset = queryset.filter(service_order__client_id=client_id)

        # Filter by month
        mes = self.request.query_params.get('mes')
        if mes:
            queryset = queryset.filter(mes=mes)

        return queryset

    def partial_update(self, request, *args, **kwargs):
        """
        Actualización parcial de factura con control de campos según estado DTE.

        Campos SIEMPRE editables (incluso con DTE emitido):
        - invoice_number: Número de factura (corrección de errores)
        - due_date: Fecha de vencimiento
        - notes: Notas internas
        - pdf_file: Archivo PDF

        Campos SOLO editables si NO tiene DTE emitido:
        - invoice_type: Tipo de documento
        - issue_date: Fecha de emisión
        """
        invoice = self.get_object()

        # Campos siempre permitidos
        ALWAYS_EDITABLE = {'invoice_number', 'dte_number', 'due_date', 'notes', 'pdf_file'}

        # Campos solo editables sin DTE
        DTE_RESTRICTED = {'invoice_type', 'issue_date'}

        # Si tiene DTE emitido, filtrar los campos restringidos
        if invoice.is_dte_issued:
            restricted_fields = set(request.data.keys()) & DTE_RESTRICTED
            if restricted_fields:
                # Crear una copia mutable de request.data
                data = request.data.copy()
                # Remover campos restringidos
                for field in restricted_fields:
                    data.pop(field, None)
                # Reemplazar request.data con la copia filtrada
                request._full_data = data

        # Proceder con la actualización normal
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        """Delete invoice and unmark service order as invoiced"""
        invoice = self.get_object()

        # Check if invoice has payments (current)
        if invoice.paid_amount > 0:
            return Response(
                {'error': 'No se puede eliminar una factura con pagos registrados. Elimine los pagos primero.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # AUDITORÍA #2: Verificar historial de pagos (incluyendo eliminados)
        # para mantener trazabilidad financiera
        from .models import InvoicePayment
        payment_history = InvoicePayment.all_objects.filter(invoice=invoice)
        if payment_history.exists():
            deleted_payments = payment_history.filter(is_deleted=True)
            if deleted_payments.exists():
                return Response(
                    {
                        'error': 'Esta factura tiene historial de pagos eliminados. '
                                'Para mantener la trazabilidad financiera, use la opción de anular factura en lugar de eliminar.',
                        'code': 'PAYMENT_HISTORY_EXISTS',
                        'deleted_payments_count': deleted_payments.count()
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        # Verificar si está emitido como DTE y ha pasado el periodo de gracia de 24 horas
        if invoice.is_dte_issued:
            if not invoice.can_delete_without_credit_note():
                hours_remaining = invoice.get_hours_until_locked()
                return Response(
                    {
                        'error': 'Esta factura tiene DTE emitido y ha pasado el plazo de 24 horas. '
                                'Solo puede anularse mediante Nota de Crédito.',
                        'code': 'DTE_LOCKED',
                        'dte_issued_at': invoice.dte_issued_at.isoformat() if invoice.dte_issued_at else None
                    },
                    status=status.HTTP_400_BAD_REQUEST
                )
            else:
                # Permitir eliminación pero advertir que está en el periodo de gracia
                hours_remaining = invoice.get_hours_until_locked()
                # Continuar con la eliminación (se registrará en los logs)

        try:
            service_order = invoice.service_order
            
            # Delete invoice
            invoice.delete()

            # Unmark service order as invoiced if it has no other invoices
            if service_order and not service_order.invoices.exists():
                service_order.facturado = False
                service_order.save()

            return Response(
                {'message': 'Factura eliminada correctamente'},
                status=status.HTTP_204_NO_CONTENT
            )

        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['post'])
    def mark_as_dte(self, request, pk=None):
        """
        Marca una pre-factura como DTE emitido.
        Una vez marcada, solo se pueden hacer notas de crédito.
        REQUIERE que se haya subido el PDF de la factura.
        """
        invoice = self.get_object()
        
        if invoice.is_dte_issued:
            return Response(
                {'error': 'Esta factura ya fue marcada como DTE emitido.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar que existe el PDF de la factura antes de marcar como emitido
        if not invoice.pdf_file:
            return Response(
                {
                    'error': 'No se puede marcar como DTE emitido sin haber subido el PDF de la factura.',
                    'code': 'PDF_REQUIRED'
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        dte_number = request.data.get('dte_number', '')
        
        # Registrar en historial
        from .models import InvoiceEditHistory
        from django.utils import timezone
        
        InvoiceEditHistory.objects.create(
            invoice=invoice,
            edit_type='dte_marked',
            description=f'Factura marcada como DTE emitido. Número DTE: {dte_number or "N/A"}',
            previous_values={'is_dte_issued': False, 'invoice_number': invoice.invoice_number},
            new_values={'is_dte_issued': True, 'dte_number': dte_number, 'invoice_number': dte_number},
            user=request.user
        )
        
        invoice.is_dte_issued = True
        invoice.dte_issued_at = timezone.now()  # Establecer timestamp de emisión
        invoice.dte_number = dte_number
        if dte_number:
            invoice.invoice_number = dte_number
        invoice.save()
        
        return Response({
            'message': 'Factura marcada como DTE emitido correctamente.',
            'invoice_number': invoice.invoice_number,
            'dte_number': dte_number
        })

    @action(detail=True, methods=['post'])
    def void(self, request, pk=None):
        """
        AUDITORÍA #5: Anular una factura (void/cancel).

        Este endpoint permite anular facturas que tienen pagos o historial,
        a diferencia de destroy() que solo permite eliminar facturas limpias.

        COMPORTAMIENTO:
        - Marca la factura como status='cancelled'
        - Establece balance=0 para liberar crédito del cliente
        - NO elimina los pagos (se mantienen para auditoría)
        - Registra el historial de anulación
        - Libera los items (cargos y gastos) para que puedan refacturarse

        REQUISITOS:
        - Motivo de anulación obligatorio
        - Usuario autenticado

        NOTA: Los pagos ya realizados quedan registrados. Si el cliente
        pagó en exceso, se debe gestionar como saldo a favor por fuera del sistema.
        """
        from django.utils import timezone
        from .models import InvoiceEditHistory, OrderCharge
        from apps.transfers.models import Transfer

        invoice = self.get_object()

        # Validar que no esté ya anulada
        if invoice.status == 'cancelled':
            return Response(
                {'error': 'Esta factura ya está anulada.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar motivo de anulación
        void_reason = request.data.get('reason', '').strip()
        if not void_reason:
            return Response(
                {'error': 'Debe proporcionar un motivo para anular la factura.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            with transaction.atomic():
                # Obtener lock sobre la factura
                inv = Invoice.objects.select_for_update().get(pk=pk)

                # Guardar valores anteriores para historial
                previous_values = {
                    'status': inv.status,
                    'balance': str(inv.balance),
                    'paid_amount': str(inv.paid_amount),
                    'credited_amount': str(inv.credited_amount)
                }

                # 1. Liberar cargos (OrderCharge) vinculados a esta factura
                released_charges = []
                for charge in inv.charges.filter(is_deleted=False):
                    released_charges.append({
                        'id': charge.id,
                        'service': charge.service.name,
                        'total': str(charge.total)
                    })
                    charge.invoice = None
                    charge.billing_status = 'disponible'
                    charge.save(skip_order_validation=True)

                # 2. Liberar gastos (Transfer) vinculados a esta factura
                released_expenses = []
                for transfer in inv.billed_transfers.filter(is_deleted=False):
                    released_expenses.append({
                        'id': transfer.id,
                        'description': transfer.description[:50],
                        'total': str(transfer.get_customer_total())
                    })
                    transfer.invoice = None
                    transfer.billing_status = 'disponible'
                    transfer.save()

                # 3. Marcar factura como anulada
                inv.status = 'cancelled'
                inv.balance = Decimal('0.00')  # Liberar crédito

                # Agregar campos de anulación si no existen
                inv.notes = f"{inv.notes}\n\n--- ANULADA ---\nFecha: {timezone.now().strftime('%Y-%m-%d %H:%M')}\nMotivo: {void_reason}\nUsuario: {request.user.get_full_name() or request.user.username}"

                inv.save()

                # 4. Registrar en historial
                InvoiceEditHistory.objects.create(
                    invoice=inv,
                    edit_type='dte_marked',  # Reutilizamos este tipo para anulación
                    description=f'FACTURA ANULADA: {void_reason}',
                    previous_values=previous_values,
                    new_values={
                        'status': 'cancelled',
                        'balance': '0.00',
                        'void_reason': void_reason,
                        'released_charges': len(released_charges),
                        'released_expenses': len(released_expenses)
                    },
                    user=request.user
                )

                # 5. Registrar en historial de la OS
                from .models import OrderHistory
                OrderHistory.objects.create(
                    service_order=inv.service_order,
                    event_type='updated',
                    description=f'Factura {inv.invoice_number} ANULADA. Motivo: {void_reason}',
                    user=request.user,
                    metadata={
                        'invoice_id': inv.id,
                        'invoice_number': inv.invoice_number,
                        'void_reason': void_reason,
                        'released_items': {
                            'charges': released_charges,
                            'expenses': released_expenses
                        }
                    }
                )

            return Response({
                'message': f'Factura {invoice.invoice_number} anulada correctamente.',
                'invoice_id': invoice.id,
                'invoice_number': invoice.invoice_number,
                'void_reason': void_reason,
                'credit_released': True,
                'items_released': {
                    'charges': len(released_charges),
                    'expenses': len(released_expenses)
                },
                'note': 'Los items liberados están disponibles para facturar nuevamente.'
            })

        except Exception as e:
            return Response(
                {'error': f'Error al anular la factura: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['patch'])
    def edit_expense(self, request, pk=None):
        """
        Editar un gasto facturado (solo si es pre-factura).

        RESTRICCIONES DE INTEGRIDAD:
        - El Monto Base (amount) NO es editable si tiene pagos o está bloqueado
        - Solo se permite editar: Margen de Utilidad y configuración de IVA
        - Los cambios se sincronizan bidireccionalmente con la OS

        CAMPOS PERMITIDOS:
        - customer_markup_percentage: Margen de utilidad (%)
        - customer_iva_type: Tratamiento fiscal (gravado/exento/no_sujeto)
        - customer_applies_iva: Legacy boolean (se sincroniza automáticamente)
        """
        invoice = self.get_object()

        if invoice.is_dte_issued:
            return Response(
                {'error': 'No se puede editar. Esta factura ya tiene DTE emitido. Use notas de crédito.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        transfer_id = request.data.get('transfer_id')
        if not transfer_id:
            return Response(
                {'error': 'Se requiere transfer_id'},
                status=status.HTTP_400_BAD_REQUEST
            )

        from apps.transfers.models import Transfer
        
        with transaction.atomic():
            try:
                transfer = Transfer.objects.select_for_update().get(id=transfer_id, invoice=invoice)
            except Transfer.DoesNotExist:
                return Response(
                    {'error': 'Gasto no encontrado en esta factura'},
                    status=status.HTTP_404_NOT_FOUND
                )

            # RESTRICCIÓN: El monto base NO es editable
            if 'amount' in request.data:
                if not transfer.is_amount_editable():
                    return Response(
                        {'error': 'El monto base no es editable. Viene del pago al proveedor y no puede modificarse.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            # Guardar valores anteriores para audit trail
            previous_values = {
                'amount': str(transfer.amount),
                'customer_markup_percentage': str(transfer.customer_markup_percentage),
                'customer_iva_type': transfer.customer_iva_type,
                'customer_applies_iva': transfer.customer_applies_iva
            }

            changes_made = []
            iva_type_already_set = False

            # Actualizar margen si se proporciona
            if 'customer_markup_percentage' in request.data:
                old_markup = transfer.customer_markup_percentage
                new_markup = Decimal(str(request.data['customer_markup_percentage']))
                if old_markup != new_markup:
                    transfer.customer_markup_percentage = new_markup
                    changes_made.append(f'Margen: {old_markup}% → {new_markup}%')

            # Actualizar tipo de IVA si se proporciona (nuevo campo)
            if 'customer_iva_type' in request.data:
                old_iva_type = transfer.customer_iva_type
                new_iva_type = request.data['customer_iva_type']
                if old_iva_type != new_iva_type:
                    transfer.customer_iva_type = new_iva_type
                    # Sincronizar customer_applies_iva con el tipo de IVA
                    transfer.customer_applies_iva = (new_iva_type == 'gravado')
                    changes_made.append(f'Tipo IVA: {old_iva_type} → {new_iva_type}')
                iva_type_already_set = True

            # Compatibilidad: actualizar legacy applies_iva (solo si no se envió customer_iva_type)
            if 'customer_applies_iva' in request.data and not iva_type_already_set:
                old_applies = transfer.customer_applies_iva
                new_applies = request.data['customer_applies_iva']
                if old_applies != new_applies:
                    # Sincronizar con customer_iva_type (usar no_sujeto en lugar de exento)
                    transfer.customer_iva_type = 'gravado' if new_applies else 'no_sujeto'
                    transfer.customer_applies_iva = new_applies
                    changes_made.append(f'Aplica IVA: {old_applies} → {new_applies}')

            transfer.save()

            # Registrar en historial con detalle de cambios
            from .models import InvoiceEditHistory

            if changes_made:
                InvoiceEditHistory.objects.create(
                    invoice=invoice,
                    edit_type='expense_edited',
                    description=f'Gasto editado: {transfer.description[:50]}... | Cambios: {", ".join(changes_made)}',
                    previous_values=previous_values,
                    new_values={
                        'amount': str(transfer.amount),
                        'customer_markup_percentage': str(transfer.customer_markup_percentage),
                        'customer_iva_type': transfer.customer_iva_type,
                        'customer_applies_iva': transfer.customer_applies_iva
                    },
                    user=request.user
                )

            # Recalcular totales de la factura (sincronización automática)
            invoice.calculate_totals()

        return Response({
            'message': 'Gasto actualizado correctamente',
            'transfer_id': transfer.id,
            'changes': changes_made,
            'new_values': {
                'customer_markup_percentage': str(transfer.customer_markup_percentage),
                'customer_iva_type': transfer.customer_iva_type,
                'base_price': str(transfer.get_customer_base_price()),
                'iva_amount': str(transfer.get_customer_iva_amount()),
                'total': str(transfer.get_customer_total())
            }
        })

    @action(detail=True, methods=['patch'])
    def edit_charge(self, request, pk=None):
        """
        Editar un cargo/servicio facturado (solo si es pre-factura).
        Los cambios se reflejan tanto en la factura como en la OS.
        """
        invoice = self.get_object()
        
        if invoice.is_dte_issued:
            return Response(
                {'error': 'No se puede editar. Esta factura ya tiene DTE emitido. Use notas de crédito.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        charge_id = request.data.get('charge_id')
        if not charge_id:
            return Response(
                {'error': 'Se requiere charge_id'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        from .models import OrderCharge
        try:
            charge = OrderCharge.objects.get(id=charge_id, invoice=invoice)
        except OrderCharge.DoesNotExist:
            return Response(
                {'error': 'Cargo no encontrado en esta factura'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Guardar valores anteriores para historial
        previous_values = {
            'quantity': charge.quantity,
            'unit_price': str(charge.unit_price),
            'discount': str(charge.discount),
            'iva_type': charge.iva_type
        }
        
        # Actualizar campos si se proporcionan
        if 'quantity' in request.data:
            charge.quantity = int(request.data['quantity'])
        if 'unit_price' in request.data:
            charge.unit_price = Decimal(str(request.data['unit_price']))
        if 'discount' in request.data:
            discount_val = request.data['discount']
            if discount_val == '' or discount_val is None:
                charge.discount = Decimal('0.00')
            else:
                try:
                    charge.discount = Decimal(str(discount_val))
                except (ValueError, TypeError, Exception):
                    charge.discount = Decimal('0.00')
        if 'description' in request.data:
            charge.description = request.data['description']
        if 'iva_type' in request.data:
            charge.iva_type = request.data['iva_type']
        if 'applies_iva' in request.data:
            # Toggle logic for simple boolean from frontend
            charge.iva_type = 'gravado' if request.data['applies_iva'] else 'exento'
        
        # El save() del modelo recalcula subtotal, iva_amount, total
        charge.save()
        
        # Registrar en historial
        from .models import InvoiceEditHistory
        InvoiceEditHistory.objects.create(
            invoice=invoice,
            edit_type='charge_edited',
            description=f'Cargo editado: {charge.service.name}',
            previous_values=previous_values,
            new_values={
                'quantity': charge.quantity,
                'unit_price': str(charge.unit_price),
                'discount': str(charge.discount),
                'iva_type': charge.iva_type
            },
            user=request.user
        )
        
        # Recalcular totales de la factura
        invoice.calculate_totals()
        
        return Response({
            'message': 'Cargo actualizado correctamente',
            'charge_id': charge.id
        })

    @action(detail=True, methods=['post'])
    def remove_item(self, request, pk=None):
        """
        Quita un item de la factura (servicio o gasto).

        REVERSIBILIDAD:
        - El item vuelve automáticamente a la OS como "Disponible para Facturar"
        - El billing_status se actualiza a 'disponible'
        - Solo funciona si la factura está en estado "Pendiente" (pre-factura)

        RESTRICCIÓN:
        - No se puede remover items de facturas con DTE emitido
        """
        invoice = self.get_object()

        if invoice.is_dte_issued:
            return Response(
                {'error': 'No se puede editar. Esta factura ya tiene DTE emitido. Use notas de crédito.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        item_type = request.data.get('item_type')  # 'charge' o 'expense'
        item_id = request.data.get('item_id')

        if not item_type or not item_id:
            return Response(
                {'error': 'Se requiere item_type (charge/expense) e item_id'},
                status=status.HTTP_400_BAD_REQUEST
            )

        from .models import InvoiceEditHistory

        with transaction.atomic():
            if item_type == 'charge':
                from .models import OrderCharge
                try:
                    charge = OrderCharge.objects.select_for_update().get(id=item_id, invoice=invoice)

                    # Guardar información para historial
                    description = f'{charge.service.name} - {charge.description}'
                    previous_values = {
                        'charge_id': item_id,
                        'linked_to_invoice': True,
                        'invoice_number': invoice.invoice_number,
                        'billing_status': charge.billing_status,
                        'total': str(charge.total)
                    }

                    # Desvincular de la factura (vuelve a estar disponible)
                    charge.invoice = None
                    charge.billing_status = 'disponible'
                    charge.save(skip_order_validation=True)  # Permitir modificación para desvinculación

                    InvoiceEditHistory.objects.create(
                        invoice=invoice,
                        edit_type='charge_removed',
                        description=f'Servicio removido y retornado a OS: {description}',
                        previous_values=previous_values,
                        new_values={
                            'linked_to_invoice': False,
                            'billing_status': 'disponible',
                            'available_in_os': True
                        },
                        user=request.user
                    )

                except OrderCharge.DoesNotExist:
                    return Response({'error': 'Cargo no encontrado'}, status=status.HTTP_404_NOT_FOUND)

            elif item_type == 'expense':
                from apps.transfers.models import Transfer
                try:
                    transfer = Transfer.objects.select_for_update().get(id=item_id, invoice=invoice)

                    description = transfer.description[:100]
                    previous_values = {
                        'transfer_id': item_id,
                        'linked_to_invoice': True,
                        'invoice_number': invoice.invoice_number,
                        'billing_status': transfer.billing_status,
                        'total': str(transfer.get_customer_total())
                    }

                    # Desvincular de la factura (vuelve a estar disponible)
                    transfer.invoice = None
                    transfer.billing_status = 'disponible'
                    transfer.save()

                    InvoiceEditHistory.objects.create(
                        invoice=invoice,
                        edit_type='expense_removed',
                        description=f'Gasto removido y retornado a OS: {description}',
                        previous_values=previous_values,
                        new_values={
                            'linked_to_invoice': False,
                            'billing_status': 'disponible',
                            'available_in_os': True
                        },
                        user=request.user
                    )

                except Transfer.DoesNotExist:
                    return Response({'error': 'Gasto no encontrado'}, status=status.HTTP_404_NOT_FOUND)
            else:
                return Response({'error': 'item_type debe ser "charge" o "expense"'}, status=status.HTTP_400_BAD_REQUEST)

            # Recalcular totales de la factura
            invoice.calculate_totals()

            # Verificar si la factura quedó vacía
            charges_count = invoice.charges.filter(is_deleted=False).count()
            expenses_count = invoice.billed_transfers.filter(is_deleted=False).count()
            
            # Si la factura quedó sin items, eliminarla automáticamente
            if charges_count == 0 and expenses_count == 0:
                invoice_number = invoice.invoice_number
                invoice.delete()
                return Response({
                    'message': f'Pre-factura {invoice_number} eliminada porque no tenía más items.',
                    'invoice_deleted': True,
                    'item_returned_to_os': True
                })

        return Response({
            'message': f'Item removido de la factura. Ahora está disponible para facturar nuevamente.',
            'item_returned_to_os': True,
            'invoice_deleted': False,
            'invoice_items_remaining': {
                'charges': charges_count,
                'expenses': expenses_count
            }
        })

    @action(detail=True, methods=['post'])
    def add_items(self, request, pk=None):
        """
        Agregar items a una factura existente (solo si es pre-factura).
        Útil para vincular items a facturas existentes que no tienen items vinculados.
        """
        invoice = self.get_object()
        
        if invoice.is_dte_issued:
            return Response(
                {'error': 'No se puede modificar. Esta factura ya tiene DTE emitido.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        charge_ids = request.data.get('charge_ids', [])
        transfer_ids = request.data.get('transfer_ids', [])
        
        if not charge_ids and not transfer_ids:
            return Response(
                {'error': 'Debe proporcionar charge_ids o transfer_ids'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        service_order = invoice.service_order
        added_charges = 0
        added_expenses = 0
        
        # Vincular cargos
        if charge_ids:
            from .models import OrderCharge
            result = OrderCharge.objects.filter(
                id__in=charge_ids,
                service_order=service_order,
                invoice__isnull=True
            ).update(invoice=invoice)
            added_charges = result
        
        # Vincular gastos
        if transfer_ids:
            from apps.transfers.models import Transfer
            result = Transfer.objects.filter(
                id__in=transfer_ids,
                service_order=service_order,
                invoice__isnull=True
            ).update(invoice=invoice)
            added_expenses = result
        
        # Recalcular totales
        invoice.calculate_totals()
        
        # Registrar en historial
        from .models import InvoiceEditHistory
        InvoiceEditHistory.objects.create(
            invoice=invoice,
            edit_type='charge_added',
            description=f'Items agregados: {added_charges} servicios, {added_expenses} gastos',
            new_values={'charge_ids': charge_ids, 'transfer_ids': transfer_ids},
            user=request.user
        )
        
        return Response({
            'message': f'Items agregados correctamente. {added_charges} servicios, {added_expenses} gastos.',
            'added_charges': added_charges,
            'added_expenses': added_expenses
        })

    @action(detail=True, methods=['get'])
    def available_items(self, request, pk=None):
        """
        Obtener items disponibles para agregar a esta factura.
        Solo muestra items de la misma OS que aún no están facturados.
        """
        invoice = self.get_object()
        service_order = invoice.service_order
        
        items = []
        
        # Cargos no facturados
        from .models import OrderCharge
        charges = OrderCharge.objects.filter(
            service_order=service_order,
            invoice__isnull=True,
            is_deleted=False
        ).select_related('service')
        
        for charge in charges:
            items.append({
                'id': charge.id,
                'type': 'charge',
                'description': f"{charge.service.name} - {charge.description or ''}",
                'amount': float(charge.subtotal),
                'iva': float(charge.iva_amount),
                'total': float(charge.total)
            })
        
        # Gastos no facturados
        from apps.transfers.models import Transfer
        from decimal import Decimal
        
        transfers = Transfer.objects.filter(
            service_order=service_order,
            transfer_type__in=['cargos', 'costos', 'terceros'],
            invoice__isnull=True,
            is_deleted=False
        ).select_related('provider')
        
        for t in transfers:
            markup = t.customer_markup_percentage or Decimal('0.00')
            cost = t.amount
            base_price = cost * (1 + markup / Decimal('100.00'))
            
            if t.customer_applies_iva:
                iva = base_price * Decimal('0.13')
            else:
                iva = Decimal('0.00')
            
            items.append({
                'id': t.id,
                'type': 'expense',
                'description': f"{t.description} ({t.provider.name if t.provider else 'N/A'})",
                'amount': float(base_price),
                'iva': float(iva),
                'total': float(base_price + iva)
            })
        
        return Response(items)

    @action(detail=True, methods=['get'])
    def edit_history(self, request, pk=None):
        """Obtener el historial de ediciones de una factura"""
        invoice = self.get_object()
        
        from .models import InvoiceEditHistory
        history = InvoiceEditHistory.objects.filter(invoice=invoice).select_related('user')
        
        data = []
        for h in history:
            data.append({
                'id': h.id,
                'edit_type': h.edit_type,
                'edit_type_display': h.get_edit_type_display(),
                'description': h.description,
                'previous_values': h.previous_values,
                'new_values': h.new_values,
                'user': h.user.get_full_name() if h.user else 'Sistema',
                'created_at': h.created_at.isoformat()
            })
        
        return Response(data)

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Get invoicing summary statistics with enhanced KPIs"""
        queryset = self.get_queryset()
        # Exclude cancelled invoices from KPIs to ensure mathematical consistency
        # Total Facturado should only reflect valid, active invoices.
        active_queryset = queryset.exclude(status='cancelled')
        
        from django.utils import timezone
        today = timezone.now().date()

        # Basic totals
        total_invoiced = active_queryset.aggregate(
            total=Sum('total_amount')
        )['total'] or Decimal('0')

        total_pending = active_queryset.filter(balance__gt=0).aggregate(
            total=Sum('balance')
        )['total'] or Decimal('0')

        # Para que Recuperado + Por Cobrar ≈ Total Facturado, debemos sumar
        # Paid Amount + Credited Amount + Retencion.
        # Si solo sumamos Paid Amount, faltan las NC y Retenciones.
        totals = active_queryset.aggregate(
            paid=Sum('paid_amount'),
            credited=Sum('credited_amount'),
            retained=Sum('retencion'),
            services=Sum('total_services'),
            expenses=Sum('total_third_party')
        )
        
        # "Recuperado" = Pagado + Acreditado + Retenido (lo que ya no se debe)
        total_paid_real = (totals['paid'] or Decimal('0'))
        total_credited = totals['credited'] or Decimal('0')
        total_retained = totals['retained'] or Decimal('0')
        
        total_collected = total_paid_real + total_credited + total_retained

        # Overdue invoices
        overdue_queryset = active_queryset.filter(
            due_date__lt=today,
            balance__gt=0
        )
        total_overdue = overdue_queryset.aggregate(
            total=Sum('balance')
        )['total'] or Decimal('0')
        overdue_count = overdue_queryset.count()

        # Status counts
        pending_count = active_queryset.filter(balance__gt=0).count()
        paid_count = active_queryset.filter(status='paid').count()
        partial_count = active_queryset.filter(status='partial').count()
        cancelled_count = queryset.filter(status='cancelled').count() # Keep using original queryset to count cancelled

        # Additional KPIs
        from datetime import timedelta
        week_end = today + timedelta(days=7)
        due_this_week = active_queryset.filter(
            due_date__gte=today,
            due_date__lte=week_end,
            balance__gt=0
        ).count()

        total_services = totals['services'] or Decimal('0')
        total_third_party_expenses = totals['expenses'] or Decimal('0')

        # Contar notas de crédito
        credit_notes_count = CreditNote.objects.filter(
            invoice__in=active_queryset,
            is_deleted=False
        ).count()

        return Response({
            'total_invoiced': str(total_invoiced),
            'total_pending': str(total_pending),
            'total_collected': str(total_collected), # Now includes Credited + Retained
            'total_overdue': str(total_overdue),
            'total_services': str(total_services),
            'total_third_party_expenses': str(total_third_party_expenses),
            'total_credited': str(total_credited),
            'credit_notes_count': credit_notes_count,
            'pending_count': pending_count,
            'paid_count': paid_count,
            'partial_count': partial_count,
            'overdue_count': overdue_count,
            'cancelled_count': cancelled_count,
            'due_this_week': due_this_week,
            'total_invoices': active_queryset.count(),
        })

    @action(detail=False, methods=['get'], permission_classes=[IsOperativo])
    def export_excel(self, request):
        """Export invoices to Excel with professional formatting"""
        from django.utils import timezone
        from datetime import datetime

        queryset = self.get_queryset()

        # Apply filters from request
        client_id = request.query_params.get('client')
        if client_id:
            queryset = queryset.filter(service_order__client_id=client_id)

        status_filter = request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        date_from = request.query_params.get('dateFrom')
        if date_from:
            queryset = queryset.filter(issue_date__gte=date_from)

        date_to = request.query_params.get('dateTo')
        if date_to:
            queryset = queryset.filter(issue_date__lte=date_to)

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cuentas por Cobrar"

        # Estilos profesionales - Diseño GPRO
        header_fill = PatternFill(start_color="0F2E4D", end_color="0F2E4D", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        title_font = Font(size=16, bold=True, color="0F2E4D")
        subtitle_font = Font(size=12, bold=True, color="1A4C7A")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        currency_format = '"$"#,##0.00'

        # === ENCABEZADO ===
        ws['A1'] = "CUENTAS POR COBRAR"
        ws['A1'].font = title_font
        ws.merge_cells('A1:N1')

        ws['A2'] = "GPRO LOGISTIC - Agencia Aduanal"
        ws['A2'].font = Font(size=11, color="666666")
        ws.merge_cells('A2:N2')

        ws['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A3'].font = Font(size=9, italic=True, color="999999")

        # === TABLA DE DATOS ===
        start_row = 5
        ws.cell(row=start_row, column=1, value="DETALLE DE FACTURAS").font = subtitle_font
        ws.merge_cells(f'A{start_row}:N{start_row}')

        # Headers de la tabla (añadidos DUCA, BL, PO, Total Servicios, Total Gastos)
        headers = [
            'No. Factura', 'Cliente', 'Orden de Servicio', 'PO', 'DUCA', 'BL', 'CCF',
            'Fecha Emisión', 'Fecha Vencimiento', 'Total Servicios', 'Total Gastos',
            'Total Factura', 'Pagado', 'Saldo', 'Estado'
        ]
        header_row = start_row + 1

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Data rows
        status_display = {
            'pending': 'Pendiente',
            'partial': 'Pago Parcial',
            'paid': 'Pagada',
            'overdue': 'Vencida',
            'cancelled': 'Anulada',
        }

        today = timezone.now().date()
        data_row = header_row + 1
        totals = {'total_services': 0, 'total_third_party': 0, 'total': 0, 'paid': 0, 'balance': 0}

        for invoice in queryset:
            # Columna 1: No. Factura
            ws.cell(row=data_row, column=1, value=invoice.invoice_number).border = thin_border
            
            # Columna 2: Cliente
            ws.cell(row=data_row, column=2, value=invoice.service_order.client.name if invoice.service_order else '').border = thin_border
            
            # Columna 3: Orden de Servicio
            ws.cell(row=data_row, column=3, value=invoice.service_order.order_number if invoice.service_order else '').border = thin_border
            
            # Columna 4: PO
            po_value = invoice.service_order.purchase_order if invoice.service_order else ''
            ws.cell(row=data_row, column=4, value=po_value).border = thin_border
            
            # Columna 5: DUCA (desde la orden de servicio)
            duca_value = invoice.service_order.duca if invoice.service_order and invoice.service_order.duca else ''
            ws.cell(row=data_row, column=5, value=duca_value).border = thin_border
            
            # Columna 6: BL (desde la orden de servicio)
            bl_value = invoice.service_order.bl_reference if invoice.service_order and invoice.service_order.bl_reference else ''
            ws.cell(row=data_row, column=6, value=bl_value).border = thin_border
            
            # Columna 7: CCF
            ws.cell(row=data_row, column=7, value=invoice.ccf or '').border = thin_border
            
            # Columna 8: Fecha Emisión
            ws.cell(row=data_row, column=8, value=invoice.issue_date.strftime('%d/%m/%Y') if invoice.issue_date else '').border = thin_border
            
            # Columna 9: Fecha Vencimiento
            ws.cell(row=data_row, column=9, value=invoice.due_date.strftime('%d/%m/%Y') if invoice.due_date else '').border = thin_border

            # Columna 10: Total Servicios
            total_services_cell = ws.cell(row=data_row, column=10, value=float(invoice.total_services))
            total_services_cell.number_format = currency_format
            total_services_cell.border = thin_border
            total_services_cell.alignment = Alignment(horizontal='right')

            # Columna 11: Total Gastos a Terceros
            total_third_party_cell = ws.cell(row=data_row, column=11, value=float(invoice.total_third_party))
            total_third_party_cell.number_format = currency_format
            total_third_party_cell.border = thin_border
            total_third_party_cell.alignment = Alignment(horizontal='right')

            # Columna 12: Total Factura
            total_cell = ws.cell(row=data_row, column=12, value=float(invoice.total_amount))
            total_cell.number_format = currency_format
            total_cell.border = thin_border
            total_cell.alignment = Alignment(horizontal='right')

            # Columna 13: Pagado
            paid_cell = ws.cell(row=data_row, column=13, value=float(invoice.paid_amount))
            paid_cell.number_format = currency_format
            paid_cell.border = thin_border
            paid_cell.alignment = Alignment(horizontal='right')

            # Columna 14: Saldo
            balance_cell = ws.cell(row=data_row, column=14, value=float(invoice.balance))
            balance_cell.number_format = currency_format
            balance_cell.border = thin_border
            balance_cell.alignment = Alignment(horizontal='right')

            # Columna 15: Estado
            ws.cell(row=data_row, column=15, value=status_display.get(invoice.status, invoice.status)).border = thin_border

            # Sumar totales
            totals['total_services'] += float(invoice.total_services)
            totals['total_third_party'] += float(invoice.total_third_party)
            totals['total'] += float(invoice.total_amount)
            totals['paid'] += float(invoice.paid_amount)
            totals['balance'] += float(invoice.balance)

            data_row += 1

        # Fila de totales
        ws.cell(row=data_row, column=1, value="TOTALES").font = Font(bold=True)
        ws.cell(row=data_row, column=1).border = thin_border
        for col in range(2, 10):
            ws.cell(row=data_row, column=col).border = thin_border

        # Total Servicios
        total_services_sum = ws.cell(row=data_row, column=10, value=totals['total_services'])
        total_services_sum.number_format = currency_format
        total_services_sum.font = Font(bold=True)
        total_services_sum.border = thin_border
        total_services_sum.alignment = Alignment(horizontal='right')

        # Total Gastos a Terceros
        total_third_party_sum = ws.cell(row=data_row, column=11, value=totals['total_third_party'])
        total_third_party_sum.number_format = currency_format
        total_third_party_sum.font = Font(bold=True)
        total_third_party_sum.border = thin_border
        total_third_party_sum.alignment = Alignment(horizontal='right')

        # Total Factura
        total_total = ws.cell(row=data_row, column=12, value=totals['total'])
        total_total.number_format = currency_format
        total_total.font = Font(bold=True)
        total_total.border = thin_border
        total_total.alignment = Alignment(horizontal='right')

        # Total Pagado
        total_paid = ws.cell(row=data_row, column=13, value=totals['paid'])
        total_paid.number_format = currency_format
        total_paid.font = Font(bold=True)
        total_paid.border = thin_border
        total_paid.alignment = Alignment(horizontal='right')

        # Total Saldo
        total_balance = ws.cell(row=data_row, column=14, value=totals['balance'])
        total_balance.number_format = currency_format
        total_balance.font = Font(bold=True)
        total_balance.border = thin_border
        total_balance.alignment = Alignment(horizontal='right')

        ws.cell(row=data_row, column=15).border = thin_border

        # Ajustar anchos de columna
        column_widths = [18, 30, 18, 15, 16, 16, 15, 14, 14, 16, 16, 16, 14, 14, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=GPRO_CXC_Facturas.xlsx'
        wb.save(response)
        return response

    @action(detail=True, methods=['post'])
    def add_payment(self, request, pk=None):
        """
        Add a payment (abono) to an invoice with distributed lock.
        
        Soporta dos modos:
        1. Pago simple (modo legacy): Solo se especifica el monto total
        2. Pago por items: Se especifica el desglose por cada item (item_allocations)
        
        Para pago por items, enviar:
        {
            "amount": 1000.00,  # Monto total del pago
            "payment_date": "2025-01-04",
            "payment_method": "transferencia",
            "bank": 1,  # opcional
            "reference": "REF-001",  # opcional
            "notes": "Pago parcial",  # opcional
            "item_allocations": [
                {"item_type": "service", "item_id": 1, "amount": 500.00},
                {"item_type": "expense", "item_id": 2, "amount": 500.00}
            ]
        }
        """
        from .models import PaymentItemAllocation, OrderCharge
        from apps.transfers.models import Transfer
        
        invoice = self.get_object()

        if invoice.balance <= 0:
            return Response(
                {'error': 'Esta factura ya está completamente pagada'},
                status=status.HTTP_400_BAD_REQUEST
            )

        def process_payment():
            # Re-fetch invoice inside lock to ensure fresh data
            inv = Invoice.objects.select_for_update().get(pk=pk)
            
            # SEGURIDAD: No permitir pagos a facturas anuladas
            if inv.status == 'cancelled':
                raise ValueError('No se pueden registrar pagos a una factura ANULADA')
            
            # Re-validate balance inside lock
            if inv.balance <= 0:
                raise ValueError('Esta factura ya está completamente pagada')
            
            amount = Decimal(str(request.data.get('amount', 0)))
            if amount <= 0:
                raise ValueError('El monto debe ser mayor a cero')

            # FIX DECIMALES: Redondear balance para comparación
            # Si el balance es 8.3754... y el usuario paga 8.38, debe permitirse.
            from decimal import ROUND_HALF_UP
            rounded_balance = inv.balance.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

            # Usar tolerancia pequeña para evitar bloqueos por milésimas
            if amount > rounded_balance:
                raise ValueError(f'El monto excede el saldo pendiente (${rounded_balance})')

            # Get bank if provided
            bank_id = request.data.get('bank')
            bank = None
            if bank_id:
                try:
                    bank = Bank.objects.get(pk=bank_id)
                except Bank.DoesNotExist:
                    pass

            # Obtener asignaciones por item si se proporcionaron
            item_allocations_data = request.data.get('item_allocations', [])
            
            # Si viene como string JSON (de FormData), parsearlo
            import json
            if isinstance(item_allocations_data, str):
                try:
                    item_allocations_data = json.loads(item_allocations_data)
                except json.JSONDecodeError:
                    item_allocations_data = []
            
            # Validar que la suma de asignaciones sea igual al monto total (si hay asignaciones)
            if item_allocations_data:
                total_allocated = sum(
                    Decimal(str(alloc.get('amount', 0))) 
                    for alloc in item_allocations_data
                )
                if total_allocated != amount:
                    raise ValueError(
                        f'La suma de asignaciones (${total_allocated}) no coincide '
                        f'con el monto del pago (${amount})'
                    )
                
                # Validar que cada item exista y pertenezca a la factura
                for alloc in item_allocations_data:
                    item_type = alloc.get('item_type')
                    item_id = alloc.get('item_id')
                    alloc_amount = Decimal(str(alloc.get('amount', 0)))
                    
                    if alloc_amount <= 0:
                        raise ValueError(f'El monto de asignación debe ser mayor a cero')
                    
                    if item_type == 'service':
                        try:
                            charge = OrderCharge.objects.get(id=item_id, invoice=inv)
                            # Calcular pendiente del item
                            item_paid = PaymentItemAllocation.objects.filter(
                                charge=charge,
                                payment__is_deleted=False,
                                is_deleted=False
                            ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
                            item_pending = charge.total - item_paid
                            if alloc_amount > item_pending:
                                raise ValueError(
                                    f'El monto asignado al servicio "{charge.service.name}" '
                                    f'(${alloc_amount}) excede su pendiente (${item_pending})'
                                )
                        except OrderCharge.DoesNotExist:
                            raise ValueError(f'Servicio #{item_id} no encontrado en esta factura')
                    
                    elif item_type == 'expense':
                        try:
                            expense = Transfer.objects.get(id=item_id, invoice=inv)
                            # Calcular pendiente del item
                            item_paid = PaymentItemAllocation.objects.filter(
                                expense=expense,
                                payment__is_deleted=False,
                                is_deleted=False
                            ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
                            item_total = expense.get_customer_total()
                            item_pending = item_total - item_paid
                            if alloc_amount > item_pending:
                                raise ValueError(
                                    f'El monto asignado al gasto "{expense.description[:30]}" '
                                    f'(${alloc_amount}) excede su pendiente (${item_pending})'
                                )
                        except Transfer.DoesNotExist:
                            raise ValueError(f'Gasto #{item_id} no encontrado en esta factura')
                    else:
                        raise ValueError(f'Tipo de item inválido: {item_type}')

            # VALIDACIÓN CRÍTICA: El pago no puede exceder el balance pendiente
            # Se usa el balance redondeado calculado arriba
            if amount > rounded_balance:
                raise ValueError(
                    f'El monto del pago (${amount}) excede el saldo pendiente de la factura (${rounded_balance}). '
                    f'No se puede registrar un pago mayor al saldo pendiente.'
                )
            
            # Validaciones especiales para pago por retención
            payment_method = request.data.get('payment_method', 'transferencia')
            numero_comprobante_retencion = request.data.get('numero_comprobante_retencion', '')
            
            if payment_method == 'retencion':
                # Validar que la factura tenga retención
                if inv.retencion <= 0:
                    raise ValueError('Esta factura no tiene retención aplicable. No puede registrar un comprobante F-910.')
                
                # FIX DE PRECISIÓN: Comparar valores redondeados a 2 decimales
                # Esto soluciona problemas donde la DB tiene 10.450000001 y el usuario envía 10.45
                from decimal import ROUND_HALF_UP
                
                stored_retention = inv.retencion.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                input_amount = amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                
                # Validar que el monto sea exactamente la retención
                if input_amount != stored_retention:
                    raise ValueError(
                        f'El monto del comprobante de retención debe ser exactamente ${stored_retention}. '
                        f'Monto ingresado: ${input_amount}'
                    )
                
                # VALIDACIÓN CRÍTICA: No permitir duplicar comprobantes de retención
                existing_retention = inv.payments.filter(
                    payment_method='retencion',
                    is_deleted=False
                ).exists()
                
                if existing_retention:
                    raise ValueError('Ya existe un comprobante de retención F-910 registrado para esta factura. No se pueden registrar múltiples comprobantes de retención.')
                
                # Validar que se proporcione el número de comprobante
                if not numero_comprobante_retencion:
                    raise ValueError('Debe proporcionar el número del comprobante F-910.')

            # Crear el pago
            payment = InvoicePayment.objects.create(
                invoice=inv,
                amount=amount,
                payment_date=request.data.get('payment_date'),
                payment_method=payment_method,
                reference_number=request.data.get('reference', ''),
                numero_comprobante_retencion=numero_comprobante_retencion,
                retention_generation_code=request.data.get('retention_generation_code', ''),
                retention_reception_stamp=request.data.get('retention_reception_stamp', ''),
                bank=bank,
                notes=request.data.get('notes', ''),
                receipt_file=request.FILES.get('receipt_file'),
                created_by=request.user
            )

            # Crear asignaciones por item si se proporcionaron
            allocations_created = []
            if item_allocations_data:
                for alloc in item_allocations_data:
                    item_type = alloc.get('item_type')
                    item_id = alloc.get('item_id')
                    alloc_amount = Decimal(str(alloc.get('amount', 0)))
                    alloc_notes = alloc.get('notes', '')
                    
                    allocation_data = {
                        'payment': payment,
                        'amount': alloc_amount,
                        'notes': alloc_notes
                    }
                    
                    if item_type == 'service':
                        allocation_data['charge_id'] = item_id
                    elif item_type == 'expense':
                        allocation_data['expense_id'] = item_id
                    
                    allocation = PaymentItemAllocation.objects.create(**allocation_data)
                    allocations_created.append({
                        'id': allocation.id,
                        'item_type': item_type,
                        'item_id': item_id,
                        'amount': str(alloc_amount)
                    })

            # El modelo InvoicePayment.save() ya actualiza paid_amount de la factura
            # Refrescar la instancia para obtener el balance actualizado
            inv.refresh_from_db()

            return payment, inv.balance, allocations_created

        try:
            # Apply distributed lock if Redis is available
            if LOCKS_ENABLED and distributed_lock:
                try:
                    with distributed_lock(f'invoice_payment_{pk}', timeout=30):
                        with transaction.atomic():
                            payment, new_balance, allocations = process_payment()
                except LockAcquisitionError:
                    return Response(
                        {'error': 'Otro usuario está procesando esta factura. Intente nuevamente.'},
                        status=status.HTTP_409_CONFLICT
                    )
            else:
                with transaction.atomic():
                    payment, new_balance, allocations = process_payment()

            response_data = {
                'message': 'Pago registrado exitosamente',
                'payment_id': payment.id,
                'new_balance': str(new_balance)
            }
            
            if allocations:
                response_data['item_allocations'] = allocations
                response_data['message'] = f'Pago registrado exitosamente con {len(allocations)} asignaciones por item'
            
            return Response(response_data, status=status.HTTP_201_CREATED)

        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['get'])
    def payment_items(self, request, pk=None):
        """
        Obtener los items de una factura con su estado de pago.
        Útil para el modal de registro de pagos por item.
        
        Retorna servicios y gastos con:
        - Monto total del item
        - Monto ya pagado (por asignaciones específicas)
        - Monto pendiente (ajustado por pagos generales)
        
        IMPORTANTE: Los pagos sin asignación específica (como retenciones) se distribuyen
        proporcionalmente entre items GRAVADOS, ya que la retención solo aplica sobre base gravada.
        """
        from .models import PaymentItemAllocation, OrderCharge
        from apps.transfers.models import Transfer
        
        invoice = self.get_object()
        items = []
        
        # Obtener servicios (cargos)
        charges = invoice.charges.filter(is_deleted=False).select_related('service')
        for charge in charges:
            paid_allocated = PaymentItemAllocation.objects.filter(
                charge=charge,
                payment__is_deleted=False,
                is_deleted=False
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
            
            items.append({
                'id': f'service_{charge.id}',
                'item_type': 'service',
                'item_id': charge.id,
                'name': charge.service.name,
                'description': charge.description or '',
                'total': str(charge.total),
                'subtotal': str(charge.subtotal),  # Base sin IVA
                'paid_allocated': str(paid_allocated),
                'iva_type': charge.iva_type,
            })
        
        # Obtener gastos (expenses/transfers)
        transfers = invoice.billed_transfers.filter(is_deleted=False).select_related('provider')
        for transfer in transfers:
            total = transfer.get_customer_total()
            base_price = transfer.get_customer_base_price()
            paid_allocated = PaymentItemAllocation.objects.filter(
                expense=transfer,
                payment__is_deleted=False,
                is_deleted=False
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
            
            items.append({
                'id': f'expense_{transfer.id}',
                'item_type': 'expense',
                'item_id': transfer.id,
                'name': transfer.description[:50] if transfer.description else 'Gasto',
                'description': transfer.provider.name if transfer.provider else 'Sin proveedor',
                'total': str(total),
                'subtotal': str(base_price),  # Base sin IVA
                'paid_allocated': str(paid_allocated),
                'iva_type': getattr(transfer, 'customer_iva_type', 'gravado' if transfer.customer_applies_iva else 'exento'),
            })
        
        # Calcular pagos asignados vs pagos totales
        total_paid_allocated = sum(Decimal(item['paid_allocated']) for item in items)
        total_paid_invoice = invoice.paid_amount  # Incluye pagos con y sin asignación
        unallocated_payments = total_paid_invoice - total_paid_allocated
        
        # Si hay pagos sin asignar (ej: retenciones), distribuirlos proporcionalmente
        # SOLO entre items gravados (ya que retención solo aplica sobre base gravada)
        if unallocated_payments > 0:
            # Calcular total de items gravados
            gravado_items = [item for item in items if item['iva_type'] == 'gravado']
            total_gravado = sum(Decimal(item['total']) for item in gravado_items)
            
            if total_gravado > 0:
                # Distribuir pagos no asignados proporcionalmente entre items gravados
                for item in items:
                    paid_allocated = Decimal(item['paid_allocated'])
                    
                    if item['iva_type'] == 'gravado':
                        # Calcular proporción de este item sobre el total gravado
                        item_total = Decimal(item['total'])
                        proportion = item_total / total_gravado if total_gravado > 0 else Decimal('0')
                        distributed_payment = unallocated_payments * proportion
                        
                        total_paid = paid_allocated + distributed_payment
                    else:
                        # Items exentos/no sujetos: solo pagos asignados explícitamente
                        total_paid = paid_allocated
                    
                    item['paid'] = str(total_paid)
                    item['pending'] = str(Decimal(item['total']) - total_paid)
            else:
                # No hay items gravados, no distribuir
                for item in items:
                    item['paid'] = item['paid_allocated']
                    item['pending'] = str(Decimal(item['total']) - Decimal(item['paid_allocated']))
        else:
            # No hay pagos sin asignar
            for item in items:
                item['paid'] = item['paid_allocated']
                item['pending'] = str(Decimal(item['total']) - Decimal(item['paid_allocated']))
        
        # Calcular totales
        total_invoice = sum(Decimal(item['total']) for item in items)
        total_paid = sum(Decimal(item['paid']) for item in items)
        total_pending = sum(Decimal(item['pending']) for item in items)
        
        return Response({
            'invoice_id': invoice.id,
            'invoice_number': invoice.invoice_number,
            'balance': str(invoice.balance),
            'items': items,
            'summary': {
                'total_items': len(items),
                'total_invoice': str(total_invoice),
                'total_paid_by_items': str(total_paid),
                'total_pending_by_items': str(total_pending),
                'unallocated_payments': str(unallocated_payments),
                'paid_allocated': str(total_paid_allocated),
            }
        })

    @action(detail=True, methods=['post'])
    def add_credit_note(self, request, pk=None):
        """Register a credit note for an invoice with distributed lock"""
        invoice = self.get_object()
        
        if invoice.status == 'cancelled':
             return Response({'error': 'No se pueden aplicar NC a facturas anuladas'}, status=status.HTTP_400_BAD_REQUEST)
        
        def process_credit_note():
            # Re-fetch invoice inside lock
            inv = Invoice.objects.select_for_update().get(pk=pk)
            
            if inv.status == 'cancelled':
                raise ValueError('No se pueden aplicar NC a facturas anuladas')
             
            amount = Decimal(str(request.data.get('amount', 0)))
            if amount <= 0:
                 raise ValueError('El monto debe ser mayor a cero')
                 
            # Validate against pending balance
            if amount > inv.balance:
                raise ValueError(f'El monto de la NC no puede superar el saldo pendiente de la factura (${inv.balance})')

            credit_note = CreditNote.objects.create(
                invoice=inv,
                note_number=request.data.get('note_number'),
                amount=amount,
                reason=request.data.get('reason', ''),
                issue_date=request.data.get('issue_date'),
                pdf_file=request.FILES.get('pdf_file'),
                created_by=request.user
            )
            
            return credit_note

        try:
            # Apply distributed lock if Redis is available
            if LOCKS_ENABLED and distributed_lock:
                try:
                    with distributed_lock(f'invoice_credit_note_{pk}', timeout=30):
                        with transaction.atomic():
                            credit_note = process_credit_note()
                except LockAcquisitionError:
                    return Response(
                        {'error': 'Otro usuario está procesando esta factura. Intente nuevamente.'},
                        status=status.HTTP_409_CONFLICT
                    )
            else:
                with transaction.atomic():
                    credit_note = process_credit_note()
            
            return Response({
                'message': 'Nota de Crédito registrada exitosamente',
                'credit_note_id': credit_note.id
            }, status=status.HTTP_201_CREATED)
            
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)


class InvoicePaymentViewSet(viewsets.ModelViewSet):
    """ViewSet for managing invoice payments"""
    permission_classes = [IsOperativo]
    serializer_class = InvoicePaymentSerializer
    filterset_fields = ['invoice', 'payment_method']
    search_fields = ['reference', 'invoice__invoice_number']
    ordering = ['-payment_date']

    def get_queryset(self):
        return InvoicePayment.objects.all().select_related(
            'invoice', 'invoice__service_order', 'invoice__service_order__client', 'created_by', 'bank'
        )

    def destroy(self, request, *args, **kwargs):
        """Delete a payment and update invoice balance with atomic transaction"""
        payment = self.get_object()

        # CORREGIDO: Usar 'cancelled' que es el estado valido, no 'cancelada'
        if payment.invoice.status == 'cancelled':
            return Response(
                {'error': 'No se pueden eliminar pagos de una factura cancelada'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # CORREGIDO: Usar transaccion atomica con select_for_update para prevenir race conditions
            with transaction.atomic():
                # Re-fetch invoice with lock
                invoice = Invoice.objects.select_for_update().get(pk=payment.invoice_id)

                # Recalcular paid_amount desde los pagos activos (excluyendo el que se elimina)
                from django.db.models import Sum
                remaining_payments = invoice.payments.filter(is_deleted=False).exclude(pk=payment.pk)
                new_paid_amount = remaining_payments.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

                invoice.paid_amount = new_paid_amount
                
                # CORRECCIÓN: Usar la misma lógica que el modelo (total - pagado - notas_credito)
                # La retención se maneja como un pago aparte, no se resta del balance base.
                invoice.balance = invoice.total_amount - invoice.paid_amount - invoice.credited_amount

                # Actualizar estado
                if invoice.balance <= 0:
                    invoice.status = 'paid'
                elif invoice.paid_amount > 0:
                    invoice.status = 'partial'
                else:
                    invoice.status = 'pending'

                invoice.save()

                # Delete payment (soft delete)
                payment.delete()

            return Response(status=status.HTTP_204_NO_CONTENT)

        except Exception as e:
            return Response(
                {'error': f'Error al eliminar el pago: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'], permission_classes=[IsOperativo])
    def export_excel(self, request):
        """Exportar listado de pagos a Excel con formato profesional"""
        from datetime import datetime
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        queryset = self.filter_queryset(self.get_queryset())

        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pagos Recibidos"

        # Estilos profesionales - Diseño GPRO
        header_fill = PatternFill(start_color="0F2E4D", end_color="0F2E4D", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        title_font = Font(size=16, bold=True, color="0F2E4D")
        subtitle_font = Font(size=12, bold=True, color="1A4C7A")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        currency_format = '"$"#,##0.00'

        # === ENCABEZADO ===
        ws['A1'] = "REPORTE DE PAGOS RECIBIDOS"
        ws['A1'].font = title_font
        ws.merge_cells('A1:H1')

        ws['A2'] = "GPRO LOGISTIC - Agencia Aduanal"
        ws['A2'].font = Font(size=11, color="666666")
        ws.merge_cells('A2:H2')

        ws['A3'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A3'].font = Font(size=9, italic=True, color="999999")

        # === TABLA DE DATOS ===
        start_row = 5
        ws.cell(row=start_row, column=1, value="DETALLE DE PAGOS").font = subtitle_font
        ws.merge_cells(f'A{start_row}:H{start_row}')

        # Headers de la tabla
        headers = ['Fecha Pago', 'No. Factura', 'Cliente', 'Método Pago',
                   'Referencia', 'Banco', 'Monto', 'Registrado Por']
        header_row = start_row + 1

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Datos
        data_row = header_row + 1
        total_amount = 0

        for payment in queryset:
            ws.cell(row=data_row, column=1, value=payment.payment_date.strftime('%d/%m/%Y')).border = thin_border
            ws.cell(row=data_row, column=2, value=payment.invoice.invoice_number).border = thin_border
            ws.cell(row=data_row, column=3, value=payment.invoice.service_order.client.name if payment.invoice.service_order.client else '').border = thin_border
            ws.cell(row=data_row, column=4, value=payment.get_payment_method_display()).border = thin_border
            ws.cell(row=data_row, column=5, value=payment.reference_number or '').border = thin_border
            ws.cell(row=data_row, column=6, value=payment.bank.name if payment.bank else '').border = thin_border

            # Columna Monto
            amount_cell = ws.cell(row=data_row, column=7, value=float(payment.amount))
            amount_cell.number_format = currency_format
            amount_cell.border = thin_border
            amount_cell.alignment = Alignment(horizontal='right')

            ws.cell(row=data_row, column=8, value=payment.created_by.get_full_name() if payment.created_by else '').border = thin_border

            total_amount += float(payment.amount)
            data_row += 1

        # Fila de totales
        ws.cell(row=data_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=data_row, column=1).border = thin_border
        for col in range(2, 7):
            ws.cell(row=data_row, column=col).border = thin_border

        total_cell = ws.cell(row=data_row, column=7, value=total_amount)
        total_cell.number_format = currency_format
        total_cell.font = Font(bold=True)
        total_cell.border = thin_border
        total_cell.alignment = Alignment(horizontal='right')

        ws.cell(row=data_row, column=8).border = thin_border

        # Ajustar anchos de columna
        column_widths = [14, 18, 30, 18, 20, 20, 15, 20]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # Generar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=GPRO_Pagos_Recibidos_{datetime.now().strftime("%Y%m%d")}.xlsx'

        wb.save(response)
        return response


class CreditNoteViewSet(viewsets.ModelViewSet):
    """ViewSet for managing credit notes"""
    queryset = CreditNote.objects.all()
    serializer_class = CreditNoteSerializer
    permission_classes = [IsOperativo]
    filterset_fields = ['invoice']
    
    def perform_destroy(self, instance):
        instance.delete() # Trigger soft delete and recalculation


from rest_framework.views import APIView
from datetime import datetime
from dateutil.relativedelta import relativedelta


class RetentionControlView(APIView):
    """
    Vista para el Control de Retenciones F-910.
    Proporciona KPIs y lista de facturas con retención para grandes contribuyentes.
    """
    permission_classes = [IsOperativo2]  # Solo Operativo2 y Admins
    
    def get(self, request):
        # Obtener parámetros de filtro
        try:
            year = int(request.query_params.get('year', datetime.now().year))
            month = int(request.query_params.get('month', datetime.now().month))
            client_id = request.query_params.get('client_id')
            status_filter = request.query_params.get('status')  # 'pending', 'received', 'all'
        except (ValueError, TypeError):
            year = datetime.now().year
            month = datetime.now().month
            client_id = None
            status_filter = 'all'
        
        # Filtrar facturas con retención
        base_qs = Invoice.objects.filter(
            retencion__gt=0
        ).exclude(
            status='cancelled'
        ).select_related(
            'service_order__client',
            'created_by'
        ).prefetch_related(
            'payments'
        )
        
        # Filtro por período
        if month == 0:  # Vista anual
            base_qs = base_qs.filter(issue_date__year=year)
        else:
            base_qs = base_qs.filter(
                issue_date__year=year,
                issue_date__month=month
            )
        
        # Filtro por cliente
        if client_id:
            base_qs = base_qs.filter(service_order__client_id=client_id)
            
        # === EXPORTACIÓN A EXCEL ===
        if request.query_params.get('export') == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Control de Retenciones"

            # Estilos profesionales - Diseño GPRO
            header_fill = PatternFill(start_color="0F2E4D", end_color="0F2E4D", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=10)
            title_font = Font(size=16, bold=True, color="0F2E4D")
            subtitle_font = Font(size=12, bold=True, color="1A4C7A")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            currency_format = '"$"#,##0.00'

            # Header Info
            ws['A1'] = "REPORTE DE CONTROL DE RETENCIONES (F-910)"
            ws['A1'].font = title_font
            ws.merge_cells('A1:L1')

            ws['A2'] = "GPRO LOGISTIC - Agencia Aduanal"
            ws['A2'].font = Font(size=11, color="666666")
            ws.merge_cells('A2:P2')

            # Meses en español
            months_es = {
                1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
                5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
                9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
            }
            period_text = f"Año {year}" if month == 0 else f"{months_es.get(month, '')} {year}"
            ws['A3'] = f"Período: {period_text} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws['A3'].font = Font(size=9, italic=True, color="999999")
            ws.merge_cells('A3:P3')

            # Table Headers
            headers = [
                'Factura', 'OS', 'PO', 'DUCA', 'BL', 'Fecha Emisión', 'Cliente', 'NIT Cliente', 
                'Total Factura', 'Monto Retención (1%)', 'Estado',
                'No. Comprobante', 'Fecha Recepción', 'Código Generación', 
                'Sello Recepción', 'Registrado Por'
            ]
            
            start_row = 5
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border

            # Data Rows
            data_row = start_row + 1
            total_retention = 0
            
            status_map = {'received': 'Recibido', 'pending': 'Pendiente'}

            for invoice in base_qs:
                # Get retention payment info
                retention_payment = invoice.payments.filter(
                    payment_method='retencion',
                    is_deleted=False
                ).first()
                
                has_comprobante = retention_payment is not None
                status = 'received' if has_comprobante else 'pending'
                
                # Apply status filter for export too
                if status_filter and status_filter != 'all' and status != status_filter:
                    continue

                # Columns
                ws.cell(row=data_row, column=1, value=invoice.invoice_number).border = thin_border
                
                # Campos de OS
                os = invoice.service_order
                ws.cell(row=data_row, column=2, value=os.order_number if os else '').border = thin_border
                ws.cell(row=data_row, column=3, value=os.purchase_order if os else '').border = thin_border
                ws.cell(row=data_row, column=4, value=os.duca if os else '').border = thin_border
                ws.cell(row=data_row, column=5, value=os.bl_reference if os else '').border = thin_border
                
                ws.cell(row=data_row, column=6, value=invoice.issue_date.strftime('%d/%m/%Y')).border = thin_border
                ws.cell(row=data_row, column=7, value=os.client.name if os and os.client else '').border = thin_border
                ws.cell(row=data_row, column=8, value=os.client.nit if os and os.client else '').border = thin_border
                
                # Financials
                total_cell = ws.cell(row=data_row, column=9, value=float(invoice.total_amount))
                total_cell.number_format = currency_format
                total_cell.border = thin_border
                
                ret_cell = ws.cell(row=data_row, column=10, value=float(invoice.retencion))
                ret_cell.number_format = currency_format
                ret_cell.border = thin_border
                
                # Status
                ws.cell(row=data_row, column=11, value=status_map.get(status, status)).border = thin_border
                
                # Certificate Details
                if has_comprobante:
                    ws.cell(row=data_row, column=12, value=retention_payment.numero_comprobante_retencion).border = thin_border
                    ws.cell(row=data_row, column=13, value=retention_payment.payment_date.strftime('%d/%m/%Y')).border = thin_border
                    ws.cell(row=data_row, column=14, value=retention_payment.retention_generation_code or '').border = thin_border
                    ws.cell(row=data_row, column=15, value=retention_payment.retention_reception_stamp or '').border = thin_border
                    ws.cell(row=data_row, column=16, value=retention_payment.created_by.get_full_name() if retention_payment.created_by else '').border = thin_border
                else:
                    for c in range(12, 17):
                        ws.cell(row=data_row, column=c, value='-').border = thin_border
                
                total_retention += float(invoice.retencion)
                data_row += 1

            # Totals
            ws.cell(row=data_row, column=9, value="TOTAL RETENCIONES").font = Font(bold=True)
            ws.cell(row=data_row, column=9).border = thin_border
            
            total_ret_cell = ws.cell(row=data_row, column=10, value=total_retention)
            total_ret_cell.number_format = currency_format
            total_ret_cell.font = Font(bold=True)
            total_ret_cell.border = thin_border

            # Auto-width (ajustado para 16 columnas)
            column_widths = [15, 12, 15, 20, 15, 12, 30, 15, 15, 18, 12, 20, 15, 35, 35, 20]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=control_retenciones_{year}_{month}.xlsx'
            wb.save(response)
            return response
        
        # Calcular KPIs
        total_retenciones = Decimal('0.00')
        comprobantes_recibidos_count = 0
        comprobantes_recibidos_monto = Decimal('0.00')
        pendientes_count = 0
        pendientes_monto = Decimal('0.00')
        
        invoices_data = []
        
        for invoice in base_qs:
            # Verificar si tiene comprobante F-910 registrado
            has_comprobante = invoice.payments.filter(
                payment_method='retencion',
                is_deleted=False
            ).exists()
            
            comprobante_payment = invoice.payments.filter(
                payment_method='retencion',
                is_deleted=False
            ).first()
            
            total_retenciones += invoice.retencion
            
            if has_comprobante:
                comprobantes_recibidos_count += 1
                comprobantes_recibidos_monto += invoice.retencion
                item_status = 'received'
            else:
                pendientes_count += 1
                pendientes_monto += invoice.retencion
                item_status = 'pending'
            
            # Aplicar filtro de estado
            if status_filter and status_filter != 'all' and item_status != status_filter:
                continue
            
            invoice_item = {
                'id': invoice.id,
                'invoice_number': invoice.invoice_number,
                'os_number': invoice.service_order.order_number if invoice.service_order else 'N/A',
                'issue_date': invoice.issue_date.isoformat(),
                'client_id': invoice.service_order.client.id if invoice.service_order.client else None,
                'client_name': invoice.service_order.client.name if invoice.service_order.client else 'N/A',
                'client_nit': invoice.service_order.client.nit if invoice.service_order.client else '',
                'retencion': float(invoice.retencion),
                'total_amount': float(invoice.total_amount),
                'status': item_status,
                'has_comprobante': has_comprobante,
                'comprobante_data': None
            }
            
            if comprobante_payment:
                invoice_item['comprobante_data'] = {
                    'id': comprobante_payment.id,
                    'numero_retencion': comprobante_payment.numero_comprobante_retencion,
                    'generation_code': comprobante_payment.retention_generation_code,
                    'reception_stamp': comprobante_payment.retention_reception_stamp,
                    'payment_date': comprobante_payment.payment_date.isoformat(),
                    'receipt_file': comprobante_payment.receipt_file.url if comprobante_payment.receipt_file else None,
                    'notes': comprobante_payment.notes,
                    'created_by': comprobante_payment.created_by.get_full_name() if comprobante_payment.created_by else None
                }
            
            invoices_data.append(invoice_item)
        
        # Calcular tasa de recuperación
        tasa_recuperacion = 0
        if total_retenciones > 0:
            tasa_recuperacion = (float(comprobantes_recibidos_monto) / float(total_retenciones)) * 100
        
        response_data = {
            'kpis': {
                'total_retenciones': float(total_retenciones),
                'comprobantes_recibidos': {
                    'count': comprobantes_recibidos_count,
                    'monto': float(comprobantes_recibidos_monto)
                },
                'pendientes': {
                    'count': pendientes_count,
                    'monto': float(pendientes_monto)
                },
                'tasa_recuperacion': round(tasa_recuperacion, 2)
            },
            'invoices': invoices_data,
            'filters': {
                'year': year,
                'month': month,
                'client_id': client_id,
                'status': status_filter
            }
        }
        
        return Response(response_data)
